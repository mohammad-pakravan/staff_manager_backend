from rest_framework import status, generics, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from rest_framework.exceptions import PermissionDenied
from drf_spectacular.utils import extend_schema, extend_schema_view
from drf_spectacular.types import OpenApiTypes
from  ..food_management.permissions import UserHumanResourcesPermission
from .models import Gathering, User
from .serializers import GatheringSerializer, UserSerializer, UserRegistrationSerializer, LoginSerializer, LoginResponseSerializer
from rest_framework.views import APIView
from datetime import datetime
from rest_framework import filters

import pandas as pd
from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.response import Response
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils.encoding import iri_to_uri
from django.conf import settings
import requests



def send_to_external_api(phone_number,request_type):
    """Helper function to send phone number to external API"""
    try:
        external_url = getattr(settings, 'EXTERNAL_API_URL', f'https://auth.metafa.ir/auth/service/{request_type}')
        
        payload = {
            'phone': str(phone_number)
        }
        
        headers = {
            "accept":"application/json",
            "x-service-secret" : "BackEndSecretKey",
            "Content-Type" : "application/json"
        }
        
        # Add timeout from settings or use default
        timeout = getattr(settings, 'EXTERNAL_API_TIMEOUT', 10)
        
        response = requests.post(
            external_url,
            json=payload,
            headers=headers,
            timeout=timeout
        )


        # Try to parse JSON response
        if response.headers.get('Content-Type', '').startswith('application/json'):
            return response.json()
        else:
            return {
                'status_code': response.status_code,
                'content': response.text[:500]  # Limit content length
            }
            
    except requests.exceptions.Timeout:
        return {'error': 'Request timeout', 'details': 'External API did not respond in time'}
    except requests.exceptions.ConnectionError:
        return {'error': 'Connection error', 'details': 'Could not connect to external API'}
    except requests.exceptions.RequestException as e:
        return {'error': 'Request failed', 'details': str(e)}
    except ValueError as e:  # JSON parsing error
        return {'error': 'Invalid response', 'details': 'Could not parse JSON response'}
    except Exception as e:
        return {'error': 'Unexpected error', 'details': str(e)}


@extend_schema(
    summary='Refresh Token',
    description='ایجاد access token جدید با استفاده از refresh token از HttpOnly cookie. نیازی به ارسال body نیست.',
    tags=['Authentication'],
    request=None,
    responses={
        200: {
            'description': 'Token جدید ایجاد شد',
            'content': {
                'application/json': {
                    'example': {
                        'message': 'Token جدید ایجاد شد'
                    }
                }
            }
        },
        400: {
            'description': 'Refresh token یافت نشد یا نامعتبر است',
            'content': {
                'application/json': {
                    'example': {
                        'error': 'Refresh token یافت نشد'
                    }
                }
            }
        }
    }
)
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def refresh_token_view(request):
    """Refresh token با استفاده از HttpOnly cookie"""
    # دریافت refresh token از cookie
    refresh_token = request.COOKIES.get('refresh_token')
    
    if not refresh_token:
        return Response(
            {'error': 'Refresh token یافت نشد'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        refresh = RefreshToken(refresh_token)
        access_token = str(refresh.access_token)
        
        # ساخت Response
        response = Response({
            'message': 'Token جدید ایجاد شد'
        })
        
        # تنظیم access token جدید در HttpOnly cookie
        from django.conf import settings
        cookie_samesite = 'Lax'  # برای development با proxy
        
        response.set_cookie(
            key='access_token',
            value=access_token,
            httponly=True,
            secure=False,
            samesite=cookie_samesite,
            max_age=60 * 60,  # 1 ساعت
            path='/'
        )
        
        # اگر refresh token rotate شود، آن را هم بروزرسانی می‌کنیم
        if refresh:
            response.set_cookie(
                key='refresh_token',
                value=str(refresh),
                httponly=True,
                secure=False,
                samesite=cookie_samesite,
                max_age=7 * 24 * 60 * 60,  # 7 روز
                path='/'
            )
        
        return response
    except Exception as e:
        return Response(
            {'error': 'Token نامعتبر است'},
            status=status.HTTP_400_BAD_REQUEST
        )


 

@extend_schema(
    summary='Login',
    description='لاگین با HttpOnly cookies برای امنیت بیشتر. توکن‌ها در HttpOnly cookies قرار می‌گیرند.',
    tags=['Authentication'],
    request=LoginSerializer,
    responses={
        200: {
            'description': 'لاگین موفق',
            'content': {
                'application/json': {
                    'example': {
                        'user': {
                            'id': 1,
                            'username': 'admin',
                            'email': 'admin@example.com',
                            'first_name': 'محمد',
                            'last_name': 'پاکروان',
                            'role': 'sys_admin',
                            'role_display': 'System Admin'
                        },
                        'message': 'لاگین با موفقیت انجام شد',
                        'note': 'Tokens are stored in HttpOnly cookies and sent automatically with each request'
                    }
                }
            }
        },
        400: {'description': 'ورودی نامعتبر یا اطلاعات اشتباه'}
    }
)
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login_view(request):
    """
    لاگین با HttpOnly cookies برای امنیت بیشتر
    
    توکن‌ها در HttpOnly cookies قرار می‌گیرند و در response body برگردانده نمی‌شوند.
    برای استفاده در Frontend مرورگر، cookies به صورت خودکار ارسال می‌شوند.
    """
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        # ارسال درخواست به متافا برای ورود
        phone_number = getattr(user, 'phone_number', None)
        
        if phone_number:
            metfa_response = send_to_external_api(phone_number,request_type="login")
        else:
            metfa_response = {'error': 'Phone number not found for user'}
        # ساخت Response - توکن‌ها در cookies هستند، نه در response body
        response = Response({
            'user': LoginResponseSerializer(user).data,
            'message': 'لاگین با موفقیت انجام شد',
            'note': 'Tokens are stored in HttpOnly cookies and sent automatically with each request',
            'metfa_response': metfa_response
        })
        
        # تنظیم access token در HttpOnly cookie
        is_development = settings.DEBUG
        
        # در development، از Lax استفاده می‌کنیم (برای proxy)
        # در production، از Lax یا None با secure=True استفاده می‌کنیم
        cookie_samesite = 'Lax'  # برای development با proxy
        
        response.set_cookie(
            key='access_token',
            value=str(refresh.access_token),
            httponly=True,  # JavaScript نمی‌تواند به آن دسترسی داشته باشد
            secure=False,  # در production باید True باشد (HTTPS)
            samesite=cookie_samesite,
            max_age=60 * 60,  # 1 ساعت (مطابق با ACCESS_TOKEN_LIFETIME)
            path='/'
        )
        
        # تنظیم refresh token در HttpOnly cookie
        response.set_cookie(
            key='refresh_token',
            value=str(refresh),
            httponly=True,
            secure=False,  # در production باید True باشد
            samesite=cookie_samesite,
            max_age=7 * 24 * 60 * 60,  # 7 روز (مطابق با REFRESH_TOKEN_LIFETIME)
            path='/'
        )
        
        return response
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(
    summary='Get Current User',
    description='دریافت اطلاعات کاربر فعلی',
    tags=['Authentication'],
    responses={
        200: UserSerializer,
        401: {'description': 'Unauthorized'}
    }
)
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def me_view(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


@extend_schema(
    summary='Logout',
    description='لاگ اوت با پاک کردن HttpOnly cookies',
    tags=['Authentication'],
    request={
        'application/json': {
            'type': 'object',
            'properties': {
                'refresh': {
                    'type': 'string',
                    'description': 'Refresh token (optional, can be from cookie)',
                    'required': False
                }
            }
        }
    },
    responses={
        200: {
            'description': 'لاگ اوت موفق',
            'content': {
                'application/json': {
                    'example': {
                        'message': 'با موفقیت خارج شدید.'
                    }
                }
            }
        }
    }
)
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def logout_view(request):
    """لاگ اوت با پاک کردن HttpOnly cookies"""
    try:
        # دریافت refresh token از cookie یا request body
        refresh_token = request.COOKIES.get('refresh_token') or request.data.get('refresh')
        
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
        
        # ساخت Response
        response = Response({'message': 'با موفقیت خارج شدید.'})
        
        # پاک کردن cookies
        response.delete_cookie('access_token', path='/')
        response.delete_cookie('refresh_token', path='/')
        
        return response
    except Exception as e:
        # حتی اگر token معتبر نباشد، cookies را پاک می‌کنیم
        response = Response({'message': 'با موفقیت خارج شدید.'})
        response.delete_cookie('access_token', path='/')
        response.delete_cookie('refresh_token', path='/')
        return response


@extend_schema_view(
    get=extend_schema(
        summary='List Users',
        description='لیست کاربران سیستم (فقط برای System Admin)',
        tags=['Users'],
        responses={200: UserSerializer(many=True)}
    ),
    post=extend_schema(
        summary='Create User',
        description='ایجاد کاربر جدید و تخصیص مراکز (فقط برای System Admin)',
        tags=['Users'],
        request=UserRegistrationSerializer,
        responses={
            201: UserSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'}
        }
    )
)
class UserListCreateView(generics.ListCreateAPIView):
    queryset = User.objects.all().prefetch_related('centers')
    permission_classes = [permissions.IsAuthenticated]

    def _ensure_sys_admin(self, user):
        if user.role != User.Role.SYS_ADMIN:
            raise PermissionDenied('Only System Admin can manage users.')

    def get_queryset(self):
        user = self.request.user
        
        if user.role == 'sys_admin':
            return User.objects.all().prefetch_related('centers')
        
        if user.role in ['hr', 'admin_food']:
            user_centers = user.centers.all()
            return User.objects.filter(
                centers__in=user_centers,
                is_active=True
            ).distinct().prefetch_related('centers')
        
        raise PermissionDenied('دسترسی ندارید')

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UserRegistrationSerializer
        return UserSerializer

    def create(self, request, *args, **kwargs):
        self._ensure_sys_admin(request.user)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        output_serializer = UserSerializer(user, context=self.get_serializer_context())
        headers = self.get_success_headers(output_serializer.data)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED, headers=headers)




@extend_schema_view(
    get=extend_schema(
        summary='List Gatherings',
        description='لیست گردهمایی‌ها',
        tags=['Gatherings'],
        responses={200: GatheringSerializer(many=True)}
    ),
    post=extend_schema(
        summary='Create Gathering',
        description='ایجاد گردهمایی جدید',
        tags=['Gatherings'],
        request=GatheringSerializer,
        responses={
            201: GatheringSerializer,
            400: {'description': 'Validation error'}
        }
    )
)
class GatheringListCreateView(generics.ListCreateAPIView):
    """
    View for listing and creating gatherings
    """
  
    serializer_class = GatheringSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [ filters.SearchFilter]
    filterset_fields = ['user', 'center']
    search_fields = ['name', 'last_name', 'personal_code']
    def get_queryset(self):
        """
        Return only the current user's gatherings
        """
        return Gathering.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        """
        Save the gathering with the current user if not provided
        """
        # If user is not provided in request, use current user
        if 'user' not in serializer.validated_data:
            serializer.save(user=self.request.user)
        else:
            serializer.save()





@extend_schema_view(
    get=extend_schema(
        summary='Export All Gatherings to Excel',
        description='خروجی اکسل از تمامی گردهمایی‌ها',
        tags=['Gatherings'],
        responses={
            200: {'description': 'Excel file'},
            404: {'description': 'Not found'}
        }
    )
)
class ExportAllGatheringsView(APIView):
    """Export all gatherings to Excel"""
    permission_classes = [UserHumanResourcesPermission]
    
    def get(self, request):
        gatherings = Gathering.objects.all().select_related('user')
        
        if not gatherings.exists():
            return Response(
                {'error': 'هیچ گردهمایی یافت نشد'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Prepare data for DataFrame
        data = []
        for gathering in gatherings:
            data.append({
                'نام': gathering.name,
                'نام خانوادگی': gathering.last_name,
                'کد ملی': gathering.personal_code,
                'مرکز': gathering.center,
                'تعداد اعضای خانواده': gathering.family_members_count,
                'کاربر ایجاد کننده': gathering.user.username,
                'تاریخ ایجاد': gathering.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'تاریخ آخرین ویرایش': gathering.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = BytesIO()
        
        # Create Excel writer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='گردهمایی‌ها', index=False)
            
            # Access the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['گردهمایی‌ها']
            
            # Style the header row (Persian/RTL support)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
                # Auto-adjust column width
                column_letter = get_column_letter(col_num)
                max_length = 0
                for idx, value in enumerate(df[column_title]):
                    max_length = max(max_length, len(str(value)))
                max_length = max(max_length, len(column_title))
                worksheet.column_dimensions[column_letter].width = min(max_length + 10, 50)
            
            # Style data rows
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Prepare the response
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"all_gatherings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        encoded_filename = iri_to_uri(filename)
        response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
        
        return response
    """
    View for retrieving, updating and deleting a gathering
    """
    queryset = Gathering.objects.all()
    serializer_class = GatheringSerializer
    permission_classes = [UserHumanResourcesPermission]

    @extend_schema(
        summary='Export Gathering to Excel',
        description='خروجی اکسل از اطلاعات گردهمایی',
        tags=['Gatherings'],
        responses={
            200: {'description': 'Excel file'},
            404: {'description': 'Not found'}
        }
    )
    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, pk=None):
        """Export a single gathering to Excel"""
        try:
            gathering = self.get_object()
            
            # Create a DataFrame from the gathering data
            data = {
                'نام': [gathering.name],
                'نام خانوادگی': [gathering.last_name],
                'کد ملی': [gathering.personal_code],
                'مرکز': [gathering.center],
                'تعداد اعضای خانواده': [gathering.family_members_count],
                'کاربر ایجاد کننده': [gathering.user.username],
                'تاریخ ایجاد': [gathering.created_at.strftime('%Y-%m-%d %H:%M:%S')],
                'تاریخ آخرین ویرایش': [gathering.updated_at.strftime('%Y-%m-%d %H:%M:%S')]
            }
            
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            output = BytesIO()
            
            # Create Excel writer
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='گردهمایی', index=False)
                
                # Access the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['گردهمایی']
                
                # Style the header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    
                    # Auto-adjust column width
                    column_letter = get_column_letter(col_num)
                    max_length = max(
                        df[column_title].astype(str).apply(len).max(),
                        len(column_title)
                    )
                    worksheet.column_dimensions[column_letter].width = min(max_length + 10, 50)
                
                # Style data rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Prepare the response
            output.seek(0)
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Create filename with gathering info
            filename = f"gathering_{gathering.personal_code}_{gathering.name}_{gathering.last_name}.xlsx"
            encoded_filename = iri_to_uri(filename)
            response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"'
            
            return response
            
        except Gathering.DoesNotExist:
            return Response(
                {'error': 'گردهمایی یافت نشد'},
                status=status.HTTP_404_NOT_FOUND
            )