from rest_framework import generics, status, permissions
from .permissions import (
    IsFoodAdminOrSystemAdmin,
    IsFoodAdminSystemAdminOrEmployee,
    FoodManagementPermission,
    StatisticsPermission
)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, extend_schema_view
from drf_spectacular.types import OpenApiTypes
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.shortcuts import get_object_or_404
from datetime import datetime, timedelta
import jdatetime
from jalali_date import datetime2jalali, date2jalali
from django.http import HttpResponse
from io import BytesIO
from apps.core.pagination import CustomPageNumberPagination


def parse_date_filter(date_str):
    """تبدیل تاریخ شمسی یا میلادی به فرمت مناسب برای فیلتر"""
    if not date_str:
        return None
    
    try:
        # اگر تاریخ شمسی است (فرمت: 1404/08/02)
        if '/' in date_str and len(date_str.split('/')) == 3:
            parts = date_str.split('/')
            if len(parts[0]) == 4:  # سال 4 رقمی (شمسی)
                jalali_date = jdatetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
                return jalali_date.togregorian()
        
        # اگر تاریخ میلادی است (فرمت: 2025-10-24)
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


# Optional imports for export functionality
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from .models import (
    BaseMeal, DailyMenu, DailyMenuMealOption,
    FoodReservation, FoodReport, GuestReservation
)
# برای سازگاری با کدهای قبلی
Meal = BaseMeal
from apps.centers.models import Center
from .serializers import (
    RestaurantSerializer, MealSerializer,
    DailyMenuSerializer, FoodReservationSerializer,
    FoodReservationCreateSerializer, FoodReportSerializer,
    MealStatisticsSerializer,
    GuestReservationSerializer, GuestReservationCreateSerializer,
    SimpleFoodReservationSerializer, SimpleGuestReservationSerializer,
    MealOptionReportSerializer, BaseMealReportSerializer, UserReportSerializer,
    DateReportSerializer, DetailedReservationReportSerializer, ComprehensiveReportSerializer,
    DailyMenuMealOptionSerializer as MealOptionSerializer,
    SimpleBaseMealSerializer, SimpleRestaurantSerializer,
    DailyMenuMealUpdateSerializer,
    SimpleEmployeeDailyMenuSerializer
)
from .models import Restaurant


# ========== Meal Management ==========

@extend_schema_view(
    get=extend_schema(
        operation_id='meal_list',
        summary='List Meals',
        description='Get list of all base meals (food groups). Meal options (DailyMenuMealOption) are managed per DailyMenu via admin panel.',
        tags=['Meals']
    ),
    post=extend_schema(
        operation_id='meal_create',
        summary='Create Meal',
        description='Create new base meal (food group). After creation, add meal options (DailyMenuMealOption) via admin panel when creating/editing a DailyMenu. (only for food admins and system admins). Returns simplified meal data without restaurant details.',
        tags=['Meals'],
        request=MealSerializer,
        responses={
            201: SimpleBaseMealSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'}
        }
    )
)
class MealListCreateView(generics.ListCreateAPIView):
    """لیست و ایجاد غذاها"""
    queryset = Meal.objects.all()
    serializer_class = MealSerializer
    permission_classes = [FoodManagementPermission]
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        # ادمین سیستم همه غذاها را می‌بیند
        # ادمین غذا فقط غذاهای رستوران‌های مراکز خود را می‌بیند
        # کاربران عادی فقط غذاهای مرکز خود را می‌بینند
        user = self.request.user
        if user.role == 'sys_admin':
            return Meal.objects.all()
        elif user.role == 'admin_food':
            # ادمین غذا: فقط غذاهای رستوران‌هایی که به مراکز ادمین غذا متصل هستند
            if user.centers.exists():
                return Meal.objects.filter(
                    restaurant__centers__in=user.centers.all()
                ).distinct()
            return Meal.objects.none()
        elif user.centers.exists():
            return Meal.objects.filter(is_active=True, restaurant__centers__in=user.centers.all()).distinct()
        return Meal.objects.none()
    
    def create(self, request, *args, **kwargs):
        # فقط ادمین‌های غذا و سیستم می‌توانند غذا ایجاد کنند
        user = request.user
        if user.role not in ['admin_food', 'sys_admin']:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("فقط ادمین‌های غذا می‌توانند غذا ایجاد کنند")
        
        # استفاده از serializer کامل برای ایجاد
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        meal = serializer.save()
        
        # برگرداندن response با serializer ساده
        response_serializer = SimpleBaseMealSerializer(meal, context={'request': request})
        headers = self.get_success_headers(response_serializer.data)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED, headers=headers)


@extend_schema_view(
    get=extend_schema(
        operation_id='meal_detail',
        summary='Get Meal Details',
        description='Get base meal details. Returns simplified meal data without restaurant information.',
        tags=['Meals'],
        responses={
            200: SimpleBaseMealSerializer,
            404: {'description': 'Meal not found'}
        }
    ),
    put=extend_schema(
        operation_id='meal_update',
        summary='Update Meal',
        description='Update base meal. Note: Meal options (DailyMenuMealOption) should be managed via admin panel when creating/editing a DailyMenu. (only for admins). Returns simplified meal data without restaurant information.',
        tags=['Meals'],
        request=MealSerializer,
        responses={
            200: SimpleBaseMealSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'},
            404: {'description': 'Meal not found'}
        }
    ),
    patch=extend_schema(
        operation_id='meal_partial_update',
        summary='Partial Update Meal',
        description='Partially update base meal. Note: Meal options (DailyMenuMealOption) should be managed via admin panel when creating/editing a DailyMenu. (only for admins). Returns simplified meal data without restaurant information.',
        tags=['Meals'],
        request=MealSerializer,
        responses={
            200: SimpleBaseMealSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'},
            404: {'description': 'Meal not found'}
        }
    ),
    delete=extend_schema(
        operation_id='meal_delete',
        summary='Delete Meal',
        description='Delete meal (only for admins)',
        tags=['Meals']
    )
)
class MealDetailView(generics.RetrieveUpdateDestroyAPIView):
    """جزئیات، ویرایش و حذف غذا"""
    queryset = Meal.objects.all()
    serializer_class = MealSerializer
    permission_classes = [FoodManagementPermission]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'sys_admin':
            return Meal.objects.all()
        elif user.role == 'admin_food':
            # ادمین غذا: فقط غذاهای رستوران‌هایی که به مراکز ادمین غذا متصل هستند
            if user.centers.exists():
                return Meal.objects.filter(
                    restaurant__centers__in=user.centers.all()
                ).distinct()
            return Meal.objects.none()
        # کاربران عادی فقط غذاهای مرکز خود را می‌بینند
        elif user.centers.exists():
            return Meal.objects.filter(restaurant__centers__in=user.centers.all()).distinct()
        else:
            return Meal.objects.none()
    
    def retrieve(self, request, *args, **kwargs):
        """بازگرداندن جزئیات غذا با serializer ساده"""
        instance = self.get_object()
        serializer = SimpleBaseMealSerializer(instance, context={'request': request})
        return Response(serializer.data)
    
    def update(self, request, *args, **kwargs):
        """به‌روزرسانی کامل غذا"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        # استفاده از serializer کامل برای validation و update
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        meal = serializer.save()
        
        # برگرداندن response با serializer ساده
        response_serializer = SimpleBaseMealSerializer(meal, context={'request': request})
        return Response(response_serializer.data)
    
    def partial_update(self, request, *args, **kwargs):
        """به‌روزرسانی جزئی غذا"""
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)


# ========== Restaurant Meals ==========

@extend_schema(
    operation_id='restaurant_meals',
    summary='Get Meals by Restaurant',
    description='Get list of meals for a specific restaurant. Food admin can only see meals of restaurants that belong to their assigned centers. Returns only meal data without restaurant information. No pagination.',
    tags=['Meals'],
    parameters=[
        {
            'name': 'restaurant_id',
            'in': 'path',
            'description': 'ID of the restaurant',
            'required': True,
            'schema': {'type': 'integer'}
        }
    ],
    responses={
        200: SimpleBaseMealSerializer(many=True),
        403: {'description': 'Permission denied'},
        404: {'description': 'Restaurant not found'}
    }
)
@api_view(['GET'])
@permission_classes([FoodManagementPermission])
def restaurant_meals(request, restaurant_id):
    """لیست غذاهای یک رستوران خاص - فقط رستوران‌هایی که کاربر به آن‌ها دسترسی دارد"""
    user = request.user
    
    # ساخت queryset بر اساس دسترسی کاربر
    if user.role == 'sys_admin':
        # System Admin به همه رستوران‌ها دسترسی دارد
        restaurants_qs = Restaurant.objects.all()
    elif user.role == 'admin_food':
        # ادمین غذا: فقط رستوران‌هایی که به مراکز ادمین غذا متصل هستند
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_400_BAD_REQUEST)
        restaurants_qs = Restaurant.objects.filter(centers__in=user.centers.all()).distinct()
    else:
        # کاربران عادی: فقط رستوران‌های مراکز خود
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_400_BAD_REQUEST)
        restaurants_qs = Restaurant.objects.filter(centers__in=user.centers.all(), is_active=True).distinct()
    
    # بررسی اینکه رستوران در لیست رستوران‌های قابل دسترسی کاربر است
    try:
        restaurant = restaurants_qs.get(id=restaurant_id)
    except Restaurant.DoesNotExist:
        return Response({
            'error': 'رستوران یافت نشد یا شما به این رستوران دسترسی ندارید'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # دریافت غذاهای رستوران
    meals = Meal.objects.filter(restaurant=restaurant)
    
    # برای کاربران عادی فقط غذاهای فعال
    if user.role not in ['sys_admin', 'admin_food']:
        meals = meals.filter(is_active=True)
    
    # استفاده از serializer ساده (بدون اطلاعات رستوران)
    serializer = SimpleBaseMealSerializer(meals, many=True, context={'request': request})
    return Response(serializer.data)


# ========== Meal Type Management ==========


# ========== Restaurant Management ==========

@extend_schema_view(
    get=extend_schema(
        operation_id='restaurant_list',
        summary='List Restaurants',
        description='Get list of all restaurants. Admins see all restaurants, employees see only their center restaurants.',
        tags=['Food Management'],
        responses={200: RestaurantSerializer(many=True)}
    ),
    post=extend_schema(
        operation_id='restaurant_create',
        summary='Create Restaurant',
        description='Create a new restaurant for a center (Admin only)',
        tags=['Food Management'],
        request=RestaurantSerializer,
        responses={
            201: RestaurantSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'}
        }
    )
)
class RestaurantListCreateView(generics.ListCreateAPIView):
    """لیست و ایجاد رستوران‌ها"""
    serializer_class = RestaurantSerializer
    permission_classes = [FoodManagementPermission]

    def get_queryset(self):
        user = self.request.user
        if user.is_admin:
            return Restaurant.objects.all()
        # Employees see only their centers' restaurants
        if user.centers.exists():
            return Restaurant.objects.filter(centers__in=user.centers.all(), is_active=True).distinct()
        return Restaurant.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        # برای ادمین غذا: می‌تواند رستوران برای مراکز خودش و مراکز دیگر ایجاد کند
        # برای ادمین سیستم: می‌تواند رستوران برای هر مرکزی ایجاد کند
        # هیچ محدودیتی اعمال نمی‌کنیم - ادمین غذا می‌تواند هر مرکزی را انتخاب کند
        serializer.save()


# ========== Meal Option Management ==========
# MealOption حذف شد - از DailyMenuMealOption استفاده کنید
# این view ها حذف شدند چون MealOption دیگر وجود ندارد


@extend_schema_view(
    get=extend_schema(
        operation_id='restaurant_detail',
        summary='Get Restaurant Details',
        description='Get details of a specific restaurant',
        tags=['Food Management']
    ),
    put=extend_schema(
        operation_id='restaurant_update',
        summary='Update Restaurant',
        description='Update restaurant completely (Food Admin & System Admin only)',
        tags=['Food Management'],
        request=RestaurantSerializer,
        responses={
            200: RestaurantSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'},
            404: {'description': 'Restaurant not found'}
        }
    ),
    patch=extend_schema(
        operation_id='restaurant_partial_update',
        summary='Partial Update Restaurant',
        description='Partially update restaurant (Food Admin & System Admin only)',
        tags=['Food Management'],
        request=RestaurantSerializer,
        responses={
            200: RestaurantSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'},
            404: {'description': 'Restaurant not found'}
        }
    ),
    delete=extend_schema(
        operation_id='restaurant_delete',
        summary='Delete Restaurant',
        description='Delete restaurant (Food Admin & System Admin only)',
        tags=['Food Management']
    )
)
class RestaurantDetailView(generics.RetrieveUpdateDestroyAPIView):
    """جزئیات رستوران"""
    serializer_class = RestaurantSerializer
    permission_classes = [FoodManagementPermission]

    def get_queryset(self):
        user = self.request.user
        if user.is_admin:
            return Restaurant.objects.all()
        # Employees see only their centers' restaurants
        if user.centers.exists():
            return Restaurant.objects.filter(centers__in=user.centers.all(), is_active=True).distinct()
        return Restaurant.objects.none()


# ========== Weekly Menu Management ==========

# ========== Admin Food Restaurants ==========

@extend_schema(
    operation_id='admin_food_restaurants',
    summary='Get Restaurants for Food Admin',
    description='Get list of restaurants that belong to the food admin\'s centers. Food admin can see all restaurants of their assigned centers. No center parameter needed. Returns simplified restaurant data.',
    tags=['Food Management'],
    responses={200: SimpleRestaurantSerializer(many=True)}
)
@api_view(['GET'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def admin_food_restaurants(request):
    """لیست رستوران‌های مراکز ادمین غذا - خروجی ساده"""
    user = request.user
    
    # بررسی اینکه کاربر ادمین غذا است
    if user.role not in ['admin_food', 'sys_admin']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # اگر sys_admin است، همه رستوران‌ها را برگردان
    if user.role == 'sys_admin':
        restaurants = Restaurant.objects.all()
    else:
        # برای admin_food، فقط رستوران‌های مراکز خودش
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        restaurants = Restaurant.objects.filter(
            centers__in=user.centers.all()
        ).distinct()
    
    # استفاده از serializer ساده
    serializer = SimpleRestaurantSerializer(restaurants, many=True, context={'request': request})
    return Response(serializer.data)


# ========== Admin Food Meals by Date ==========

@extend_schema(
    operation_id='admin_food_meals_by_date',
    summary='Get/Update Meals by Date for Food Admin',
    description='GET: Get list of meals that exist in daily menus for a specific date. Food admin can only see meals of restaurants that belong to their assigned centers. Returns only meal data without restaurant information. No pagination.\n\nPOST: Add or update a single meal with its options in daily menu for a specific date. Requires restaurant_id, base_meal_id, and meal_options array (title, description, price, quantity).',
    tags=['Food Management'],
    parameters=[
        {
            'name': 'date',
            'in': 'query',
            'description': 'Date (format: YYYY-MM-DD or YYYY/MM/DD)',
            'required': True,
            'schema': {'type': 'string'}
        }
    ],
    request=DailyMenuMealUpdateSerializer,
    responses={
        200: SimpleBaseMealSerializer(many=True),
        201: DailyMenuSerializer,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET', 'POST'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def admin_food_meals_by_date(request):
    """لیست و به‌روزرسانی غذاهای موجود در منو برای یک تاریخ مشخص - برای ادمین غذا"""
    user = request.user
    
    # بررسی اینکه کاربر ادمین غذا است
    if user.role not in ['admin_food', 'sys_admin']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت تاریخ
    date = request.query_params.get('date')
    if not date:
        return Response({
            'error': 'تاریخ الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # تبدیل تاریخ شمسی یا میلادی به فرمت مناسب
    parsed_date = parse_date_filter(date)
    if not parsed_date:
        return Response({
            'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # GET: لیست غذاها
    if request.method == 'GET':
        # دریافت منوهای روزانه برای آن تاریخ
        if user.role == 'sys_admin':
            # System Admin: همه منوها
            daily_menus = DailyMenu.objects.filter(date=parsed_date, is_available=True)
        else:
            # Food Admin: فقط منوهای رستوران‌های مراکز خود
            if not user.centers.exists():
                return Response({
                    'error': 'کاربر مرکز مشخصی ندارد'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            daily_menus = DailyMenu.objects.filter(
                date=parsed_date,
                is_available=True,
                restaurant__centers__in=user.centers.all()
            ).distinct()
        
        # استخراج base_meal ها از meal_options موجود در منوها
        meal_ids = set()
        for daily_menu in daily_menus.select_related('restaurant').prefetch_related('menu_meal_options__base_meal'):
            for meal_option in daily_menu.menu_meal_options.all():
                if meal_option.base_meal:
                    meal_ids.add(meal_option.base_meal.id)
        
        # دریافت غذاها
        if meal_ids:
            meals = Meal.objects.filter(id__in=meal_ids)
        else:
            meals = Meal.objects.none()
        
        # استفاده از serializer ساده (بدون اطلاعات رستوران)
        serializer = SimpleBaseMealSerializer(meals, many=True, context={'request': request})
        return Response(serializer.data)
    
    # POST: افزودن/ویرایش یک غذا در منو
    elif request.method == 'POST':
        # اعتبارسنجی داده‌ها
        serializer = DailyMenuMealUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        restaurant_id = serializer.validated_data['restaurant_id']
        base_meal_id = serializer.validated_data['base_meal_id']
        meal_options_data = serializer.validated_data['meal_options']
        
        # بررسی دسترسی به رستوران
        try:
            restaurant = Restaurant.objects.get(id=restaurant_id)
        except Restaurant.DoesNotExist:
            return Response({
                'error': 'رستوران یافت نشد'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # بررسی دسترسی ادمین غذا به رستوران
        if user.role == 'admin_food':
            if not user.centers.exists():
                return Response({
                    'error': 'کاربر مرکز مشخصی ندارد'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # بررسی اینکه رستوران به مراکز ادمین غذا متصل است
            restaurant_centers = restaurant.centers.all()
            user_centers = user.centers.all()
            if not any(center in restaurant_centers for center in user_centers):
                return Response({
                    'error': 'شما به این رستوران دسترسی ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # بررسی وجود base_meal
        try:
            base_meal = BaseMeal.objects.get(id=base_meal_id)
        except BaseMeal.DoesNotExist:
            return Response({
                'error': f'غذای پایه با شناسه {base_meal_id} یافت نشد'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # پیدا کردن یا ایجاد DailyMenu
        daily_menu, created = DailyMenu.objects.get_or_create(
            restaurant=restaurant,
            date=parsed_date,
            defaults={'is_available': True}
        )
        
        # اضافه کردن base_meal به base_meals ManyToMany (اگر وجود نداشته باشد)
        daily_menu.base_meals.add(base_meal)
        
        # حذف meal_options قدیمی برای این base_meal در این daily_menu
        DailyMenuMealOption.objects.filter(
            daily_menu=daily_menu,
            base_meal=base_meal
        ).delete()
        
        # ایجاد meal_options جدید
        for option_data in meal_options_data:
            DailyMenuMealOption.objects.create(
                daily_menu=daily_menu,
                base_meal=base_meal,
                title=option_data['title'],
                description=option_data.get('description', ''),
                price=option_data['price'],
                quantity=option_data['quantity'],
                is_default=False,
                sort_order=0
            )
        
        # بارگذاری مجدد daily_menu با تمام روابط
        daily_menu.refresh_from_db()
        daily_menu = DailyMenu.objects.prefetch_related(
            'menu_meal_options__base_meal',
            'restaurant__centers'
        ).get(id=daily_menu.id)
        
        # استفاده از DailyMenuSerializer برای برگرداندن داده‌های کامل
        serializer = DailyMenuSerializer(daily_menu, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)


# ========== Remove Meal from Daily Menu ==========

@extend_schema(
    operation_id='admin_food_remove_meal_from_menu',
    summary='Remove Meal from Daily Menu',
    description='Remove a base meal and all its meal options from daily menu for a specific date. Food admin can only remove meals from restaurants that belong to their assigned centers.',
    tags=['Food Management'],
    parameters=[
        {
            'name': 'date',
            'in': 'query',
            'description': 'Date (format: YYYY-MM-DD or YYYY/MM/DD)',
            'required': True,
            'schema': {'type': 'string'}
        },
        {
            'name': 'restaurant_id',
            'in': 'query',
            'description': 'Restaurant ID',
            'required': True,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'base_meal_id',
            'in': 'query',
            'description': 'Base Meal ID to remove',
            'required': True,
            'schema': {'type': 'integer'}
        }
    ],
    responses={
        200: DailyMenuSerializer,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'},
        404: {'description': 'Not found'}
    }
)
@api_view(['DELETE'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def admin_food_remove_meal_from_menu(request):
    """حذف غذا از منوی روزانه - برای ادمین غذا"""
    user = request.user
    
    # بررسی اینکه کاربر ادمین غذا است
    if user.role not in ['admin_food', 'sys_admin']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت پارامترها
    date = request.query_params.get('date')
    restaurant_id = request.query_params.get('restaurant_id')
    base_meal_id = request.query_params.get('base_meal_id')
    
    if not date:
        return Response({
            'error': 'تاریخ الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not restaurant_id:
        return Response({
            'error': 'شناسه رستوران الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not base_meal_id:
        return Response({
            'error': 'شناسه غذای پایه الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # تبدیل تاریخ
    parsed_date = parse_date_filter(date)
    if not parsed_date:
        return Response({
            'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # تبدیل restaurant_id و base_meal_id
    try:
        restaurant_id = int(restaurant_id)
        base_meal_id = int(base_meal_id)
    except (ValueError, TypeError):
        return Response({
            'error': 'شناسه رستوران و غذای پایه باید عدد باشند'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # بررسی وجود رستوران
    try:
        restaurant = Restaurant.objects.get(id=restaurant_id)
    except Restaurant.DoesNotExist:
        return Response({
            'error': 'رستوران یافت نشد'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # بررسی دسترسی ادمین غذا به رستوران
    if user.role == 'admin_food':
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # بررسی اینکه رستوران به مراکز ادمین غذا متصل است
        restaurant_centers = restaurant.centers.all()
        user_centers = user.centers.all()
        if not any(center in restaurant_centers for center in user_centers):
            return Response({
                'error': 'شما به این رستوران دسترسی ندارید'
            }, status=status.HTTP_403_FORBIDDEN)
    
    # بررسی وجود base_meal
    try:
        base_meal = BaseMeal.objects.get(id=base_meal_id)
    except BaseMeal.DoesNotExist:
        return Response({
            'error': 'غذای پایه یافت نشد'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # پیدا کردن DailyMenu
    try:
        daily_menu = DailyMenu.objects.get(
            restaurant=restaurant,
            date=parsed_date
        )
    except DailyMenu.DoesNotExist:
        return Response({
            'error': 'منوی روزانه برای این تاریخ و رستوران یافت نشد'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # حذف تمام meal_options مربوط به این base_meal
    deleted_count = DailyMenuMealOption.objects.filter(
        daily_menu=daily_menu,
        base_meal=base_meal
    ).delete()[0]
    
    # حذف base_meal از base_meals ManyToMany
    daily_menu.base_meals.remove(base_meal)
    
    # بارگذاری مجدد daily_menu با تمام روابط
    daily_menu.refresh_from_db()
    daily_menu = DailyMenu.objects.prefetch_related(
        'menu_meal_options__base_meal',
        'restaurant__centers'
    ).get(id=daily_menu.id)
    
    # استفاده از DailyMenuSerializer برای برگرداندن داده‌های کامل
    serializer = DailyMenuSerializer(daily_menu, context={'request': request})
    return Response({
        'message': f'غذا و {deleted_count} اپشن آن با موفقیت از منو حذف شد',
        'deleted_meal_options_count': deleted_count,
        'daily_menu': serializer.data
    }, status=status.HTTP_200_OK)


# ========== Daily Menu Views ==========

class DailyMenuListView(generics.ListAPIView):
    """لیست منوهای روزانه"""
    serializer_class = DailyMenuSerializer
    permission_classes = [FoodManagementPermission]

    def get_queryset(self):
        user = self.request.user
        center_id = self.request.query_params.get('center')
        date = self.request.query_params.get('date')
        
        queryset = DailyMenu.objects.filter(is_available=True)
        
        # فیلتر بر اساس مرکز
        if center_id:
            try:
                center_id = int(center_id)
                queryset = queryset.filter(restaurant__centers__id=center_id).distinct()
            except (ValueError, TypeError):
                # Invalid center_id, return empty queryset
                queryset = queryset.none()
        elif user.centers.exists() and not user.is_admin:
            queryset = queryset.filter(restaurant__centers__in=user.centers.all()).distinct()
        
        # فیلتر بر اساس تاریخ
        if date:
            try:
                from datetime import datetime
                # Validate date format
                datetime.strptime(date, '%Y-%m-%d')
                queryset = queryset.filter(date=date)
            except (ValueError, TypeError):
                # Invalid date format, return empty queryset
                queryset = queryset.none()
        else:
            # اگر تاریخ مشخص نشده، منوهای هفته جاری را نشان بده
            today = timezone.now().date()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            queryset = queryset.filter(date__range=[week_start, week_end])
        
        # بهینه‌سازی با prefetch_related برای جلوگیری از تکرار query ها
        queryset = queryset.select_related('restaurant').prefetch_related(
            'restaurant__centers',
            'menu_meal_options',
            'menu_meal_options__base_meal',
            'menu_dessert_options',
            'menu_dessert_options__base_dessert'
        )
        
        return queryset.order_by('date', 'restaurant__name')


# ========== Food Reservation Views ==========

@extend_schema_view(
    get=extend_schema(
        operation_id='food_reservation_list',
        summary='List Food Reservations',
        description='Get list of food reservations (admins: all, users: own reservations)',
        tags=['Reservations']
    ),
    post=extend_schema(
        operation_id='food_reservation_create',
        summary='Create Food Reservation',
        description='Create new food reservation for user',
        tags=['Reservations'],
        request=FoodReservationCreateSerializer,
        responses={
            201: SimpleFoodReservationSerializer,
            400: {'description': 'Validation error'},
            403: {'description': 'Permission denied'}
        }
    )
)
class FoodReservationListCreateView(generics.ListCreateAPIView):
    """لیست و ایجاد رزروهای غذا"""
    permission_classes = [FoodManagementPermission]
    pagination_class = CustomPageNumberPagination

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return FoodReservationCreateSerializer
        return FoodReservationSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role in ['admin_food', 'sys_admin']:
            return FoodReservation.objects.all()
        return FoodReservation.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class FoodReservationDetailView(generics.RetrieveUpdateDestroyAPIView):
    """جزئیات رزرو غذا"""
    serializer_class = SimpleFoodReservationSerializer
    permission_classes = [FoodManagementPermission]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['admin_food', 'sys_admin']:
            return FoodReservation.objects.all()
        return FoodReservation.objects.filter(user=user)


@extend_schema(
    operation_id='user_reservation_limits',
    summary='Check User Reservation Limits',
    description='Check available and remaining reservations for user on specific date',
    responses={
        200: OpenApiTypes.OBJECT,
        400: OpenApiTypes.OBJECT
    }
)
@api_view(['GET'])
@permission_classes([FoodManagementPermission])
def user_reservation_limits(request):
    """بررسی محدودیت‌های رزرو کاربر - محدودیت رزرو برداشته شده است"""
    user = request.user
    date = request.query_params.get('date')
    
    if not date:
        return Response({
            'error': 'تاریخ الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # تبدیل تاریخ شمسی یا میلادی به فرمت مناسب
    parsed_date = parse_date_filter(date)
    if not parsed_date:
        return Response({
            'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    current_reservations = FoodReservation.get_user_date_reservations_count(user, parsed_date)
    
    return Response({
        'user': {
            'id': user.id,
            'username': user.username,
            'full_name': f"{user.first_name} {user.last_name}".strip(),
        },
        'date': date,
        'current_reservations': current_reservations,
        'unlimited': True,  # محدودیت برداشته شده است
        'can_reserve': True  # همیشه می‌تواند رزرو کند
    })


@api_view(['POST'])
@permission_classes([FoodManagementPermission])
def cancel_reservation(request, reservation_id):
    """لغو رزرو غذا"""
    try:
        reservation = FoodReservation.objects.get(
            id=reservation_id,
            user=request.user
        )
        
        if reservation.cancel():
            return Response({
                'message': 'رزرو با موفقیت لغو شد.'
            })
        else:
            return Response({
                'error': 'امکان لغو این رزرو وجود ندارد.'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    except FoodReservation.DoesNotExist:
        return Response({
            'error': 'رزرو مورد نظر یافت نشد.'
        }, status=status.HTTP_404_NOT_FOUND)


# ========== Guest Reservation Views ==========

@extend_schema_view(
    get=extend_schema(
        operation_id='guest_reservation_list',
        summary='List Guest Reservations',
        description='Get list of guest reservations (admins: all, users: own guest reservations)',
        tags=['Guest Reservations']
    ),
    post=extend_schema(
        operation_id='guest_reservation_create',
        summary='Create Guest Reservation',
        description='Create food reservation for guest',
        tags=['Guest Reservations']
    )
)
class GuestReservationListCreateView(generics.ListCreateAPIView):
    """لیست و ایجاد رزروهای مهمان"""
    permission_classes = [FoodManagementPermission]
    pagination_class = CustomPageNumberPagination

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return GuestReservationCreateSerializer
        return GuestReservationSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role in ['admin_food', 'sys_admin']:
            return GuestReservation.objects.all()
        return GuestReservation.objects.filter(host_user=user)

    def perform_create(self, serializer):
        serializer.save(host_user=self.request.user)


class GuestReservationDetailView(generics.RetrieveUpdateDestroyAPIView):
    """جزئیات رزرو مهمان"""
    serializer_class = SimpleGuestReservationSerializer
    permission_classes = [FoodManagementPermission]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['admin_food', 'sys_admin']:
            return GuestReservation.objects.all()
        return GuestReservation.objects.filter(host_user=user)


@extend_schema(
    operation_id='user_guest_reservation_limits',
    summary='Check User Guest Reservation Limits',
    description='Check available and remaining guest reservations for user on specific date',
    responses={
        200: OpenApiTypes.OBJECT,
        400: OpenApiTypes.OBJECT
    }
)
@api_view(['GET'])
@permission_classes([FoodManagementPermission])
def user_guest_reservation_limits(request):
    """بررسی محدودیت‌های رزرو مهمان کاربر - محدودیت رزرو برداشته شده است"""
    user = request.user
    date = request.query_params.get('date')
    
    if not date:
        return Response({
            'error': 'تاریخ الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # تبدیل تاریخ شمسی یا میلادی به فرمت مناسب
    parsed_date = parse_date_filter(date)
    if not parsed_date:
        return Response({
            'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    current_guest_reservations = GuestReservation.get_user_date_guest_reservations_count(user, parsed_date)
    
    return Response({
        'user': {
            'id': user.id,
            'username': user.username,
            'full_name': f"{user.first_name} {user.last_name}".strip(),
        },
        'date': date,
        'current_guest_reservations': current_guest_reservations,
        'unlimited': True,  # محدودیت برداشته شده است
        'can_reserve_guest': True  # همیشه می‌تواند رزرو کند
    })


@extend_schema(
    operation_id='cancel_guest_reservation',
    summary='Cancel Guest Reservation',
    description='Cancel guest reservation by host user',
    responses={
        200: OpenApiTypes.OBJECT,
        400: OpenApiTypes.OBJECT,
        404: OpenApiTypes.OBJECT
    }
)
@api_view(['POST'])
@permission_classes([FoodManagementPermission])
def cancel_guest_reservation(request, reservation_id):
    """لغو رزرو مهمان"""
    try:
        reservation = GuestReservation.objects.get(
            id=reservation_id,
            host_user=request.user
        )
        
        if reservation.cancel():
            return Response({
                'message': 'رزرو مهمان با موفقیت لغو شد.'
            })
        else:
            return Response({
                'error': 'امکان لغو این رزرو مهمان وجود ندارد.'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    except GuestReservation.DoesNotExist:
        return Response({
            'error': 'رزرو مهمان مورد نظر یافت نشد.'
        }, status=status.HTTP_404_NOT_FOUND)


# ========== Statistics and Reports ==========

@extend_schema(
    operation_id='comprehensive_statistics',
    summary='Comprehensive Statistics',
    description='Get comprehensive statistics including: base meals, meal options (DailyMenuMealOption), restaurants, users, daily menus, reservations, and guest reservations. Supports filters by center_id, user_id, start_date, and end_date. Employees see statistics for all their assigned centers.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'user_id',
            'in': 'query',
            'description': 'فیلتر بر اساس کاربر',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def comprehensive_statistics(request):
    """آمار جامع با امکان فیلتر"""
    user = request.user
    
    # اگر کاربر ادمین نیست، فقط آمار مراکز خودش را ببیند
    if not user.is_admin:
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    user_id = request.query_params.get('user_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # ساخت queryset های پایه
    base_meals_qs = BaseMeal.objects.all()
    meal_options_qs = DailyMenuMealOption.objects.all()
    restaurants_qs = Restaurant.objects.all()
    from apps.accounts.models import User
    users_qs = User.objects.all()
    reservations_qs = FoodReservation.objects.all()
    guest_reservations_qs = GuestReservation.objects.all()
    daily_menus_qs = DailyMenu.objects.all()
    centers_qs = Center.objects.all()
    
    # فیلتر بر اساس مرکز
    if center_id:
        base_meals_qs = base_meals_qs.filter(restaurant__centers__id=center_id).distinct()
        meal_options_qs = meal_options_qs.filter(base_meal__restaurant__centers__id=center_id).distinct()
        restaurants_qs = restaurants_qs.filter(centers__id=center_id).distinct()
        users_qs = users_qs.filter(centers__id=center_id).distinct()
        reservations_qs = reservations_qs.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations_qs = guest_reservations_qs.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        daily_menus_qs = daily_menus_qs.filter(restaurant__centers__id=center_id).distinct()
        centers_qs = centers_qs.filter(id=center_id)
    elif not user.is_admin:
        # اگر ادمین نیست، فقط مراکز خودش
        user_centers = user.centers.all()
        if user_centers.exists():
            base_meals_qs = base_meals_qs.filter(restaurant__centers__in=user_centers).distinct()
            meal_options_qs = meal_options_qs.filter(base_meal__restaurant__centers__in=user_centers).distinct()
            restaurants_qs = restaurants_qs.filter(centers__in=user_centers).distinct()
            users_qs = users_qs.filter(centers__in=user_centers).distinct()
            reservations_qs = reservations_qs.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
            guest_reservations_qs = guest_reservations_qs.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
            daily_menus_qs = daily_menus_qs.filter(restaurant__centers__in=user_centers).distinct()
            centers_qs = centers_qs.filter(id__in=user_centers.values_list('id', flat=True))
        else:
            centers_qs = Center.objects.none()
    
    # فیلتر بر اساس کاربر
    if user_id:
        reservations_qs = reservations_qs.filter(user_id=user_id)
        guest_reservations_qs = guest_reservations_qs.filter(host_user_id=user_id)
    
    # فیلتر بر اساس تاریخ
    if start_date:
        reservations_qs = reservations_qs.filter(daily_menu__date__gte=start_date)
        guest_reservations_qs = guest_reservations_qs.filter(daily_menu__date__gte=start_date)
        daily_menus_qs = daily_menus_qs.filter(date__gte=start_date)
    
    if end_date:
        reservations_qs = reservations_qs.filter(daily_menu__date__lte=end_date)
        guest_reservations_qs = guest_reservations_qs.filter(daily_menu__date__lte=end_date)
        daily_menus_qs = daily_menus_qs.filter(date__lte=end_date)
    
    # محاسبه آمار
    from decimal import Decimal
    today = timezone.now().date()
    
    # آمار غذاها
    total_base_meals = base_meals_qs.count()
    active_base_meals = base_meals_qs.filter(is_active=True).count()
    base_meal_ids = {
        'all': list(base_meals_qs.values_list('id', flat=True)),
        'active': list(base_meals_qs.filter(is_active=True).values_list('id', flat=True)),
        'inactive': list(base_meals_qs.filter(is_active=False).values_list('id', flat=True))
    }
    
    # آمار اپشن‌ها
    total_meal_options = meal_options_qs.count()
    meal_option_ids = {
        'all': list(meal_options_qs.values_list('id', flat=True))
    }
    
    # آمار رستوران‌ها
    total_restaurants = restaurants_qs.count()
    active_restaurants = restaurants_qs.filter(is_active=True).count()
    restaurant_ids = {
        'all': list(restaurants_qs.values_list('id', flat=True)),
        'active': list(restaurants_qs.filter(is_active=True).values_list('id', flat=True)),
        'inactive': list(restaurants_qs.filter(is_active=False).values_list('id', flat=True))
    }
    
    # آمار کاربران
    total_users = users_qs.count()
    active_users = users_qs.filter(is_active=True).count()
    user_ids = {
        'all': list(users_qs.values_list('id', flat=True)),
        'active': list(users_qs.filter(is_active=True).values_list('id', flat=True)),
        'inactive': list(users_qs.filter(is_active=False).values_list('id', flat=True))
    }
    
    # آمار منوهای روزانه
    total_daily_menus = daily_menus_qs.count()
    active_daily_menus = daily_menus_qs.filter(is_available=True).count()
    daily_menu_ids = {
        'all': list(daily_menus_qs.values_list('id', flat=True)),
        'active': list(daily_menus_qs.filter(is_available=True).values_list('id', flat=True)),
        'inactive': list(daily_menus_qs.filter(is_available=False).values_list('id', flat=True))
    }
    
    # آمار رزروها
    total_reservations = reservations_qs.count()
    reserved_reservations = reservations_qs.filter(status='reserved').count()
    cancelled_reservations = reservations_qs.filter(status='cancelled').count()
    served_reservations = reservations_qs.filter(status='served').count()
    today_reservations = reservations_qs.filter(daily_menu__date=today).count()
    
    # لیست ID های رزروها
    reservation_ids = {
        'all': list(reservations_qs.values_list('id', flat=True)),
        'reserved': list(reservations_qs.filter(status='reserved').values_list('id', flat=True)),
        'cancelled': list(reservations_qs.filter(status='cancelled').values_list('id', flat=True)),
        'served': list(reservations_qs.filter(status='served').values_list('id', flat=True)),
        'today': list(reservations_qs.filter(daily_menu__date=today).values_list('id', flat=True))
    }
    
    # آمار رزروهای مهمان
    total_guest_reservations = guest_reservations_qs.count()
    reserved_guest_reservations = guest_reservations_qs.filter(status='reserved').count()
    cancelled_guest_reservations = guest_reservations_qs.filter(status='cancelled').count()
    served_guest_reservations = guest_reservations_qs.filter(status='served').count()
    today_guest_reservations = guest_reservations_qs.filter(daily_menu__date=today).count()
    
    # لیست ID های رزروهای مهمان
    guest_reservation_ids = {
        'all': list(guest_reservations_qs.values_list('id', flat=True)),
        'reserved': list(guest_reservations_qs.filter(status='reserved').values_list('id', flat=True)),
        'cancelled': list(guest_reservations_qs.filter(status='cancelled').values_list('id', flat=True)),
        'served': list(guest_reservations_qs.filter(status='served').values_list('id', flat=True)),
        'today': list(guest_reservations_qs.filter(daily_menu__date=today).values_list('id', flat=True))
    }
    
    # محاسبه مبالغ
    total_amount = Decimal('0')
    reserved_amount = Decimal('0')
    served_amount = Decimal('0')
    cancelled_amount = Decimal('0')
    
    for reservation in reservations_qs.select_related('meal_option'):
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(reservation.quantity or 1))
        total_amount += amount
        if reservation.status == 'reserved':
            reserved_amount += amount
        elif reservation.status == 'served':
            served_amount += amount
        elif reservation.status == 'cancelled':
            cancelled_amount += amount
    
    for guest_reservation in guest_reservations_qs.select_related('meal_option'):
        amount = Decimal(str(guest_reservation.amount or 0))
        total_amount += amount
        if guest_reservation.status == 'reserved':
            reserved_amount += amount
        elif guest_reservation.status == 'served':
            served_amount += amount
        elif guest_reservation.status == 'cancelled':
            cancelled_amount += amount
    
    # آمار مجموع
    total_all_reservations = total_reservations + total_guest_reservations
    total_today_reservations = today_reservations + today_guest_reservations
    
    # آمار مراکز
    total_centers = centers_qs.count()
    center_ids = {
        'all': list(centers_qs.values_list('id', flat=True))
    }
    
    # ساخت response
    stats = {
        'base_meals': {
            'total': total_base_meals,
            'active': active_base_meals,
            'inactive': total_base_meals - active_base_meals,
            'ids': base_meal_ids
        },
        'meal_options': {
            'total': total_meal_options,
            'ids': meal_option_ids
        },
        'restaurants': {
            'total': total_restaurants,
            'active': active_restaurants,
            'inactive': total_restaurants - active_restaurants,
            'ids': restaurant_ids
        },
        'users': {
            'total': total_users,
            'active': active_users,
            'inactive': total_users - active_users,
            'ids': user_ids
        },
        'centers': {
            'total': total_centers,
            'ids': center_ids
        },
        'daily_menus': {
            'total': total_daily_menus,
            'active': active_daily_menus,
            'inactive': total_daily_menus - active_daily_menus,
            'ids': daily_menu_ids
        },
        'reservations': {
            'total': total_reservations,
            'reserved': reserved_reservations,
            'cancelled': cancelled_reservations,
            'served': served_reservations,
            'today': today_reservations,
            'ids': reservation_ids
        },
        'guest_reservations': {
            'total': total_guest_reservations,
            'reserved': reserved_guest_reservations,
            'cancelled': cancelled_guest_reservations,
            'served': served_guest_reservations,
            'today': today_guest_reservations,
            'ids': guest_reservation_ids
        },
        'summary': {
            'total_reservations': total_all_reservations,
            'total_today_reservations': total_today_reservations,
            'total_amount': str(total_amount),
            'reserved_amount': str(reserved_amount),
            'served_amount': str(served_amount),
            'cancelled_amount': str(cancelled_amount)
        },
        'filters': {
            'center_id': int(center_id) if center_id else None,
            'user_id': int(user_id) if user_id else None,
            'start_date': start_date.isoformat() if start_date else None,
            'end_date': end_date.isoformat() if end_date else None
        }
    }
    
    return Response(stats)


@api_view(['GET'])
@permission_classes([StatisticsPermission])
def meal_statistics(request):
    """آمار کلی غذاها (برای سازگاری با کدهای قبلی)"""
    return comprehensive_statistics(request)


# ========== New Statistics Endpoints ==========

@extend_schema(
    operation_id='meal_statistics_by_restaurant',
    summary='Meal Statistics by Restaurant and Base Meal',
    description='Get statistics of meals grouped by restaurant, base meal, and meal options. Shows reservation counts (reserved, served, cancelled, guest) for each meal option.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'restaurant_id',
            'in': 'query',
            'description': 'فیلتر بر اساس رستوران',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def meal_statistics_by_restaurant(request):
    """آمار غذاها بر اساس رستوران، غذای پایه و اپشن‌های غذا"""
    user = request.user
    
    # اگر کاربر ادمین نیست، فقط آمار مراکز خودش را ببیند
    if not user.is_admin:
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    restaurant_id = request.query_params.get('restaurant_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(meal_option__isnull=False)
    
    guest_reservations = GuestReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(meal_option__isnull=False)
    
    # فیلتر بر اساس مرکز
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    elif not user.is_admin:
        user_centers = user.centers.all()
        if user_centers.exists():
            reservations = reservations.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
            guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
    
    # فیلتر بر اساس رستوران
    if restaurant_id:
        reservations = reservations.filter(daily_menu__restaurant__id=restaurant_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__id=restaurant_id).distinct()
    
    # فیلتر بر اساس تاریخ
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    # ساختار داده: رستوران -> غذای پایه -> اپشن‌های غذا
    restaurants_data = {}
    
    # پردازش رزروهای معمولی
    for reservation in reservations:
        if not reservation.meal_option or not reservation.meal_option.base_meal:
            continue
        
        restaurant = reservation.daily_menu.restaurant if reservation.daily_menu else None
        if not restaurant:
            continue
        
        base_meal = reservation.meal_option.base_meal
        meal_option = reservation.meal_option
        
        # ساختار رستوران
        if restaurant.id not in restaurants_data:
            restaurants_data[restaurant.id] = {
                'restaurant': {
                    'id': restaurant.id,
                    'name': restaurant.name,
                    'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
                },
                'base_meals': {}
            }
        
        # ساختار غذای پایه
        if base_meal.id not in restaurants_data[restaurant.id]['base_meals']:
            restaurants_data[restaurant.id]['base_meals'][base_meal.id] = {
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or '',
                    'is_active': base_meal.is_active
                },
                'meal_options': {}
            }
        
        # ساختار اپشن غذا
        if meal_option.id not in restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options']:
            restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options'][meal_option.id] = {
                'meal_option': {
                    'id': meal_option.id,
                    'title': meal_option.title,
                    'description': meal_option.description or '',
                    'price': float(meal_option.price)
                },
                'statistics': {
                    'reserved_count': 0,
                    'served_count': 0,
                    'cancelled_count': 0,
                    'guest_count': 0,
                    'total_count': 0
                }
            }
        
        # به‌روزرسانی آمار
        stats = restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options'][meal_option.id]['statistics']
        stats['total_count'] += reservation.quantity
        
        if reservation.status == 'reserved':
            stats['reserved_count'] += reservation.quantity
        elif reservation.status == 'served':
            stats['served_count'] += reservation.quantity
        elif reservation.status == 'cancelled':
            stats['cancelled_count'] += reservation.quantity
    
    # پردازش رزروهای مهمان
    for guest_reservation in guest_reservations:
        if not guest_reservation.meal_option or not guest_reservation.meal_option.base_meal:
            continue
        
        restaurant = guest_reservation.daily_menu.restaurant if guest_reservation.daily_menu else None
        if not restaurant:
            continue
        
        base_meal = guest_reservation.meal_option.base_meal
        meal_option = guest_reservation.meal_option
        
        if restaurant.id not in restaurants_data:
            continue
        
        if base_meal.id not in restaurants_data[restaurant.id]['base_meals']:
            continue
        
        if meal_option.id not in restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options']:
            continue
        
        # به‌روزرسانی آمار مهمان
        stats = restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options'][meal_option.id]['statistics']
        stats['guest_count'] += 1
        stats['total_count'] += 1
        
        if guest_reservation.status == 'reserved':
            stats['reserved_count'] += 1
        elif guest_reservation.status == 'served':
            stats['served_count'] += 1
        elif guest_reservation.status == 'cancelled':
            stats['cancelled_count'] += 1
    
    # تبدیل به لیست
    result = []
    for restaurant_id, restaurant_data in restaurants_data.items():
        base_meals_list = []
        for base_meal_id, base_meal_data in restaurant_data['base_meals'].items():
            meal_options_list = []
            for meal_option_id, meal_option_data in base_meal_data['meal_options'].items():
                meal_options_list.append(meal_option_data)
            
            base_meal_data['meal_options'] = meal_options_list
            base_meals_list.append(base_meal_data)
        
        restaurant_data['base_meals'] = base_meals_list
        result.append(restaurant_data)
    
    return Response(result)


@extend_schema(
    operation_id='reservations_by_base_meal',
    summary='Reservations by Base Meal',
    description='Get detailed statistics of reservations for a specific base meal. Shows users who ordered, meal options they ordered, details (quantity, amount, status), and restaurant information.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'base_meal_id',
            'in': 'query',
            'description': 'شناسه غذای پایه',
            'required': True,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def reservations_by_base_meal(request):
    """آمار بر اساس غذا - کاربرانی که سفارش دادند"""
    user = request.user
    
    # دریافت base_meal_id
    base_meal_id = request.query_params.get('base_meal_id')
    if not base_meal_id:
        return Response({
            'error': 'شناسه غذای پایه الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        base_meal_id = int(base_meal_id)
    except (ValueError, TypeError):
        return Response({
            'error': 'شناسه غذای پایه باید عدد باشد'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # بررسی وجود base_meal
    try:
        base_meal = BaseMeal.objects.get(id=base_meal_id)
    except BaseMeal.DoesNotExist:
        return Response({
            'error': 'غذای پایه یافت نشد'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # اگر کاربر ادمین نیست، فقط آمار مراکز خودش را ببیند
    if not user.is_admin:
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'user', 'meal_option', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').filter(
        meal_option__base_meal_id=base_meal_id
    )
    
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'meal_option', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').filter(
        meal_option__base_meal_id=base_meal_id
    )
    
    # فیلتر بر اساس مرکز
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    elif not user.is_admin:
        user_centers = user.centers.all()
        if user_centers.exists():
            reservations = reservations.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
            guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
    
    # فیلتر بر اساس تاریخ
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    # ساختار داده: کاربر -> رزروها
    users_data = {}
    
    # پردازش رزروهای معمولی
    for reservation in reservations:
        user_obj = reservation.user
        restaurant = reservation.daily_menu.restaurant if reservation.daily_menu else None
        meal_option = reservation.meal_option
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or ''
                },
                'reservations': []
            }
        
        # افزودن رزرو
        users_data[user_obj.id]['reservations'].append({
            'id': reservation.id,
            'restaurant': {
                'id': restaurant.id if restaurant else None,
                'name': restaurant.name if restaurant else None,
                'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()] if restaurant else []
            } if restaurant else None,
            'meal_option': {
                'id': meal_option.id if meal_option else None,
                'title': meal_option.title if meal_option else None,
                'description': meal_option.description or '' if meal_option else '',
                'price': float(meal_option.price) if meal_option else 0
            } if meal_option else None,
            'quantity': reservation.quantity,
            'amount': float(reservation.amount or 0),
            'status': reservation.status,
            'reservation_date': reservation.reservation_date.isoformat() if reservation.reservation_date else None,
            'daily_menu_date': reservation.daily_menu.date.isoformat() if reservation.daily_menu and reservation.daily_menu.date else None
        })
    
    # پردازش رزروهای مهمان
    for guest_reservation in guest_reservations:
        user_obj = guest_reservation.host_user
        restaurant = guest_reservation.daily_menu.restaurant if guest_reservation.daily_menu else None
        meal_option = guest_reservation.meal_option
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or ''
                },
                'reservations': []
            }
        
        # افزودن رزرو مهمان
        users_data[user_obj.id]['reservations'].append({
            'id': guest_reservation.id,
            'guest_name': f"{guest_reservation.guest_first_name} {guest_reservation.guest_last_name}".strip(),
            'restaurant': {
                'id': restaurant.id if restaurant else None,
                'name': restaurant.name if restaurant else None,
                'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()] if restaurant else []
            } if restaurant else None,
            'meal_option': {
                'id': meal_option.id if meal_option else None,
                'title': meal_option.title if meal_option else None,
                'description': meal_option.description or '' if meal_option else '',
                'price': float(meal_option.price) if meal_option else 0
            } if meal_option else None,
            'quantity': 1,  # مهمان همیشه 1 است
            'amount': float(guest_reservation.amount or 0),
            'status': guest_reservation.status,
            'reservation_date': guest_reservation.reservation_date.isoformat() if guest_reservation.reservation_date else None,
            'daily_menu_date': guest_reservation.daily_menu.date.isoformat() if guest_reservation.daily_menu and guest_reservation.daily_menu.date else None,
            'is_guest': True
        })
    
    # تبدیل به لیست
    result = list(users_data.values())
    
    return Response({
        'base_meal': {
            'id': base_meal.id,
            'title': base_meal.title,
            'description': base_meal.description or ''
        },
        'users': result
    })


@extend_schema(
    operation_id='user_statistics_by_date_range',
    summary='User Statistics by Date Range',
    description='Get statistics of users and total amount they ordered within a date range. Shows users with their total order amounts.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': True,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': True,
            'schema': {'type': 'string'}
        },
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def user_statistics_by_date_range(request):
    """آمار بر اساس بازه تاریخی - کاربران و جمع مبلغی که سفارش داشتند"""
    user = request.user
    
    # دریافت تاریخ‌ها
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    if not start_date or not end_date:
        return Response({
            'error': 'تاریخ شروع و پایان الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    parsed_start_date = parse_date_filter(start_date)
    parsed_end_date = parse_date_filter(end_date)
    
    if not parsed_start_date or not parsed_end_date:
        return Response({
            'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if parsed_start_date > parsed_end_date:
        return Response({
            'error': 'تاریخ شروع باید قبل از تاریخ پایان باشد'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # اگر کاربر ادمین نیست، فقط آمار مراکز خودش را ببیند
    if not user.is_admin:
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').filter(
        daily_menu__date__gte=parsed_start_date,
        daily_menu__date__lte=parsed_end_date
    )
    
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').filter(
        daily_menu__date__gte=parsed_start_date,
        daily_menu__date__lte=parsed_end_date
    )
    
    # فیلتر بر اساس مرکز
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    elif not user.is_admin:
        user_centers = user.centers.all()
        if user_centers.exists():
            reservations = reservations.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
            guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
    
    # ساختار داده: کاربر -> جمع مبلغ
    users_data = {}
    
    # پردازش رزروهای معمولی
    for reservation in reservations:
        user_obj = reservation.user
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'total_amount': 0.0,
                'reservation_count': 0,
                'guest_reservation_count': 0
            }
        
        users_data[user_obj.id]['total_amount'] += float(reservation.amount or 0)
        users_data[user_obj.id]['reservation_count'] += reservation.quantity
    
    # پردازش رزروهای مهمان
    for guest_reservation in guest_reservations:
        user_obj = guest_reservation.host_user
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'total_amount': 0.0,
                'reservation_count': 0,
                'guest_reservation_count': 0
            }
        
        users_data[user_obj.id]['total_amount'] += float(guest_reservation.amount or 0)
        users_data[user_obj.id]['guest_reservation_count'] += 1
    
    # تبدیل به لیست و مرتب‌سازی بر اساس مبلغ
    from decimal import Decimal
    result = []
    for user_data in users_data.values():
        user_data['total_amount'] = Decimal(str(user_data['total_amount']))
        result.append(user_data)
    
    # مرتب‌سازی بر اساس مبلغ (نزولی)
    result.sort(key=lambda x: x['total_amount'], reverse=True)
    
    return Response({
        'start_date': start_date,
        'end_date': end_date,
        'parsed_start_date': parsed_start_date.isoformat(),
        'parsed_end_date': parsed_end_date.isoformat(),
        'users': result,
        'total_users': len(result),
        'total_amount': sum([float(u['total_amount']) for u in result])
    })


@api_view(['GET'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def center_reservations(request, center_id):
    """رزروهای یک مرکز خاص"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center = get_object_or_404(Center, id=center_id)
    date = request.query_params.get('date')
    
    queryset = FoodReservation.objects.filter(
        daily_menu__restaurant__centers=center
    )
    
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            queryset = queryset.filter(daily_menu__date=parsed_date)
    
    serializer = SimpleFoodReservationSerializer(queryset, many=True)
    return Response(serializer.data)


# ========== Export Functions ==========

@api_view(['GET'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def export_reservations_excel(request, center_id):
    """خروجی اکسل رزروها"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    if not HAS_OPENPYXL:
        return Response({
            'error': 'کتابخانه openpyxl نصب نشده است'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    center = get_object_or_404(Center, id=center_id)
    date = request.query_params.get('date')
    
    reservations = FoodReservation.objects.filter(
        daily_menu__restaurant__centers=center
    )
    
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            reservations = reservations.filter(daily_menu__date=parsed_date)
    
    # ایجاد فایل اکسل
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"رزروهای {center.name}"
    
    # هدرها
    headers = [
        'نام کاربر', 'تاریخ', 'وعده غذایی', 'نام غذا',
        'وضعیت', 'تاریخ رزرو', 'مهلت لغو'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # داده‌ها
    for row, reservation in enumerate(reservations, 2):
        ws.cell(row=row, column=1, value=reservation.user.username)
        ws.cell(row=row, column=2, value=str(reservation.daily_menu.date))
        ws.cell(row=row, column=3, value='-')
        if reservation.meal_option:
            meal_title = f"{reservation.meal_option.base_meal.title} - {reservation.meal_option.title}"
        else:
            meal_title = 'نامشخص'
        ws.cell(row=row, column=4, value=meal_title)
        ws.cell(row=row, column=5, value=reservation.get_status_display())
        ws.cell(row=row, column=6, value=str(reservation.reservation_date))
        ws.cell(row=row, column=7, value=str(reservation.cancellation_deadline))
    
    # ذخیره فایل
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reservations_{center.name}_{date or "all"}.xlsx"'
    
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def export_reservations_pdf(request, center_id):
    """خروجی PDF رزروها"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    if not HAS_REPORTLAB:
        return Response({
            'error': 'کتابخانه reportlab نصب نشده است'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    center = get_object_or_404(Center, id=center_id)
    date = request.query_params.get('date')
    
    reservations = FoodReservation.objects.filter(
        daily_menu__restaurant__centers=center
    )
    
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            reservations = reservations.filter(daily_menu__date=parsed_date)
    
    # ایجاد PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # عنوان
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 50, f"گزارش رزروهای {center.name}")
    
    if date:
        p.setFont("Helvetica", 12)
        p.drawString(100, height - 80, f"تاریخ: {date}")
    
    # هدر جدول
    y_position = height - 120
    p.setFont("Helvetica-Bold", 10)
    headers = ['نام کاربر', 'تاریخ', 'وعده', 'غذا', 'وضعیت']
    x_positions = [100, 200, 300, 400, 500]
    
    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y_position, header)
    
    # داده‌ها
    y_position -= 20
    p.setFont("Helvetica", 8)
    
    for reservation in reservations:
        if y_position < 100:  # صفحه جدید
            p.showPage()
            y_position = height - 50
        
        data = [
            reservation.user.username,
            str(reservation.daily_menu.date),
            '-',
            f"{reservation.meal_option.base_meal.title} - {reservation.meal_option.title}" if reservation.meal_option else 'نامشخص',
            reservation.get_status_display()
        ]
        
        for i, item in enumerate(data):
            p.drawString(x_positions[i], y_position, str(item))
        
        y_position -= 15
    
    p.showPage()
    p.save()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reservations_{center.name}_{date or "all"}.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


# ========== User Reservations ==========

@extend_schema(
    operation_id='user_reservations',
    summary='Get User Reservations',
    description='Get all reservations for the authenticated user',
    tags=['User Reservations']
)
@api_view(['GET'])
@permission_classes([FoodManagementPermission])
def user_reservations(request):
    """رزروهای کاربر"""
    user = request.user
    reservations = FoodReservation.objects.filter(user=user).order_by('-reservation_date')
    
    # فیلتر بر اساس تاریخ
    date = request.query_params.get('date')
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            reservations = reservations.filter(daily_menu__date=parsed_date)
    
    # فیلتر بر اساس وضعیت
    status_filter = request.query_params.get('status')
    if status_filter:
        reservations = reservations.filter(status=status_filter)
    
    serializer = SimpleFoodReservationSerializer(reservations, many=True)
    return Response(serializer.data)


@extend_schema(
    operation_id='user_guest_reservations',
    summary='Get User Guest Reservations',
    description='Get all guest reservations made by the authenticated user',
    tags=['User Reservations']
)
@api_view(['GET'])
@permission_classes([FoodManagementPermission])
def user_guest_reservations(request):
    """رزروهای مهمان کاربر"""
    user = request.user
    guest_reservations = GuestReservation.objects.filter(host_user=user).order_by('-reservation_date')
    
    # فیلتر بر اساس تاریخ
    date = request.query_params.get('date')
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            guest_reservations = guest_reservations.filter(daily_menu__date=parsed_date)
    
    # فیلتر بر اساس وضعیت
    status_filter = request.query_params.get('status')
    if status_filter:
        guest_reservations = guest_reservations.filter(status=status_filter)
    
    serializer = SimpleGuestReservationSerializer(guest_reservations, many=True)
    return Response(serializer.data)


@extend_schema(
    operation_id='user_reservations_summary',
    summary='Get User Reservations Summary',
    description='Get summary of user reservations and guest reservations with full details',
    tags=['User Reservations']
)
@api_view(['GET'])
@permission_classes([FoodManagementPermission])
def user_reservations_summary(request):
    """خلاصه رزروهای کاربر با جزئیات کامل"""
    user = request.user
    
    # رزروهای شخصی
    personal_reservations = FoodReservation.objects.filter(user=user).select_related(
        'daily_menu', 'meal_option', 'meal_option__base_meal', 'meal_option__base_meal__restaurant'
    ).prefetch_related('meal_option__base_meal__restaurant__centers')
    personal_count = personal_reservations.count()
    personal_reserved = personal_reservations.filter(status='reserved').count()
    personal_cancelled = personal_reservations.filter(status='cancelled').count()
    personal_served = personal_reservations.filter(status='served').count()
    
    # جزئیات کامل رزروهای شخصی
    personal_reserved_items = personal_reservations.filter(status='reserved')
    personal_cancelled_items = personal_reservations.filter(status='cancelled')
    personal_served_items = personal_reservations.filter(status='served')
    personal_all_items = personal_reservations.all()
    
    # رزروهای مهمان
    guest_reservations = GuestReservation.objects.filter(host_user=user).select_related(
        'daily_menu', 'meal_option', 'meal_option__base_meal', 'meal_option__base_meal__restaurant'
    ).prefetch_related('meal_option__base_meal__restaurant__centers')
    guest_count = guest_reservations.count()
    guest_reserved = guest_reservations.filter(status='reserved').count()
    guest_cancelled = guest_reservations.filter(status='cancelled').count()
    guest_served = guest_reservations.filter(status='served').count()
    
    # جزئیات کامل رزروهای مهمان
    guest_reserved_items = guest_reservations.filter(status='reserved')
    guest_cancelled_items = guest_reservations.filter(status='cancelled')
    guest_served_items = guest_reservations.filter(status='served')
    guest_all_items = guest_reservations.all()
    
    # رزروهای امروز
    today = timezone.now().date()
    today_personal = personal_reservations.filter(daily_menu__date=today).count()
    today_guest = guest_reservations.filter(daily_menu__date=today).count()
    today_personal_items = personal_reservations.filter(daily_menu__date=today)
    today_guest_items = guest_reservations.filter(daily_menu__date=today)
    
    # استفاده از serializer برای جزئیات کامل
    personal_serializer = SimpleFoodReservationSerializer
    guest_serializer = SimpleGuestReservationSerializer
    
    return Response({
        'personal_reservations': {
            'total': personal_count,
            'reserved': personal_reserved,
            'cancelled': personal_cancelled,
            'served': personal_served,
            'today': today_personal,
            'items': {
                'all': personal_serializer(personal_all_items, many=True, context={'request': request}).data,
                'reserved': personal_serializer(personal_reserved_items, many=True, context={'request': request}).data,
                'cancelled': personal_serializer(personal_cancelled_items, many=True, context={'request': request}).data,
                'served': personal_serializer(personal_served_items, many=True, context={'request': request}).data,
                'today': personal_serializer(today_personal_items, many=True, context={'request': request}).data
            }
        },
        'guest_reservations': {
            'total': guest_count,
            'reserved': guest_reserved,
            'cancelled': guest_cancelled,
            'served': guest_served,
            'today': today_guest,
            'items': {
                'all': guest_serializer(guest_all_items, many=True, context={'request': request}).data,
                'reserved': guest_serializer(guest_reserved_items, many=True, context={'request': request}).data,
                'cancelled': guest_serializer(guest_cancelled_items, many=True, context={'request': request}).data,
                'served': guest_serializer(guest_served_items, many=True, context={'request': request}).data,
                'today': guest_serializer(today_guest_items, many=True, context={'request': request}).data
            }
        },
        'total_today': today_personal + today_guest,
        'total_all': personal_count + guest_count
    })


# ========== Employee Menu and Reservation Management ==========

@extend_schema(
    operation_id='employee_daily_menus',
    summary='Get Daily Menus for Employee',
    description='Get simplified daily menus for employee\'s assigned centers on specific date (supports multiple centers). Returns only essential information: restaurant (id, name, center), date, and meals with their options (id, title, image, price, quantity, available_quantity).',
    tags=['Employee Management'],
    responses={
        200: SimpleEmployeeDailyMenuSerializer(many=True),
        400: {'description': 'Validation error'}
    }
)
@api_view(['GET'])
@permission_classes([FoodManagementPermission])
def employee_daily_menus(request):
    """منوهای روزانه برای کارمند - خروجی ساده"""
    user = request.user
    
    # بررسی اینکه کاربر مرکز دارد
    if not user.centers.exists():
        return Response(
            {'error': 'کاربر مرکز مشخصی ندارد'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # دریافت تاریخ از query parameter
    date = request.query_params.get('date')
    if not date:
        return Response(
            {'error': 'تاریخ الزامی است'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # تبدیل تاریخ شمسی یا میلادی به فرمت مناسب
    parsed_date = parse_date_filter(date)
    if not parsed_date:
        return Response(
            {'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # دریافت منوهای روزانه برای مرکز کاربر در تاریخ مشخص
    # استفاده از distinct() برای جلوگیری از تکرار منوها وقتی یک رستوران به چندین مرکز متصل است
    daily_menus = DailyMenu.objects.filter(
        restaurant__centers__in=user.centers.all(),
        date=parsed_date,
        is_available=True
    ).select_related('restaurant').prefetch_related(
        'restaurant__centers',
        'menu_meal_options', 
        'menu_meal_options__base_meal',
        'menu_dessert_options',
        'menu_dessert_options__base_dessert'
    ).distinct().order_by('restaurant__name', 'date')
    
    serializer = SimpleEmployeeDailyMenuSerializer(daily_menus, many=True, context={'request': request})
    return Response(serializer.data)


@extend_schema(
    operation_id='employee_reservations',
    summary='Employee Reservations',
    description='Get all reservations for the authenticated employee or create a new reservation. Reservations use meal_option (DailyMenuMealOption) which is the actual meal with price and options.',
    tags=['Employee Management'],
    request=FoodReservationCreateSerializer,
    responses={
        200: SimpleFoodReservationSerializer(many=True),
        201: SimpleFoodReservationSerializer,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET', 'POST'])
@permission_classes([FoodManagementPermission])
def employee_reservations(request):
    """رزروهای کارمند"""
    user = request.user
    
    # بررسی اینکه کاربر مرکز دارد
    if not user.centers.exists():
        return Response(
            {'error': 'کاربر مرکز مشخصی ندارد'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    if request.method == 'GET':
        # لیست رزروها
        reservations = FoodReservation.objects.filter(user=user).order_by('-reservation_date')
        
        # فیلتر بر اساس تاریخ
        date = request.query_params.get('date')
        if date:
            parsed_date = parse_date_filter(date)
            if parsed_date:
                reservations = reservations.filter(daily_menu__date=parsed_date)
        
        # فیلتر بر اساس وضعیت
        status_filter = request.query_params.get('status')
        if status_filter:
            reservations = reservations.filter(status=status_filter)
        
        serializer = SimpleFoodReservationSerializer(reservations, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        # ایجاد رزرو جدید
        serializer = FoodReservationCreateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            # بررسی اینکه daily_menu متعلق به یکی از مراکز کاربر است
            daily_menu = serializer.validated_data['daily_menu']
            # بررسی اینکه آیا رستوران به یکی از مراکز کاربر مرتبط است
            restaurant_centers = daily_menu.restaurant.centers.all() if daily_menu.restaurant else []
            if not any(user.has_center(center) for center in restaurant_centers):
                return Response(
                    {'error': 'شما نمی‌توانید برای این مرکز رزرو کنید'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
            # ایجاد رزرو
            reservation = serializer.save(user=user)
            response_serializer = SimpleFoodReservationSerializer(reservation)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(
    operation_id='employee_create_guest_reservation',
    summary='Create Guest Reservation (Employee)',
    description='Create a guest reservation for the authenticated employee using meal_option (DailyMenuMealOption) which is the actual meal with price and options.',
    tags=['Employee Management']
)
@api_view(['POST'])
@permission_classes([FoodManagementPermission])
def employee_create_guest_reservation(request):
    """ایجاد رزرو مهمان برای کارمند"""
    user = request.user
    
    # بررسی اینکه کاربر مرکز دارد
    if not user.centers.exists():
        return Response(
            {'error': 'کاربر مرکز مشخصی ندارد'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    serializer = GuestReservationCreateSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        # بررسی اینکه daily_menu متعلق به یکی از مراکز کاربر است
        daily_menu = serializer.validated_data['daily_menu']
        # بررسی اینکه آیا رستوران به یکی از مراکز کاربر مرتبط است
        restaurant_centers = daily_menu.restaurant.centers.all() if daily_menu.restaurant else []
        if not any(user.has_center(center) for center in restaurant_centers):
            return Response(
                {'error': 'شما نمی‌توانید برای این مرکز رزرو کنید'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # ایجاد رزرو مهمان
        guest_reservation = serializer.save(host_user=user)
        response_serializer = SimpleGuestReservationSerializer(guest_reservation)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(
    operation_id='employee_update_reservation',
    summary='Update Food Reservation (Employee)',
    description='Update a food reservation using meal_option (DailyMenuMealOption). (only own reservations)',
    tags=['Employee Management'],
    request=FoodReservationCreateSerializer,
    responses={
        200: SimpleFoodReservationSerializer,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'},
        404: {'description': 'Reservation not found'}
    }
)
@api_view(['PUT', 'PATCH'])
@permission_classes([FoodManagementPermission])
def employee_update_reservation(request, reservation_id):
    """ویرایش رزرو غذا برای کارمند"""
    user = request.user
    
    try:
        reservation = FoodReservation.objects.get(id=reservation_id, user=user)
    except FoodReservation.DoesNotExist:
        return Response(
            {'error': 'رزرو یافت نشد'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    # بررسی اینکه رزرو قابل ویرایش است
    if reservation.status != 'reserved':
        return Response(
            {'error': 'فقط رزروهای فعال قابل ویرایش هستند'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    serializer = FoodReservationCreateSerializer(
        reservation, 
        data=request.data, 
        partial=request.method == 'PATCH',
        context={'request': request}
    )
    
    if serializer.is_valid():
        # بررسی اینکه daily_menu متعلق به یکی از مراکز کاربر است
        daily_menu = serializer.validated_data['daily_menu']
        # بررسی اینکه آیا رستوران به یکی از مراکز کاربر مرتبط است
        restaurant_centers = daily_menu.restaurant.centers.all() if daily_menu.restaurant else []
        if not any(user.has_center(center) for center in restaurant_centers):
            return Response(
                {'error': 'شما نمی‌توانید برای این مرکز رزرو کنید'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer.save()
        response_serializer = SimpleFoodReservationSerializer(reservation)
        return Response(response_serializer.data)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(
    operation_id='employee_update_guest_reservation',
    summary='Update Guest Reservation (Employee)',
    description='Update a guest reservation using meal_option (DailyMenuMealOption). (only own guest reservations)',
    tags=['Employee Management'],
    request=GuestReservationCreateSerializer,
    responses={
        200: SimpleGuestReservationSerializer,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'},
        404: {'description': 'Guest reservation not found'}
    }
)
@api_view(['PUT', 'PATCH'])
@permission_classes([FoodManagementPermission])
def employee_update_guest_reservation(request, guest_reservation_id):
    """ویرایش رزرو مهمان برای کارمند"""
    user = request.user
    
    try:
        guest_reservation = GuestReservation.objects.get(id=guest_reservation_id, host_user=user)
    except GuestReservation.DoesNotExist:
        return Response(
            {'error': 'رزرو مهمان یافت نشد'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    # بررسی اینکه رزرو قابل ویرایش است
    if guest_reservation.status != 'reserved':
        return Response(
            {'error': 'فقط رزروهای فعال قابل ویرایش هستند'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    serializer = GuestReservationCreateSerializer(
        guest_reservation, 
        data=request.data, 
        partial=request.method == 'PATCH',
        context={'request': request}
    )
    
    if serializer.is_valid():
        # بررسی اینکه daily_menu متعلق به یکی از مراکز کاربر است
        daily_menu = serializer.validated_data['daily_menu']
        # بررسی اینکه آیا رستوران به یکی از مراکز کاربر مرتبط است
        restaurant_centers = daily_menu.restaurant.centers.all() if daily_menu.restaurant else []
        if not any(user.has_center(center) for center in restaurant_centers):
            return Response(
                {'error': 'شما نمی‌توانید برای این مرکز رزرو کنید'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer.save()
        response_serializer = SimpleGuestReservationSerializer(guest_reservation)
        return Response(response_serializer.data)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(
    operation_id='employee_cancel_reservation',
    summary='Cancel Food Reservation (Employee)',
    description='Cancel a food reservation (only own reservations)',
    tags=['Employee Management']
)
@api_view(['POST'])
@permission_classes([FoodManagementPermission])
def employee_cancel_reservation(request, reservation_id):
    """لغو رزرو غذا برای کارمند"""
    user = request.user
    
    try:
        reservation = FoodReservation.objects.get(id=reservation_id, user=user)
    except FoodReservation.DoesNotExist:
        return Response(
            {'error': 'رزرو یافت نشد'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    if reservation.cancel():
        response_serializer = SimpleFoodReservationSerializer(reservation)
        return Response(response_serializer.data)
    else:
        return Response(
            {'error': 'رزرو قابل لغو نیست'}, 
            status=status.HTTP_400_BAD_REQUEST
        )


@extend_schema(
    operation_id='employee_cancel_guest_reservation',
    summary='Cancel Guest Reservation (Employee)',
    description='Cancel a guest reservation (only own guest reservations)',
    tags=['Employee Management']
)
@api_view(['POST'])
@permission_classes([FoodManagementPermission])
def employee_cancel_guest_reservation(request, guest_reservation_id):
    """لغو رزرو مهمان برای کارمند"""
    user = request.user
    
    try:
        guest_reservation = GuestReservation.objects.get(id=guest_reservation_id, host_user=user)
    except GuestReservation.DoesNotExist:
        return Response(
            {'error': 'رزرو مهمان یافت نشد'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    if guest_reservation.cancel():
        response_serializer = SimpleGuestReservationSerializer(guest_reservation)
        return Response(response_serializer.data)
    else:
        return Response(
            {'error': 'رزرو مهمان قابل لغو نیست'}, 
            status=status.HTTP_400_BAD_REQUEST
        )


# ========== Report Views ==========

@extend_schema(
    summary='Report by Meal Option',
    description='Complete report of reservations by meal options (DailyMenuMealOption)',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_meal_option(request):
    """گزارش بر اساس DailyMenuMealOption"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.all()
    
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
    
    # گروه‌بندی بر اساس DailyMenuMealOption
    from django.db.models import Sum, Count, Q
    
    meal_options_data = {}
    
    for reservation in reservations.select_related('meal_option', 'meal_option__base_meal', 'meal_option__daily_menu', 'meal_option__daily_menu__restaurant', 'daily_menu__restaurant').prefetch_related('daily_menu__restaurant__centers'):
        if not reservation.meal_option:
            continue
        
        meal_option = reservation.meal_option
        meal_option_id = meal_option.id
        
        if meal_option_id not in meal_options_data:
            # دریافت restaurant از meal_option
            restaurant = None
            if meal_option.base_meal and meal_option.base_meal.restaurant:
                restaurant = meal_option.base_meal.restaurant
            elif meal_option.daily_menu and meal_option.daily_menu.restaurant:
                restaurant = meal_option.daily_menu.restaurant
            
            meal_options_data[meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': meal_option.title,
                'base_meal_title': meal_option.base_meal.title if meal_option.base_meal else '',
                'restaurant_name': restaurant.name if restaurant else '',
                'restaurant_id': restaurant.id if restaurant else None,
                'center_name': ', '.join([c.name for c in restaurant.centers.all()]) if restaurant and restaurant.centers.exists() else '',
                'center_id': restaurant.centers.first().id if restaurant and restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'served_amount': 0,
            }
        
        data = meal_options_data[meal_option_id]
        data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
            data['served_amount'] += float(reservation.amount or 0)
        
        data['total_amount'] += float(reservation.amount or 0)
    
    from decimal import Decimal
    for data in meal_options_data.values():
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['reserved_amount'] = Decimal(str(data['reserved_amount']))
        data['served_amount'] = Decimal(str(data['served_amount']))
    
    serializer = MealOptionReportSerializer(list(meal_options_data.values()), many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Report by Base Meal',
    description='Complete report of reservations by base meals',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_base_meal(request):
    """گزارش بر اساس BaseMeal"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__restaurant',
        'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(meal_option__isnull=False)
    
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
    
    # گروه‌بندی بر اساس BaseMeal
    base_meals_data = {}
    meal_options_by_base_meal = {}
    
    for reservation in reservations:
        if not reservation.meal_option or not reservation.meal_option.base_meal:
            continue
        
        base_meal = reservation.meal_option.base_meal
        base_meal_id = base_meal.id
        meal_option_id = reservation.meal_option.id
        
        if base_meal_id not in base_meals_data:
            base_meals_data[base_meal_id] = {
                'base_meal_id': base_meal_id,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.restaurant.name if reservation.meal_option.restaurant else '',
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'meal_options_count': 0,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
            }
            meal_options_by_base_meal[base_meal_id] = {}
        
        data = base_meals_data[base_meal_id]
        data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
        
        data['total_amount'] += float(reservation.amount or 0)
        
        # جمع‌آوری داده‌های MealOption
        if meal_option_id not in meal_options_by_base_meal[base_meal_id]:
            meal_options_by_base_meal[base_meal_id][meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': reservation.meal_option.title,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.restaurant.name if reservation.meal_option.restaurant else '',
                'restaurant_id': reservation.meal_option.restaurant.id if reservation.meal_option.restaurant else None,
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': reservation.daily_menu.restaurant.centers.first().id if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'served_amount': 0,
            }
            data['meal_options_count'] += 1
        
        meal_option_data = meal_options_by_base_meal[base_meal_id][meal_option_id]
        meal_option_data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            meal_option_data['reserved_count'] += 1
            meal_option_data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            meal_option_data['cancelled_count'] += 1
        elif reservation.status == 'served':
            meal_option_data['served_count'] += 1
            meal_option_data['served_amount'] += float(reservation.amount or 0)
        
        meal_option_data['total_amount'] += float(reservation.amount or 0)
    
    from decimal import Decimal
    result = []
    for base_meal_id, data in base_meals_data.items():
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['meal_options'] = []
        
        for meal_option_data in meal_options_by_base_meal[base_meal_id].values():
            meal_option_data['total_amount'] = Decimal(str(meal_option_data['total_amount']))
            meal_option_data['reserved_amount'] = Decimal(str(meal_option_data['reserved_amount']))
            meal_option_data['served_amount'] = Decimal(str(meal_option_data['served_amount']))
            data['meal_options'].append(meal_option_data)
        
        result.append(data)
    
    serializer = BaseMealReportSerializer(result, many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Report by User',
    description='Complete report of reservations by users',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_user(request):
    """گزارش بر اساس کاربر"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').all()
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').all()
    
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    # گروه‌بندی بر اساس کاربر
    users_data = {}
    
    # رزروهای معمولی
    for reservation in reservations:
        user_id = reservation.user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': reservation.user.username,
                'full_name': reservation.user.get_full_name(),
                'employee_number': reservation.user.employee_number or '',
                'center_name': ', '.join([c.name for c in reservation.user.centers.all()]) if reservation.user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
            }
        
        data = users_data[user_id]
        data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
        
        data['total_amount'] += float(reservation.amount or 0)
    
    # رزروهای مهمان
    for guest_reservation in guest_reservations:
        user_id = guest_reservation.host_user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': guest_reservation.host_user.username,
                'full_name': guest_reservation.host_user.get_full_name(),
                'employee_number': guest_reservation.host_user.employee_number or '',
                'center_name': ', '.join([c.name for c in guest_reservation.host_user.centers.all()]) if guest_reservation.host_user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
            }
        
        data = users_data[user_id]
        data['total_guest_reservations'] += 1
        data['total_amount'] += float(guest_reservation.amount or 0)
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(guest_reservation.amount or 0)
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    from decimal import Decimal
    for data in users_data.values():
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['reserved_amount'] = Decimal(str(data['reserved_amount']))
    
    serializer = UserReportSerializer(list(users_data.values()), many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Report by Date',
    description='Complete report of reservations by date',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_date(request):
    """گزارش بر اساس تاریخ"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'daily_menu__restaurant__centers'
    ).all()
    guest_reservations = GuestReservation.objects.select_related(
        'daily_menu__restaurant__centers'
    ).all()
    
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    # گروه‌بندی بر اساس تاریخ
    dates_data = {}
    
    for reservation in reservations:
        if not reservation.daily_menu:
            continue
        
        date = reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': date2jalali(date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_reservations'] += 1
        
        center_name = ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': 0,
            }
        
        data['centers'][center_name]['total_reservations'] += 1
        data['centers'][center_name]['total_amount'] += float(reservation.amount or 0)
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
        
        data['total_amount'] += float(reservation.amount or 0)
    
    for guest_reservation in guest_reservations:
        if not guest_reservation.daily_menu:
            continue
        
        date = guest_reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': date2jalali(date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_guest_reservations'] += 1
        
        center_name = ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': 0,
            }
        
        data['total_amount'] += float(guest_reservation.amount or 0)
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(guest_reservation.amount or 0)
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    from decimal import Decimal
    result = []
    for date in sorted(dates_data.keys(), reverse=True):
        data = dates_data[date]
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['reserved_amount'] = Decimal(str(data['reserved_amount']))
        data['centers'] = list(data['centers'].values())
        for center_data in data['centers']:
            center_data['total_amount'] = Decimal(str(center_data['total_amount']))
        result.append(data)
    
    serializer = DateReportSerializer(result, many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Comprehensive Report',
    description='Comprehensive report including all types of reports',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def comprehensive_report(request):
    """گزارش جامع"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__restaurant',
        'daily_menu__restaurant__centers', 'user'
    ).all()
    guest_reservations = GuestReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__restaurant',
        'daily_menu__restaurant__centers', 'host_user'
    ).all()
    
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    from decimal import Decimal
    
    # آمار کلی
    total_reservations = reservations.count()
    total_guest_reservations = guest_reservations.count()
    total_amount = Decimal('0')
    reserved_amount = Decimal('0')
    served_amount = Decimal('0')
    cancelled_amount = Decimal('0')
    
    for reservation in reservations:
        amount = Decimal(str(reservation.amount or 0))
        total_amount += amount
        if reservation.status == 'reserved':
            reserved_amount += amount
        elif reservation.status == 'served':
            served_amount += amount
        elif reservation.status == 'cancelled':
            cancelled_amount += amount
    
    for guest_reservation in guest_reservations:
        amount = Decimal(str(guest_reservation.amount or 0))
        total_amount += amount
        if guest_reservation.status == 'reserved':
            reserved_amount += amount
        elif guest_reservation.status == 'served':
            served_amount += amount
        elif guest_reservation.status == 'cancelled':
            cancelled_amount += amount
    
    # گزارش بر اساس MealOption
    meal_options_data = {}
    for reservation in reservations:
        if not reservation.meal_option:
            continue
        
        meal_option = reservation.meal_option
        meal_option_id = meal_option.id
        
        if meal_option_id not in meal_options_data:
            meal_options_data[meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': meal_option.title,
                'base_meal_title': meal_option.base_meal.title if meal_option.base_meal else '',
                'restaurant_name': meal_option.restaurant.name if meal_option.restaurant else '',
                'restaurant_id': meal_option.restaurant.id if meal_option.restaurant else None,
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': reservation.daily_menu.restaurant.centers.first().id if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
        
        data = meal_options_data[meal_option_id]
        data['total_reservations'] += 1
        amount = Decimal(str(reservation.amount or 0))
        data['total_amount'] += amount
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
            data['served_amount'] += amount
    
    # گزارش بر اساس BaseMeal
    base_meals_data = {}
    meal_options_by_base_meal = {}
    
    for reservation in reservations:
        if not reservation.meal_option or not reservation.meal_option.base_meal:
            continue
        
        base_meal = reservation.meal_option.base_meal
        base_meal_id = base_meal.id
        meal_option_id = reservation.meal_option.id
        
        if base_meal_id not in base_meals_data:
            base_meals_data[base_meal_id] = {
                'base_meal_id': base_meal_id,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.restaurant.name if reservation.meal_option.restaurant else '',
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'meal_options_count': 0,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
            }
            meal_options_by_base_meal[base_meal_id] = {}
        
        data = base_meals_data[base_meal_id]
        data['total_reservations'] += 1
        amount = Decimal(str(reservation.amount or 0))
        data['total_amount'] += amount
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
        
        if meal_option_id not in meal_options_by_base_meal[base_meal_id]:
            meal_options_by_base_meal[base_meal_id][meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': reservation.meal_option.title,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.restaurant.name if reservation.meal_option.restaurant else '',
                'restaurant_id': reservation.meal_option.restaurant.id if reservation.meal_option.restaurant else None,
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': reservation.daily_menu.restaurant.centers.first().id if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
            data['meal_options_count'] += 1
        
        meal_option_data = meal_options_by_base_meal[base_meal_id][meal_option_id]
        meal_option_data['total_reservations'] += 1
        meal_option_data['total_amount'] += amount
        
        if reservation.status == 'reserved':
            meal_option_data['reserved_count'] += 1
            meal_option_data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            meal_option_data['cancelled_count'] += 1
        elif reservation.status == 'served':
            meal_option_data['served_count'] += 1
            meal_option_data['served_amount'] += amount
    
    result_base_meals = []
    for base_meal_id, data in base_meals_data.items():
        data['meal_options'] = list(meal_options_by_base_meal[base_meal_id].values())
        result_base_meals.append(data)
    
    # گزارش بر اساس کاربر
    users_data = {}
    
    for reservation in reservations:
        user_id = reservation.user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': reservation.user.username,
                'full_name': reservation.user.get_full_name(),
                'employee_number': reservation.user.employee_number or '',
                'center_name': ', '.join([c.name for c in reservation.user.centers.all()]) if reservation.user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
            }
        
        data = users_data[user_id]
        data['total_reservations'] += 1
        amount = Decimal(str(reservation.amount or 0))
        data['total_amount'] += amount
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
    
    for guest_reservation in guest_reservations:
        user_id = guest_reservation.host_user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': guest_reservation.host_user.username,
                'full_name': guest_reservation.host_user.get_full_name(),
                'employee_number': guest_reservation.host_user.employee_number or '',
                'center_name': ', '.join([c.name for c in guest_reservation.host_user.centers.all()]) if guest_reservation.host_user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
            }
        
        data = users_data[user_id]
        data['total_guest_reservations'] += 1
        amount = Decimal(str(guest_reservation.amount or 0))
        data['total_amount'] += amount
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    # گزارش بر اساس تاریخ
    dates_data = {}
    
    for reservation in reservations:
        if not reservation.daily_menu:
            continue
        
        date = reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': date2jalali(date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_reservations'] += 1
        amount = Decimal(str(reservation.amount or 0))
        data['total_amount'] += amount
        
        center_name = ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': Decimal('0'),
            }
        
        data['centers'][center_name]['total_reservations'] += 1
        data['centers'][center_name]['total_amount'] += amount
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
    
    for guest_reservation in guest_reservations:
        if not guest_reservation.daily_menu:
            continue
        
        date = guest_reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': date2jalali(date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_guest_reservations'] += 1
        amount = Decimal(str(guest_reservation.amount or 0))
        data['total_amount'] += amount
        
        center_name = ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': Decimal('0'),
            }
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    result_dates = []
    for date in sorted(dates_data.keys(), reverse=True):
        data = dates_data[date]
        data['centers'] = list(data['centers'].values())
        result_dates.append(data)
    
    # ساخت گزارش جامع
    report_data = {
        'total_reservations': total_reservations,
        'total_guest_reservations': total_guest_reservations,
        'total_amount': total_amount,
        'reserved_amount': reserved_amount,
        'served_amount': served_amount,
        'cancelled_amount': cancelled_amount,
        'by_meal_option': list(meal_options_data.values()),
        'by_base_meal': result_base_meals,
        'by_user': list(users_data.values()),
        'by_date': result_dates,
    }
    
    serializer = ComprehensiveReportSerializer(report_data)
    return Response(serializer.data)


@extend_schema(
    summary='Detailed Reservations Report',
    description='Complete detailed report of reservations with all information',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'user_id',
            'in': 'query',
            'description': 'فیلتر بر اساس کاربر',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'status',
            'in': 'query',
            'description': 'فیلتر بر اساس وضعیت (reserved, cancelled, served)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def detailed_reservations_report(request):
    """گزارش جزئیات رزروها"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    user_id = request.query_params.get('user_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    status_filter = request.query_params.get('status')
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'user', 'meal_option', 'meal_option__base_meal', 'meal_option__restaurant',
        'daily_menu', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').all()
    
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    
    if user_id:
        reservations = reservations.filter(user_id=user_id)
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
    
    if status_filter:
        reservations = reservations.filter(status=status_filter)
    
    serializer = DetailedReservationReportSerializer(reservations, many=True, context={'request': request})
    return Response(serializer.data)

