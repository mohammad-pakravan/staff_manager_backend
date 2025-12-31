"""
Views for reports app - Statistics and Reports
"""
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema
from drf_spectacular.types import OpenApiTypes
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from io import BytesIO
from decimal import Decimal
from datetime import datetime
import jdatetime

# Optional imports for export functionality
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from apps.food_management.permissions import (
    IsFoodAdminOrSystemAdmin,
    StatisticsPermission,
    UserReportPermission
)
from django.db.models import Q
from apps.food_management.models import (
    BaseMeal, BaseDessert, DailyMenu, DailyMenuMealOption, DailyMenuDessertOption,
    FoodReservation, GuestReservation, Restaurant,
    DessertReservation, GuestDessertReservation
)
from apps.centers.models import Center
from apps.food_management.utils import parse_date_filter
from apps.reports.serializers import (
    MealOptionReportSerializer,
    BaseMealReportSerializer,
    UserReportSerializer,
    DateReportSerializer,
    DetailedReservationReportSerializer,
    ComprehensiveReportSerializer,
    UserWithMealOptionSerializer,
    UserWithDessertOptionSerializer
)
from apps.reservations.serializers import SimpleFoodReservationSerializer


# ========== Helper Functions ==========

def get_accessible_centers(user):
    """
    تعیین مراکز قابل دسترسی برای کاربر
    - System Admin: None (همه مراکز)
    - Food Admin: مراکز اختصاص داده شده به Food Admin
    - Other users: مراکز خودشان
    """
    if user.role == 'sys_admin':
        return None  # دسترسی به همه مراکز
    elif user.role == 'admin_food':
        if not user.centers.exists():
            return []  # بدون مرکز
        return user.centers.all()
    else:
        if not user.centers.exists():
            return []
        return user.centers.all()


def filter_reservations_by_center(reservations, center_id=None, accessible_centers=None):
    """
    فیلتر رزروها بر اساس مرکز - شامل رزروهایی که daily_menu حذف شده است
    """
    if center_id:
        # شامل رزروهایی که daily_menu=None است یا daily_menu.restaurant.centers شامل center_id است
        reservations = reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
    elif accessible_centers is not None:
        # شامل رزروهایی که daily_menu=None است یا daily_menu.restaurant.centers در accessible_centers است
        reservations = reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__in=accessible_centers)
        ).distinct()
    return reservations


def filter_reservations_by_date(reservations, start_date=None, end_date=None):
    """
    فیلتر رزروها بر اساس تاریخ - شامل رزروهایی که daily_menu حذف شده است
    """
    if start_date:
        # شامل رزروهایی که daily_menu=None است یا daily_menu.date >= start_date
        reservations = reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__date__gte=start_date)
        )
    if end_date:
        # شامل رزروهایی که daily_menu=None است یا daily_menu.date <= end_date
        reservations = reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__date__lte=end_date)
        )
    return reservations


def filter_guest_reservations_by_center(guest_reservations, center_id=None, accessible_centers=None):
    """
    فیلتر رزروهای مهمان بر اساس مرکز - شامل رزروهایی که daily_menu حذف شده است
    """
    if center_id:
        guest_reservations = guest_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
    elif accessible_centers is not None:
        guest_reservations = guest_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__in=accessible_centers)
        ).distinct()
    return guest_reservations


def filter_guest_reservations_by_date(guest_reservations, start_date=None, end_date=None):
    """
    فیلتر رزروهای مهمان بر اساس تاریخ - شامل رزروهایی که daily_menu حذف شده است
    """
    if start_date:
        guest_reservations = guest_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__date__gte=start_date)
        )
    if end_date:
        guest_reservations = guest_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__date__lte=end_date)
        )
    return guest_reservations


# ========== Statistics Views ==========

@extend_schema(
    operation_id='comprehensive_statistics',
    summary='Comprehensive Statistics',
    description='Get comprehensive statistics including: base meals, meal options (DailyMenuMealOption), restaurants, users, daily menus, reservations, and guest reservations. Supports filters by center_id, user_id, start_date, and end_date. Employees see statistics for all their assigned centers.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'user_id',
            'in': 'query',
            'description': 'فیلتر بر اساس کاربر',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def comprehensive_statistics(request):
    """آمار جامع با امکان فیلتر"""
    user = request.user
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    user_id = request.query_params.get('user_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = None
    if user.role == 'sys_admin':
        # System Admin می‌تواند همه مراکز را ببیند
        accessible_centers = None
    elif user.role == 'admin_food':
        # Food Admin فقط مراکز خودش را می‌بیند
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_403_FORBIDDEN)
        accessible_centers = user.centers.all()
    else:
        # سایر کاربران (Employee) فقط مراکز خودش
        if not user.centers.exists():
            return Response({
                'error': 'کاربر مرکز مشخصی ندارد'
            }, status=status.HTTP_403_FORBIDDEN)
        accessible_centers = user.centers.all()
    
    # ساخت queryset های پایه
    base_meals_qs = BaseMeal.objects.all()
    meal_options_qs = DailyMenuMealOption.objects.all()
    restaurants_qs = Restaurant.objects.all()
    from apps.accounts.models import User
    users_qs = User.objects.all()
    reservations_qs = FoodReservation.objects.all()
    guest_reservations_qs = GuestReservation.objects.all()
    daily_menus_qs = DailyMenu.objects.all()
    centers_qs = Center.objects.all()
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        base_meals_qs = base_meals_qs.filter(restaurant__centers__id=center_id).distinct()
        meal_options_qs = meal_options_qs.filter(base_meal__restaurant__centers__id=center_id).distinct()
        restaurants_qs = restaurants_qs.filter(centers__id=center_id).distinct()
        users_qs = users_qs.filter(centers__id=center_id).distinct()
        reservations_qs = reservations_qs.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations_qs = guest_reservations_qs.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        daily_menus_qs = daily_menus_qs.filter(restaurant__centers__id=center_id).distinct()
        centers_qs = centers_qs.filter(id=center_id)
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        user_centers = accessible_centers
        base_meals_qs = base_meals_qs.filter(restaurant__centers__in=user_centers).distinct()
        meal_options_qs = meal_options_qs.filter(base_meal__restaurant__centers__in=user_centers).distinct()
        restaurants_qs = restaurants_qs.filter(centers__in=user_centers).distinct()
        users_qs = users_qs.filter(centers__in=user_centers).distinct()
        reservations_qs = reservations_qs.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
        guest_reservations_qs = guest_reservations_qs.filter(daily_menu__restaurant__centers__in=user_centers).distinct()
        daily_menus_qs = daily_menus_qs.filter(restaurant__centers__in=user_centers).distinct()
        centers_qs = centers_qs.filter(id__in=user_centers.values_list('id', flat=True))
    
    # فیلتر بر اساس کاربر
    if user_id:
        reservations_qs = reservations_qs.filter(user_id=user_id)
        guest_reservations_qs = guest_reservations_qs.filter(host_user_id=user_id)
    
    # فیلتر بر اساس تاریخ
    if start_date:
        reservations_qs = reservations_qs.filter(daily_menu__date__gte=start_date)
        guest_reservations_qs = guest_reservations_qs.filter(daily_menu__date__gte=start_date)
        daily_menus_qs = daily_menus_qs.filter(date__gte=start_date)
    
    if end_date:
        reservations_qs = reservations_qs.filter(daily_menu__date__lte=end_date)
        guest_reservations_qs = guest_reservations_qs.filter(daily_menu__date__lte=end_date)
        daily_menus_qs = daily_menus_qs.filter(date__lte=end_date)
    
    # محاسبه آمار
    today = timezone.now().date()
    
    # آمار غذاها
    total_base_meals = base_meals_qs.count()
    active_base_meals = base_meals_qs.filter(is_active=True).count()
    base_meal_ids = {
        'all': list(base_meals_qs.values_list('id', flat=True)),
        'active': list(base_meals_qs.filter(is_active=True).values_list('id', flat=True)),
        'inactive': list(base_meals_qs.filter(is_active=False).values_list('id', flat=True))
    }
    
    # آمار اپشن‌ها
    total_meal_options = meal_options_qs.count()
    meal_option_ids = {
        'all': list(meal_options_qs.values_list('id', flat=True))
    }
    
    # آمار رستوران‌ها
    total_restaurants = restaurants_qs.count()
    active_restaurants = restaurants_qs.filter(is_active=True).count()
    restaurant_ids = {
        'all': list(restaurants_qs.values_list('id', flat=True)),
        'active': list(restaurants_qs.filter(is_active=True).values_list('id', flat=True)),
        'inactive': list(restaurants_qs.filter(is_active=False).values_list('id', flat=True))
    }
    
    # آمار کاربران
    total_users = users_qs.count()
    active_users = users_qs.filter(is_active=True).count()
    user_ids = {
        'all': list(users_qs.values_list('id', flat=True)),
        'active': list(users_qs.filter(is_active=True).values_list('id', flat=True)),
        'inactive': list(users_qs.filter(is_active=False).values_list('id', flat=True))
    }
    
    # آمار منوهای روزانه
    total_daily_menus = daily_menus_qs.count()
    active_daily_menus = daily_menus_qs.filter(is_available=True).count()
    daily_menu_ids = {
        'all': list(daily_menus_qs.values_list('id', flat=True)),
        'active': list(daily_menus_qs.filter(is_available=True).values_list('id', flat=True)),
        'inactive': list(daily_menus_qs.filter(is_available=False).values_list('id', flat=True))
    }
    
    # آمار رزروها
    total_reservations = reservations_qs.count()
    reserved_reservations = reservations_qs.filter(status='reserved').count()
    cancelled_reservations = reservations_qs.filter(status='cancelled').count()
    served_reservations = reservations_qs.filter(status='served').count()
    today_reservations = reservations_qs.filter(daily_menu__date=today).count()
    
    # لیست ID های رزروها
    reservation_ids = {
        'all': list(reservations_qs.values_list('id', flat=True)),
        'reserved': list(reservations_qs.filter(status='reserved').values_list('id', flat=True)),
        'cancelled': list(reservations_qs.filter(status='cancelled').values_list('id', flat=True)),
        'served': list(reservations_qs.filter(status='served').values_list('id', flat=True)),
        'today': list(reservations_qs.filter(daily_menu__date=today).values_list('id', flat=True))
    }
    
    # آمار رزروهای مهمان
    total_guest_reservations = guest_reservations_qs.count()
    reserved_guest_reservations = guest_reservations_qs.filter(status='reserved').count()
    cancelled_guest_reservations = guest_reservations_qs.filter(status='cancelled').count()
    served_guest_reservations = guest_reservations_qs.filter(status='served').count()
    today_guest_reservations = guest_reservations_qs.filter(daily_menu__date=today).count()
    
    # لیست ID های رزروهای مهمان
    guest_reservation_ids = {
        'all': list(guest_reservations_qs.values_list('id', flat=True)),
        'reserved': list(guest_reservations_qs.filter(status='reserved').values_list('id', flat=True)),
        'cancelled': list(guest_reservations_qs.filter(status='cancelled').values_list('id', flat=True)),
        'served': list(guest_reservations_qs.filter(status='served').values_list('id', flat=True)),
        'today': list(guest_reservations_qs.filter(daily_menu__date=today).values_list('id', flat=True))
    }
    
    # محاسبه مبالغ
    total_amount = Decimal('0')
    reserved_amount = Decimal('0')
    served_amount = Decimal('0')
    cancelled_amount = Decimal('0')
    
    for reservation in reservations_qs.select_related('meal_option'):
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(reservation.quantity or 1))
        total_amount += amount
        if reservation.status == 'reserved':
            reserved_amount += amount
        elif reservation.status == 'served':
            served_amount += amount
        elif reservation.status == 'cancelled':
            cancelled_amount += amount
    
    for guest_reservation in guest_reservations_qs.select_related('meal_option'):
        amount = Decimal(str(guest_reservation.amount or 0))
        total_amount += amount
        if guest_reservation.status == 'reserved':
            reserved_amount += amount
        elif guest_reservation.status == 'served':
            served_amount += amount
        elif guest_reservation.status == 'cancelled':
            cancelled_amount += amount
    
    # آمار مجموع
    total_all_reservations = total_reservations + total_guest_reservations
    total_today_reservations = today_reservations + today_guest_reservations
    
    # آمار مراکز
    total_centers = centers_qs.count()
    center_ids = {
        'all': list(centers_qs.values_list('id', flat=True))
    }
    
    # ساخت response
    stats = {
        'base_meals': {
            'total': total_base_meals,
            'active': active_base_meals,
            'inactive': total_base_meals - active_base_meals,
            'ids': base_meal_ids
        },
        'meal_options': {
            'total': total_meal_options,
            'ids': meal_option_ids
        },
        'restaurants': {
            'total': total_restaurants,
            'active': active_restaurants,
            'inactive': total_restaurants - active_restaurants,
            'ids': restaurant_ids
        },
        'users': {
            'total': total_users,
            'active': active_users,
            'inactive': total_users - active_users,
            'ids': user_ids
        },
        'centers': {
            'total': total_centers,
            'ids': center_ids
        },
        'daily_menus': {
            'total': total_daily_menus,
            'active': active_daily_menus,
            'inactive': total_daily_menus - active_daily_menus,
            'ids': daily_menu_ids
        },
        'reservations': {
            'total': total_reservations,
            'reserved': reserved_reservations,
            'cancelled': cancelled_reservations,
            'served': served_reservations,
            'today': today_reservations,
            'ids': reservation_ids
        },
        'guest_reservations': {
            'total': total_guest_reservations,
            'reserved': reserved_guest_reservations,
            'cancelled': cancelled_guest_reservations,
            'served': served_guest_reservations,
            'today': today_guest_reservations,
            'ids': guest_reservation_ids
        },
        'summary': {
            'total_reservations': total_all_reservations,
            'total_today_reservations': total_today_reservations,
            'total_amount': str(total_amount),
            'reserved_amount': str(reserved_amount),
            'served_amount': str(served_amount),
            'cancelled_amount': str(cancelled_amount)
        },
        'filters': {
            'center_id': int(center_id) if center_id else None,
            'user_id': int(user_id) if user_id else None,
            'start_date': start_date.isoformat() if start_date else None,
            'end_date': end_date.isoformat() if end_date else None
        }
    }
    
    return Response(stats)


@api_view(['GET'])
@permission_classes([StatisticsPermission])
def meal_statistics(request):
    """آمار کلی غذاها (برای سازگاری با کدهای قبلی)"""
    return comprehensive_statistics(request)


# ========== Additional Statistics Views ==========

@extend_schema(
    operation_id='meal_statistics_by_restaurant',
    summary='Meal Statistics by Restaurant and Base Meal',
    description='Get statistics of meals grouped by restaurant, base meal, and meal options. Shows reservation counts (reserved, served, cancelled, guest) for each meal option.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'restaurant_id',
            'in': 'query',
            'description': 'فیلتر بر اساس رستوران',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def meal_statistics_by_restaurant(request):
    """آمار غذاها بر اساس رستوران، غذای پایه و اپشن‌های غذا"""
    user = request.user
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    restaurant_id = request.query_params.get('restaurant_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(meal_option__isnull=False)
    
    guest_reservations = GuestReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(meal_option__isnull=False)
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
    
    # فیلتر بر اساس رستوران
    if restaurant_id:
        reservations = reservations.filter(daily_menu__restaurant__id=restaurant_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__id=restaurant_id).distinct()
    
    # فیلتر بر اساس تاریخ
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    # ساختار داده: رستوران -> غذای پایه -> اپشن‌های غذا
    restaurants_data = {}
    
    # پردازش رزروهای معمولی
    for reservation in reservations:
        if not reservation.meal_option or not reservation.meal_option.base_meal:
            continue
        
        restaurant = reservation.daily_menu.restaurant if reservation.daily_menu else None
        if not restaurant:
            continue
        
        base_meal = reservation.meal_option.base_meal
        meal_option = reservation.meal_option
        
        # ساختار رستوران
        if restaurant.id not in restaurants_data:
            restaurants_data[restaurant.id] = {
                'restaurant': {
                    'id': restaurant.id,
                    'name': restaurant.name,
                    'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
                },
                'base_meals': {}
            }
        
        # ساختار غذای پایه
        if base_meal.id not in restaurants_data[restaurant.id]['base_meals']:
            restaurants_data[restaurant.id]['base_meals'][base_meal.id] = {
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or '',
                    'is_active': base_meal.is_active
                },
                'meal_options': {}
            }
        
        # ساختار اپشن غذا
        if meal_option.id not in restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options']:
            restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options'][meal_option.id] = {
                'meal_option': {
                    'id': meal_option.id,
                    'title': meal_option.title,
                    'description': meal_option.description or '',
                    'price': float(meal_option.price)
                },
                'statistics': {
                    'reserved_count': 0,
                    'served_count': 0,
                    'cancelled_count': 0,
                    'guest_count': 0,
                    'total_count': 0
                }
            }
        
        # به‌روزرسانی آمار
        stats = restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options'][meal_option.id]['statistics']
        stats['total_count'] += reservation.quantity
        
        if reservation.status == 'reserved':
            stats['reserved_count'] += reservation.quantity
        elif reservation.status == 'served':
            stats['served_count'] += reservation.quantity
        elif reservation.status == 'cancelled':
            stats['cancelled_count'] += reservation.quantity
    
    # پردازش رزروهای مهمان
    for guest_reservation in guest_reservations:
        if not guest_reservation.meal_option or not guest_reservation.meal_option.base_meal:
            continue
        
        restaurant = guest_reservation.daily_menu.restaurant if guest_reservation.daily_menu else None
        if not restaurant:
            continue
        
        base_meal = guest_reservation.meal_option.base_meal
        meal_option = guest_reservation.meal_option
        
        if restaurant.id not in restaurants_data:
            continue
        
        if base_meal.id not in restaurants_data[restaurant.id]['base_meals']:
            continue
        
        if meal_option.id not in restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options']:
            continue
        
        # به‌روزرسانی آمار مهمان
        stats = restaurants_data[restaurant.id]['base_meals'][base_meal.id]['meal_options'][meal_option.id]['statistics']
        stats['guest_count'] += 1
        stats['total_count'] += 1
        
        if guest_reservation.status == 'reserved':
            stats['reserved_count'] += 1
        elif guest_reservation.status == 'served':
            stats['served_count'] += 1
        elif guest_reservation.status == 'cancelled':
            stats['cancelled_count'] += 1
    
    # تبدیل به لیست
    result = []
    for restaurant_id, restaurant_data in restaurants_data.items():
        base_meals_list = []
        for base_meal_id, base_meal_data in restaurant_data['base_meals'].items():
            meal_options_list = []
            for meal_option_id, meal_option_data in base_meal_data['meal_options'].items():
                meal_options_list.append(meal_option_data)
            
            base_meal_data['meal_options'] = meal_options_list
            base_meals_list.append(base_meal_data)
        
        restaurant_data['base_meals'] = base_meals_list
        result.append(restaurant_data)
    
    return Response(result)


@extend_schema(
    operation_id='reservations_by_base_meal',
    summary='Reservations by Base Meal',
    description='Get detailed statistics of reservations for a specific base meal. Shows users who ordered, meal options they ordered, details (quantity, amount, status), and restaurant information.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'base_meal_id',
            'in': 'query',
            'description': 'شناسه غذای پایه',
            'required': True,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def reservations_by_base_meal(request):
    """آمار بر اساس غذا - کاربرانی که سفارش دادند"""
    user = request.user
    
    # دریافت base_meal_id
    base_meal_id = request.query_params.get('base_meal_id')
    if not base_meal_id:
        return Response({
            'error': 'شناسه غذای پایه الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        base_meal_id = int(base_meal_id)
    except (ValueError, TypeError):
        return Response({
            'error': 'شناسه غذای پایه باید عدد باشد'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # بررسی وجود base_meal
    try:
        base_meal = BaseMeal.objects.get(id=base_meal_id)
    except BaseMeal.DoesNotExist:
        return Response({
            'error': 'غذای پایه یافت نشد'
        }, status=status.HTTP_404_NOT_FOUND)
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها - حذف رزروهای کنسل شده
    reservations = FoodReservation.objects.select_related(
        'user', 'meal_option', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').filter(
        meal_option__base_meal_id=base_meal_id
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'meal_option', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').filter(
        meal_option__base_meal_id=base_meal_id
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
    
    # فیلتر بر اساس تاریخ
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    # ساختار داده: کاربر -> رزروها
    users_data = {}
    
    # پردازش رزروهای معمولی - فقط رزروهای غیر کنسل شده
    for reservation in reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if reservation.status == 'cancelled':
            continue
        
        user_obj = reservation.user
        restaurant = reservation.daily_menu.restaurant if reservation.daily_menu else None
        meal_option = reservation.meal_option
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or ''
                },
                'reservations': []
            }
        
        # افزودن رزرو
        users_data[user_obj.id]['reservations'].append({
            'id': reservation.id,
            'restaurant': {
                'id': restaurant.id if restaurant else None,
                'name': restaurant.name if restaurant else None,
                'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()] if restaurant else []
            } if restaurant else None,
            'meal_option': {
                'id': meal_option.id if meal_option else None,
                'title': meal_option.title if meal_option else None,
                'description': meal_option.description or '' if meal_option else '',
                'price': float(meal_option.price) if meal_option else 0
            } if meal_option else None,
            'quantity': reservation.quantity,
            'amount': float(reservation.amount or 0),
            'status': reservation.status,
            'reservation_date': reservation.reservation_date.isoformat() if reservation.reservation_date else None,
            'daily_menu_date': reservation.daily_menu.date.isoformat() if reservation.daily_menu and reservation.daily_menu.date else None
        })
    
    # پردازش رزروهای مهمان - فقط رزروهای غیر کنسل شده
    for guest_reservation in guest_reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if guest_reservation.status == 'cancelled':
            continue
        
        user_obj = guest_reservation.host_user
        restaurant = guest_reservation.daily_menu.restaurant if guest_reservation.daily_menu else None
        meal_option = guest_reservation.meal_option
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or ''
                },
                'reservations': []
            }
        
        # افزودن رزرو مهمان
        users_data[user_obj.id]['reservations'].append({
            'id': guest_reservation.id,
            'guest_name': f"{guest_reservation.guest_first_name} {guest_reservation.guest_last_name}".strip(),
            'restaurant': {
                'id': restaurant.id if restaurant else None,
                'name': restaurant.name if restaurant else None,
                'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()] if restaurant else []
            } if restaurant else None,
            'meal_option': {
                'id': meal_option.id if meal_option else None,
                'title': meal_option.title if meal_option else None,
                'description': meal_option.description or '' if meal_option else '',
                'price': float(meal_option.price) if meal_option else 0
            } if meal_option else None,
            'quantity': 1,  # مهمان همیشه 1 است
            'amount': float(guest_reservation.amount or 0),
            'status': guest_reservation.status,
            'reservation_date': guest_reservation.reservation_date.isoformat() if guest_reservation.reservation_date else None,
            'daily_menu_date': guest_reservation.daily_menu.date.isoformat() if guest_reservation.daily_menu and guest_reservation.daily_menu.date else None,
            'is_guest': True
        })
    
    # تبدیل به لیست
    result = list(users_data.values())
    
    return Response({
        'base_meal': {
            'id': base_meal.id,
            'title': base_meal.title,
            'description': base_meal.description or ''
        },
        'users': result
    })


@extend_schema(
    operation_id='user_statistics_by_date_range',
    summary='User Statistics by Date Range',
    description='Get statistics of users and total amount they ordered within a date range. Shows users with their total order amounts.',
    tags=['Statistics'],
    parameters=[
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': True,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': True,
            'schema': {'type': 'string'}
        },
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def user_statistics_by_date_range(request):
    """آمار بر اساس بازه تاریخی - کاربران و جمع مبلغی که سفارش داشتند"""
    user = request.user
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # دریافت تاریخ‌ها
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    if not start_date or not end_date:
        return Response({
            'error': 'تاریخ شروع و پایان الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    parsed_start_date = parse_date_filter(start_date)
    parsed_end_date = parse_date_filter(end_date)
    
    if not parsed_start_date or not parsed_end_date:
        return Response({
            'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if parsed_start_date > parsed_end_date:
        return Response({
            'error': 'تاریخ شروع باید قبل از تاریخ پایان باشد'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # دریافت فیلترها
    center_id = request.query_params.get('center_id')
    
    # فیلتر رزروها - حذف رزروهای کنسل شده
    reservations = FoodReservation.objects.select_related(
        'user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').filter(
        daily_menu__date__gte=parsed_start_date,
        daily_menu__date__lte=parsed_end_date
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').filter(
        daily_menu__date__gte=parsed_start_date,
        daily_menu__date__lte=parsed_end_date
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
    # فیلتر رزروهای دسر - حذف رزروهای کنسل شده
    dessert_reservations = DessertReservation.objects.select_related(
        'user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').filter(
        daily_menu__date__gte=parsed_start_date,
        daily_menu__date__lte=parsed_end_date
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
    guest_dessert_reservations = GuestDessertReservation.objects.select_related(
        'host_user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').filter(
        daily_menu__date__gte=parsed_start_date,
        daily_menu__date__lte=parsed_end_date
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        dessert_reservations = dessert_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
        dessert_reservations = dessert_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
        guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
    
    # ساختار داده: کاربر -> جمع مبلغ
    users_data = {}
    
    # پردازش رزروهای معمولی غذا - فقط رزروهای غیر کنسل شده
    for reservation in reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if reservation.status == 'cancelled':
            continue
        
        user_obj = reservation.user
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'total_amount': 0.0,
                'reservation_count': 0,
                'guest_reservation_count': 0,
                'dessert_reservation_count': 0,
                'guest_dessert_reservation_count': 0
            }
        
        users_data[user_obj.id]['total_amount'] += float(reservation.amount or 0)
        users_data[user_obj.id]['reservation_count'] += reservation.quantity
    
    # پردازش رزروهای مهمان غذا - فقط رزروهای غیر کنسل شده
    for guest_reservation in guest_reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if guest_reservation.status == 'cancelled':
            continue
        
        user_obj = guest_reservation.host_user
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'total_amount': 0.0,
                'reservation_count': 0,
                'guest_reservation_count': 0,
                'dessert_reservation_count': 0,
                'guest_dessert_reservation_count': 0
            }
        
        users_data[user_obj.id]['total_amount'] += float(guest_reservation.amount or 0)
        users_data[user_obj.id]['guest_reservation_count'] += 1
    
    # پردازش رزروهای دسر - فقط رزروهای غیر کنسل شده
    for dessert_reservation in dessert_reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if dessert_reservation.status == 'cancelled':
            continue
        
        user_obj = dessert_reservation.user
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'total_amount': 0.0,
                'reservation_count': 0,
                'guest_reservation_count': 0,
                'dessert_reservation_count': 0,
                'guest_dessert_reservation_count': 0
            }
        
        # محاسبه مبلغ دسر با در نظر گیری quantity
        dessert_amount = float(dessert_reservation.amount or 0) * float(dessert_reservation.quantity or 1)
        users_data[user_obj.id]['total_amount'] += dessert_amount
        users_data[user_obj.id]['dessert_reservation_count'] += dessert_reservation.quantity or 1
    
    # پردازش رزروهای مهمان دسر - فقط رزروهای غیر کنسل شده
    for guest_dessert_reservation in guest_dessert_reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if guest_dessert_reservation.status == 'cancelled':
            continue
        
        user_obj = guest_dessert_reservation.host_user
        
        if user_obj.id not in users_data:
            users_data[user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'total_amount': 0.0,
                'reservation_count': 0,
                'guest_reservation_count': 0,
                'dessert_reservation_count': 0,
                'guest_dessert_reservation_count': 0
            }
        
        users_data[user_obj.id]['total_amount'] += float(guest_dessert_reservation.amount or 0)
        users_data[user_obj.id]['guest_dessert_reservation_count'] += 1
    
    # تبدیل به لیست و مرتب‌سازی بر اساس مبلغ
    result = []
    for user_data in users_data.values():
        user_data['total_amount'] = Decimal(str(user_data['total_amount']))
        result.append(user_data)
    
    # مرتب‌سازی بر اساس مبلغ (نزولی)
    result.sort(key=lambda x: x['total_amount'], reverse=True)
    
    return Response({
        'start_date': start_date,
        'end_date': end_date,
        'parsed_start_date': parsed_start_date.isoformat(),
        'parsed_end_date': parsed_end_date.isoformat(),
        'users': result,
        'total_users': len(result),
        'total_amount': sum([float(u['total_amount']) for u in result])
    })


@api_view(['GET'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def center_reservations(request, center_id):
    """رزروهای یک مرکز خاص"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center = get_object_or_404(Center, id=center_id)
    date = request.query_params.get('date')
    
    queryset = FoodReservation.objects.filter(
        daily_menu__restaurant__centers=center
    )
    
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            queryset = queryset.filter(daily_menu__date=parsed_date)
    
    serializer = SimpleFoodReservationSerializer(queryset, many=True)
    return Response(serializer.data)


# ========== Export Functions ==========

@api_view(['GET'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def export_reservations_excel(request, center_id):
    """خروجی اکسل رزروها"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    if not HAS_OPENPYXL:
        return Response({
            'error': 'کتابخانه openpyxl نصب نشده است'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    center = get_object_or_404(Center, id=center_id)
    date = request.query_params.get('date')
    
    reservations = FoodReservation.objects.filter(
        daily_menu__restaurant__centers=center
    )
    
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            reservations = reservations.filter(daily_menu__date=parsed_date)
    
    # ایجاد فایل اکسل
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"رزروهای {center.name}"
    
    # هدرها
    headers = [
        'نام کاربر', 'تاریخ', 'وعده غذایی', 'نام غذا',
        'وضعیت', 'تاریخ رزرو', 'مهلت لغو'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # داده‌ها
    for row, reservation in enumerate(reservations, 2):
        ws.cell(row=row, column=1, value=reservation.user.username)
        ws.cell(row=row, column=2, value=str(reservation.daily_menu.date))
        ws.cell(row=row, column=3, value='-')
        if reservation.meal_option:
            meal_title = f"{reservation.meal_option.base_meal.title} - {reservation.meal_option.title}"
        else:
            meal_title = 'نامشخص'
        ws.cell(row=row, column=4, value=meal_title)
        ws.cell(row=row, column=5, value=reservation.get_status_display())
        ws.cell(row=row, column=6, value=str(reservation.reservation_date))
        ws.cell(row=row, column=7, value=str(reservation.cancellation_deadline))
    
    # ذخیره فایل
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reservations_{center.name}_{date or "all"}.xlsx"'
    
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsFoodAdminOrSystemAdmin])
def export_reservations_pdf(request, center_id):
    """خروجی PDF رزروها"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    if not HAS_REPORTLAB:
        return Response({
            'error': 'کتابخانه reportlab نصب نشده است'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    center = get_object_or_404(Center, id=center_id)
    date = request.query_params.get('date')
    
    reservations = FoodReservation.objects.filter(
        daily_menu__restaurant__centers=center
    )
    
    if date:
        parsed_date = parse_date_filter(date)
        if parsed_date:
            reservations = reservations.filter(daily_menu__date=parsed_date)
    
    # ایجاد PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # عنوان
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 50, f"گزارش رزروهای {center.name}")
    
    if date:
        p.setFont("Helvetica", 12)
        p.drawString(100, height - 80, f"تاریخ: {date}")
    
    # هدر جدول
    y_position = height - 120
    p.setFont("Helvetica-Bold", 10)
    headers = ['نام کاربر', 'تاریخ', 'وعده', 'غذا', 'وضعیت']
    x_positions = [100, 200, 300, 400, 500]
    
    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y_position, header)
    
    # داده‌ها
    y_position -= 20
    p.setFont("Helvetica", 8)
    
    for reservation in reservations:
        if y_position < 100:  # صفحه جدید
            p.showPage()
            y_position = height - 50
        
        data = [
            reservation.user.username,
            str(reservation.daily_menu.date),
            '-',
            f"{reservation.meal_option.base_meal.title} - {reservation.meal_option.title}" if reservation.meal_option else 'نامشخص',
            reservation.get_status_display()
        ]
        
        for i, item in enumerate(data):
            p.drawString(x_positions[i], y_position, str(item))
        
        y_position -= 15
    
    p.showPage()
    p.save()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reservations_{center.name}_{date or "all"}.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response


# ========== Report Views ==========

@extend_schema(
    summary='Report by Meal Option',
    description='Complete report of reservations by meal options (DailyMenuMealOption)',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_meal_option(request):
    """گزارش بر اساس DailyMenuMealOption"""
    user = request.user
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # بررسی دسترسی - فقط System Admin و Food Admin
    if user.role not in ['sys_admin', 'admin_food']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.all()
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = filter_reservations_by_center(reservations, center_id=center_id)
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = filter_reservations_by_center(reservations, accessible_centers=accessible_centers)
    
    # فیلتر بر اساس تاریخ
    reservations = filter_reservations_by_date(reservations, start_date=start_date, end_date=end_date)
    
    # گروه‌بندی بر اساس DailyMenuMealOption
    meal_options_data = {}
    
    for reservation in reservations.select_related('meal_option', 'meal_option__base_meal', 'meal_option__daily_menu', 'meal_option__daily_menu__restaurant', 'daily_menu__restaurant').prefetch_related('daily_menu__restaurant__centers'):
        if not reservation.meal_option:
            continue
        
        meal_option = reservation.meal_option
        meal_option_id = meal_option.id
        
        if meal_option_id not in meal_options_data:
            # دریافت restaurant از meal_option
            restaurant = None
            if meal_option.base_meal and meal_option.base_meal.restaurant:
                restaurant = meal_option.base_meal.restaurant
            elif meal_option.daily_menu and meal_option.daily_menu.restaurant:
                restaurant = meal_option.daily_menu.restaurant
            
            meal_options_data[meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': meal_option.title,
                'base_meal_title': meal_option.base_meal.title if meal_option.base_meal else '',
                'restaurant_name': restaurant.name if restaurant else '',
                'restaurant_id': restaurant.id if restaurant else None,
                'center_name': ', '.join([c.name for c in restaurant.centers.all()]) if restaurant and restaurant.centers.exists() else '',
                'center_id': restaurant.centers.first().id if restaurant and restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'served_amount': 0,
            }
        
        data = meal_options_data[meal_option_id]
        data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
            data['served_amount'] += float(reservation.amount or 0)
        
        data['total_amount'] += float(reservation.amount or 0)
    
    for data in meal_options_data.values():
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['reserved_amount'] = Decimal(str(data['reserved_amount']))
        data['served_amount'] = Decimal(str(data['served_amount']))
    
    serializer = MealOptionReportSerializer(list(meal_options_data.values()), many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Report by Base Meal',
    description='Complete report of reservations by base meals',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_base_meal(request):
    """گزارش بر اساس BaseMeal"""
    user = request.user
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # بررسی دسترسی - فقط System Admin و Food Admin
    if user.role not in ['sys_admin', 'admin_food']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(meal_option__isnull=False)
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = filter_reservations_by_center(reservations, center_id=center_id)
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = filter_reservations_by_center(reservations, accessible_centers=accessible_centers)
    
    # فیلتر بر اساس تاریخ
    reservations = filter_reservations_by_date(reservations, start_date=start_date, end_date=end_date)
    
    # گروه‌بندی بر اساس BaseMeal
    base_meals_data = {}
    meal_options_by_base_meal = {}
    
    for reservation in reservations:
        if not reservation.meal_option or not reservation.meal_option.base_meal:
            continue
        
        base_meal = reservation.meal_option.base_meal
        base_meal_id = base_meal.id
        meal_option_id = reservation.meal_option.id
        
        if base_meal_id not in base_meals_data:
            base_meals_data[base_meal_id] = {
                'base_meal_id': base_meal_id,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.daily_menu.restaurant.name if reservation.meal_option and reservation.meal_option.daily_menu and reservation.meal_option.daily_menu.restaurant else '',
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'meal_options_count': 0,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
            }
            meal_options_by_base_meal[base_meal_id] = {}
        
        data = base_meals_data[base_meal_id]
        data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
        
        data['total_amount'] += float(reservation.amount or 0)
        
        # جمع‌آوری داده‌های MealOption
        if meal_option_id not in meal_options_by_base_meal[base_meal_id]:
            meal_options_by_base_meal[base_meal_id][meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': reservation.meal_option.title,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.daily_menu.restaurant.name if reservation.meal_option and reservation.meal_option.daily_menu and reservation.meal_option.daily_menu.restaurant else '',
                'restaurant_id': reservation.meal_option.daily_menu.restaurant.id if reservation.meal_option and reservation.meal_option.daily_menu and reservation.meal_option.daily_menu.restaurant else None,
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': reservation.daily_menu.restaurant.centers.first().id if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'served_amount': 0,
            }
            data['meal_options_count'] += 1
        
        meal_option_data = meal_options_by_base_meal[base_meal_id][meal_option_id]
        meal_option_data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            meal_option_data['reserved_count'] += 1
            meal_option_data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            meal_option_data['cancelled_count'] += 1
        elif reservation.status == 'served':
            meal_option_data['served_count'] += 1
            meal_option_data['served_amount'] += float(reservation.amount or 0)
        
        meal_option_data['total_amount'] += float(reservation.amount or 0)
    
    result = []
    for base_meal_id, data in base_meals_data.items():
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['meal_options'] = []
        
        for meal_option_data in meal_options_by_base_meal[base_meal_id].values():
            meal_option_data['total_amount'] = Decimal(str(meal_option_data['total_amount']))
            meal_option_data['reserved_amount'] = Decimal(str(meal_option_data['reserved_amount']))
            meal_option_data['served_amount'] = Decimal(str(meal_option_data['served_amount']))
            data['meal_options'].append(meal_option_data)
        
        result.append(data)
    
    serializer = BaseMealReportSerializer(result, many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Report by User',
    description='Complete report of reservations by users',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_user(request):
    """گزارش بر اساس کاربر"""
    if not request.user.is_admin:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').all()
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').all()
    
    if center_id:
        reservations = reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_reservations = guest_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    # گروه‌بندی بر اساس کاربر
    users_data = {}
    
    # رزروهای معمولی
    for reservation in reservations:
        user_id = reservation.user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': reservation.user.username,
                'full_name': reservation.user.get_full_name(),
                'employee_number': reservation.user.employee_number or '',
                'center_name': ', '.join([c.name for c in reservation.user.centers.all()]) if reservation.user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
            }
        
        data = users_data[user_id]
        data['total_reservations'] += 1
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
        
        data['total_amount'] += float(reservation.amount or 0)
    
    # رزروهای مهمان
    for guest_reservation in guest_reservations:
        user_id = guest_reservation.host_user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': guest_reservation.host_user.username,
                'full_name': guest_reservation.host_user.get_full_name(),
                'employee_number': guest_reservation.host_user.employee_number or '',
                'center_name': ', '.join([c.name for c in guest_reservation.host_user.centers.all()]) if guest_reservation.host_user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
            }
        
        data = users_data[user_id]
        data['total_guest_reservations'] += 1
        data['total_amount'] += float(guest_reservation.amount or 0)
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(guest_reservation.amount or 0)
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    for data in users_data.values():
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['reserved_amount'] = Decimal(str(data['reserved_amount']))
    
    serializer = UserReportSerializer(list(users_data.values()), many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Report by Date',
    description='Complete report of reservations by date',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def report_by_date(request):
    """گزارش بر اساس تاریخ"""
    user = request.user
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # بررسی دسترسی - فقط System Admin و Food Admin
    if user.role not in ['sys_admin', 'admin_food']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'daily_menu__restaurant'
    ).prefetch_related(
        'daily_menu__restaurant__centers'
    ).all()
    guest_reservations = GuestReservation.objects.select_related(
        'daily_menu__restaurant'
    ).prefetch_related(
        'daily_menu__restaurant__centers'
    ).all()
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = filter_reservations_by_center(reservations, center_id=center_id)
        guest_reservations = filter_guest_reservations_by_center(guest_reservations, center_id=center_id)
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = filter_reservations_by_center(reservations, accessible_centers=accessible_centers)
        guest_reservations = filter_guest_reservations_by_center(guest_reservations, accessible_centers=accessible_centers)
    
    # فیلتر بر اساس تاریخ
    reservations = filter_reservations_by_date(reservations, start_date=start_date, end_date=end_date)
    guest_reservations = filter_guest_reservations_by_date(guest_reservations, start_date=start_date, end_date=end_date)
    
    # گروه‌بندی بر اساس تاریخ
    dates_data = {}
    
    for reservation in reservations:
        if not reservation.daily_menu:
            continue
        
        date = reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': jdatetime.date.fromgregorian(date=date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_reservations'] += 1
        
        center_name = ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': 0,
            }
        
        data['centers'][center_name]['total_reservations'] += 1
        data['centers'][center_name]['total_amount'] += float(reservation.amount or 0)
        
        if reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(reservation.amount or 0)
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif reservation.status == 'served':
            data['served_count'] += 1
        
        data['total_amount'] += float(reservation.amount or 0)
    
    for guest_reservation in guest_reservations:
        if not guest_reservation.daily_menu:
            continue
        
        date = guest_reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': jdatetime.date.fromgregorian(date=date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': 0,
                'reserved_amount': 0,
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_guest_reservations'] += 1
        
        center_name = ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': 0,
            }
        
        data['total_amount'] += float(guest_reservation.amount or 0)
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += float(guest_reservation.amount or 0)
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    result = []
    for date in sorted(dates_data.keys(), reverse=True):
        data = dates_data[date]
        data['total_amount'] = Decimal(str(data['total_amount']))
        data['reserved_amount'] = Decimal(str(data['reserved_amount']))
        data['centers'] = list(data['centers'].values())
        for center_data in data['centers']:
            center_data['total_amount'] = Decimal(str(center_data['total_amount']))
        result.append(data)
    
    serializer = DateReportSerializer(result, many=True)
    return Response(serializer.data)


@extend_schema(
    summary='Comprehensive Report',
    description='Comprehensive report including all types of reports',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def comprehensive_report(request):
    """گزارش جامع"""
    user = request.user
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # بررسی دسترسی - فقط System Admin و Food Admin
    if user.role not in ['sys_admin', 'admin_food']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    # فیلتر رزروها
    reservations = FoodReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant', 'user'
    ).prefetch_related(
        'daily_menu__restaurant__centers', 'meal_option__daily_menu__restaurant__centers'
    ).exclude(status='cancelled')

    guest_reservations = GuestReservation.objects.select_related(
        'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant', 'host_user'
    ).prefetch_related(
        'daily_menu__restaurant__centers', 'meal_option__daily_menu__restaurant__centers'
    ).exclude(status='cancelled')

    # فیلتر رزروهای دسر
    dessert_reservations = DessertReservation.objects.select_related(
        'dessert_option', 'dessert_option__base_dessert', 'dessert_option__daily_menu',
        'dessert_option__daily_menu__restaurant', 'daily_menu__restaurant', 'user'
    ).prefetch_related(
        'daily_menu__restaurant__centers', 'dessert_option__daily_menu__restaurant__centers'
    ).exclude(status='cancelled')

    guest_dessert_reservations = GuestDessertReservation.objects.select_related(
        'dessert_option', 'dessert_option__base_dessert', 'dessert_option__daily_menu',
        'dessert_option__daily_menu__restaurant', 'daily_menu__restaurant', 'host_user'
    ).prefetch_related(
        'daily_menu__restaurant__centers', 'dessert_option__daily_menu__restaurant__centers'
    ).exclude(status='cancelled')
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = filter_reservations_by_center(reservations, center_id=center_id)
        guest_reservations = filter_guest_reservations_by_center(guest_reservations, center_id=center_id)
        dessert_reservations = filter_reservations_by_center(dessert_reservations, center_id=center_id)
        guest_dessert_reservations = filter_guest_reservations_by_center(guest_dessert_reservations, center_id=center_id)
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = filter_reservations_by_center(reservations, accessible_centers=accessible_centers)
        guest_reservations = filter_guest_reservations_by_center(guest_reservations, accessible_centers=accessible_centers)
        dessert_reservations = filter_reservations_by_center(dessert_reservations, accessible_centers=accessible_centers)
        guest_dessert_reservations = filter_guest_reservations_by_center(guest_dessert_reservations, accessible_centers=accessible_centers)
    
    # فیلتر بر اساس تاریخ
    reservations = filter_reservations_by_date(reservations, start_date=start_date, end_date=end_date)
    guest_reservations = filter_guest_reservations_by_date(guest_reservations, start_date=start_date, end_date=end_date)
    dessert_reservations = filter_reservations_by_date(dessert_reservations, start_date=start_date, end_date=end_date)
    guest_dessert_reservations = filter_reservations_by_date(guest_dessert_reservations, start_date=start_date, end_date=end_date)
    
    # آمار کلی
    total_reservations = reservations.count()
    total_guest_reservations = guest_reservations.count()
    total_dessert_reservations = dessert_reservations.count()
    total_guest_dessert_reservations = guest_dessert_reservations.count()
    total_amount = Decimal('0')
    reserved_amount = Decimal('0')
    served_amount = Decimal('0')
    cancelled_amount = Decimal('0')
    
    for reservation in reservations:
        quantity = reservation.quantity or 1
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(quantity))
        total_amount += amount
        if reservation.status == 'reserved':
            reserved_amount += amount
        elif reservation.status == 'served':
            served_amount += amount
        elif reservation.status == 'cancelled':
            cancelled_amount += amount
    
    for guest_reservation in guest_reservations:
        amount = Decimal(str(guest_reservation.amount or 0))
        total_amount += amount
        if guest_reservation.status == 'reserved':
            reserved_amount += amount
        elif guest_reservation.status == 'served':
            served_amount += amount
        elif guest_reservation.status == 'cancelled':
            cancelled_amount += amount
    
    for dessert_reservation in dessert_reservations:
        amount = Decimal(str(dessert_reservation.amount or 0)) * Decimal(str(dessert_reservation.quantity or 1))
        total_amount += amount
        if dessert_reservation.status == 'reserved':
            reserved_amount += amount
        elif dessert_reservation.status == 'served':
            served_amount += amount
        elif dessert_reservation.status == 'cancelled':
            cancelled_amount += amount
    
    for guest_dessert_reservation in guest_dessert_reservations:
        amount = Decimal(str(guest_dessert_reservation.amount or 0))
        total_amount += amount
        if guest_dessert_reservation.status == 'reserved':
            reserved_amount += amount
        elif guest_dessert_reservation.status == 'served':
            served_amount += amount
        elif guest_dessert_reservation.status == 'cancelled':
            cancelled_amount += amount
    
    # گزارش بر اساس MealOption
    meal_options_data = {}
    for reservation in reservations:
        meal_option = reservation.meal_option
        meal_option_id = None
        meal_option_title = ''
        base_meal_title = ''
        restaurant_name = ''
        restaurant_id = None
        center_name = ''
        center_id = None
        
        if meal_option:
            # اگر meal_option وجود دارد
            meal_option_id = meal_option.id
            meal_option_title = meal_option.title
            base_meal_title = meal_option.base_meal.title if meal_option.base_meal else ''
            restaurant_name = meal_option.daily_menu.restaurant.name if meal_option.daily_menu and meal_option.daily_menu.restaurant else ''
            restaurant_id = meal_option.daily_menu.restaurant.id if meal_option.daily_menu and meal_option.daily_menu.restaurant else None
            center_name = ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else ''
            center_id = reservation.daily_menu.restaurant.centers.first().id if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else None
        elif reservation.meal_option_info:
            # اگر meal_option حذف شده، از meal_option_info استفاده می‌کنیم
            meal_option_id = f"deleted_{reservation.id}"  # استفاده از ID منحصر به فرد برای رزروهای حذف شده
            meal_option_title = reservation.meal_option_info
            base_meal_title = 'نامشخص'
            # سعی می‌کنیم اطلاعات را از daily_menu_info استخراج کنیم
            if reservation.daily_menu and reservation.daily_menu.restaurant:
                restaurant_name = reservation.daily_menu.restaurant.name
                restaurant_id = reservation.daily_menu.restaurant.id
                center_name = ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu.restaurant.centers.exists() else ''
                center_id = reservation.daily_menu.restaurant.centers.first().id if reservation.daily_menu.restaurant.centers.exists() else None
            elif reservation.daily_menu_info:
                # استخراج از daily_menu_info
                if 'مرکز:' in reservation.daily_menu_info:
                    center_name = reservation.daily_menu_info.split('مرکز:')[1].split('-')[0].strip()
                restaurant_name = 'رستوران حذف شده'
        else:
            # اگر هیچ اطلاعاتی وجود ندارد
            continue
        
        if meal_option_id not in meal_options_data:
            meal_options_data[meal_option_id] = {
                'meal_option_id': meal_option_id if isinstance(meal_option_id, int) else None,
                'meal_option_title': meal_option_title,
                'base_meal_title': base_meal_title,
                'restaurant_name': restaurant_name,
                'restaurant_id': restaurant_id,
                'center_name': center_name,
                'center_id': center_id,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
        
        data = meal_options_data[meal_option_id]
        quantity = reservation.quantity or 1
        data['total_reservations'] += quantity
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(quantity))
        data['total_amount'] += amount

        if reservation.status == 'reserved':
            data['reserved_count'] += quantity
            data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += quantity
        elif reservation.status == 'served':
            data['served_count'] += quantity
            data['served_amount'] += amount
    
    # اضافه کردن رزروهای مهمان غذا به گزارش MealOption
    for guest_reservation in guest_reservations:
        meal_option = guest_reservation.meal_option
        meal_option_id = None
        meal_option_title = ''
        base_meal_title = ''
        restaurant_name = ''
        restaurant_id = None
        center_name = ''
        center_id = None
        
        if meal_option:
            # اگر meal_option وجود دارد
            meal_option_id = meal_option.id
            meal_option_title = meal_option.title
            base_meal_title = meal_option.base_meal.title if meal_option.base_meal else ''
            restaurant_name = meal_option.daily_menu.restaurant.name if meal_option.daily_menu and meal_option.daily_menu.restaurant else ''
            restaurant_id = meal_option.daily_menu.restaurant.id if meal_option.daily_menu and meal_option.daily_menu.restaurant else None
            center_name = ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else ''
            center_id = guest_reservation.daily_menu.restaurant.centers.first().id if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else None
        elif guest_reservation.meal_option_info:
            # اگر meal_option حذف شده، از meal_option_info استفاده می‌کنیم
            meal_option_id = f"deleted_guest_{guest_reservation.id}"  # استفاده از ID منحصر به فرد برای رزروهای حذف شده
            meal_option_title = guest_reservation.meal_option_info
            base_meal_title = 'نامشخص'
            # سعی می‌کنیم اطلاعات را از daily_menu_info استخراج کنیم
            if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant:
                restaurant_name = guest_reservation.daily_menu.restaurant.name
                restaurant_id = guest_reservation.daily_menu.restaurant.id
                center_name = ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu.restaurant.centers.exists() else ''
                center_id = guest_reservation.daily_menu.restaurant.centers.first().id if guest_reservation.daily_menu.restaurant.centers.exists() else None
            elif guest_reservation.daily_menu_info:
                # استخراج از daily_menu_info
                if 'مرکز:' in guest_reservation.daily_menu_info:
                    center_name = guest_reservation.daily_menu_info.split('مرکز:')[1].split('-')[0].strip()
                restaurant_name = 'رستوران حذف شده'
        else:
            # اگر هیچ اطلاعاتی وجود ندارد
            continue
        
        if meal_option_id not in meal_options_data:
            meal_options_data[meal_option_id] = {
                'meal_option_id': meal_option_id if isinstance(meal_option_id, int) else None,
                'meal_option_title': meal_option_title,
                'base_meal_title': base_meal_title,
                'restaurant_name': restaurant_name,
                'restaurant_id': restaurant_id,
                'center_name': center_name,
                'center_id': center_id,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
        
        data = meal_options_data[meal_option_id]
        data['total_reservations'] += 1
        amount = Decimal(str(guest_reservation.amount or 0))
        data['total_amount'] += amount
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
            data['served_amount'] += amount
    
    # گزارش بر اساس BaseMeal
    base_meals_data = {}
    meal_options_by_base_meal = {}
    
    for reservation in reservations:
        if not reservation.meal_option:
            # اگر meal_option null است، نمی‌توانیم base_meal را مشخص کنیم
            # این رزروها در by_meal_option نمایش داده می‌شوند اما در by_base_meal نمی‌توانیم آنها را گروه‌بندی کنیم
            continue
        
        if not reservation.meal_option.base_meal:
            continue
        
        base_meal = reservation.meal_option.base_meal
        base_meal_id = base_meal.id
        meal_option_id = reservation.meal_option.id
        
        if base_meal_id not in base_meals_data:
            base_meals_data[base_meal_id] = {
                'base_meal_id': base_meal_id,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.daily_menu.restaurant.name if reservation.meal_option and reservation.meal_option.daily_menu and reservation.meal_option.daily_menu.restaurant else '',
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'meal_options_count': 0,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
            }
            meal_options_by_base_meal[base_meal_id] = {}
        
        data = base_meals_data[base_meal_id]
        quantity = reservation.quantity or 1
        data['total_reservations'] += quantity
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(quantity))
        data['total_amount'] += amount

        if reservation.status == 'reserved':
            data['reserved_count'] += quantity
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += quantity
        elif reservation.status == 'served':
            data['served_count'] += quantity
        
        if meal_option_id not in meal_options_by_base_meal[base_meal_id]:
            meal_options_by_base_meal[base_meal_id][meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': reservation.meal_option.title,
                'base_meal_title': base_meal.title,
                'restaurant_name': reservation.meal_option.daily_menu.restaurant.name if reservation.meal_option and reservation.meal_option.daily_menu and reservation.meal_option.daily_menu.restaurant else '',
                'restaurant_id': reservation.meal_option.daily_menu.restaurant.id if reservation.meal_option and reservation.meal_option.daily_menu and reservation.meal_option.daily_menu.restaurant else None,
                'center_name': ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': reservation.daily_menu.restaurant.centers.first().id if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
            data['meal_options_count'] += 1
        
        meal_option_data = meal_options_by_base_meal[base_meal_id][meal_option_id]
        meal_option_data['total_reservations'] += quantity
        meal_option_data['total_amount'] += amount

        if reservation.status == 'reserved':
            meal_option_data['reserved_count'] += quantity
            meal_option_data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            meal_option_data['cancelled_count'] += quantity
        elif reservation.status == 'served':
            meal_option_data['served_count'] += quantity
            meal_option_data['served_amount'] += amount
    
    # اضافه کردن رزروهای مهمان غذا به گزارش BaseMeal
    for guest_reservation in guest_reservations:
        if not guest_reservation.meal_option:
            # اگر meal_option null است، نمی‌توانیم base_meal را مشخص کنیم
            continue
        
        if not guest_reservation.meal_option.base_meal:
            continue
        
        base_meal = guest_reservation.meal_option.base_meal
        base_meal_id = base_meal.id
        meal_option_id = guest_reservation.meal_option.id
        
        if base_meal_id not in base_meals_data:
            base_meals_data[base_meal_id] = {
                'base_meal_id': base_meal_id,
                'base_meal_title': base_meal.title,
                'restaurant_name': guest_reservation.meal_option.daily_menu.restaurant.name if guest_reservation.meal_option and guest_reservation.meal_option.daily_menu and guest_reservation.meal_option.daily_menu.restaurant else '',
                'center_name': ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else '',
                'meal_options_count': 0,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
            }
            meal_options_by_base_meal[base_meal_id] = {}
        
        data = base_meals_data[base_meal_id]
        data['total_reservations'] += 1
        amount = Decimal(str(guest_reservation.amount or 0))
        data['total_amount'] += amount
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
        
        if meal_option_id not in meal_options_by_base_meal[base_meal_id]:
            meal_options_by_base_meal[base_meal_id][meal_option_id] = {
                'meal_option_id': meal_option_id,
                'meal_option_title': guest_reservation.meal_option.title,
                'base_meal_title': base_meal.title,
                'restaurant_name': guest_reservation.meal_option.daily_menu.restaurant.name if guest_reservation.meal_option and guest_reservation.meal_option.daily_menu and guest_reservation.meal_option.daily_menu.restaurant else '',
                'restaurant_id': guest_reservation.meal_option.daily_menu.restaurant.id if guest_reservation.meal_option and guest_reservation.meal_option.daily_menu and guest_reservation.meal_option.daily_menu.restaurant else None,
                'center_name': ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': guest_reservation.daily_menu.restaurant.centers.first().id if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
            data['meal_options_count'] += 1
        
        meal_option_data = meal_options_by_base_meal[base_meal_id][meal_option_id]
        meal_option_data['total_reservations'] += 1
        meal_option_data['total_amount'] += amount
        
        if guest_reservation.status == 'reserved':
            meal_option_data['reserved_count'] += 1
            meal_option_data['reserved_amount'] += amount
        elif guest_reservation.status == 'cancelled':
            meal_option_data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            meal_option_data['served_count'] += 1
            meal_option_data['served_amount'] += amount
    
    result_base_meals = []
    for base_meal_id, data in base_meals_data.items():
        data['meal_options'] = list(meal_options_by_base_meal[base_meal_id].values())
        result_base_meals.append(data)
    
    # گزارش بر اساس DessertOption
    dessert_options_data = {}
    for dessert_reservation in dessert_reservations:
        if not dessert_reservation.dessert_option:
            continue
        
        dessert_option = dessert_reservation.dessert_option
        dessert_option_id = dessert_option.id
        
        if dessert_option_id not in dessert_options_data:
            dessert_options_data[dessert_option_id] = {
                'dessert_option_id': dessert_option_id,
                'dessert_option_title': dessert_option.title,
                'base_dessert_title': dessert_option.base_dessert.title if dessert_option.base_dessert else '',
                'restaurant_name': dessert_option.daily_menu.restaurant.name if dessert_option.daily_menu and dessert_option.daily_menu.restaurant else '',
                'restaurant_id': dessert_option.daily_menu.restaurant.id if dessert_option.daily_menu and dessert_option.daily_menu.restaurant else None,
                'center_name': ', '.join([c.name for c in dessert_reservation.daily_menu.restaurant.centers.all()]) if dessert_reservation.daily_menu and dessert_reservation.daily_menu.restaurant and dessert_reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': dessert_reservation.daily_menu.restaurant.centers.first().id if dessert_reservation.daily_menu and dessert_reservation.daily_menu.restaurant and dessert_reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
        
        data = dessert_options_data[dessert_option_id]
        data['total_reservations'] += dessert_reservation.quantity or 1
        amount = Decimal(str(dessert_reservation.amount or 0)) * Decimal(str(dessert_reservation.quantity or 1))
        data['total_amount'] += amount
        
        if dessert_reservation.status == 'reserved':
            data['reserved_count'] += dessert_reservation.quantity or 1
            data['reserved_amount'] += amount
        elif dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += dessert_reservation.quantity or 1
        elif dessert_reservation.status == 'served':
            data['served_count'] += dessert_reservation.quantity or 1
            data['served_amount'] += amount
    
    # اضافه کردن رزروهای مهمان دسر
    for guest_dessert_reservation in guest_dessert_reservations:
        if not guest_dessert_reservation.dessert_option:
            continue
        
        dessert_option = guest_dessert_reservation.dessert_option
        dessert_option_id = dessert_option.id
        
        if dessert_option_id not in dessert_options_data:
            dessert_options_data[dessert_option_id] = {
                'dessert_option_id': dessert_option_id,
                'dessert_option_title': dessert_option.title,
                'base_dessert_title': dessert_option.base_dessert.title if dessert_option.base_dessert else '',
                'restaurant_name': dessert_option.daily_menu.restaurant.name if dessert_option.daily_menu and dessert_option.daily_menu.restaurant else '',
                'restaurant_id': dessert_option.daily_menu.restaurant.id if dessert_option.daily_menu and dessert_option.daily_menu.restaurant else None,
                'center_name': ', '.join([c.name for c in guest_dessert_reservation.daily_menu.restaurant.centers.all()]) if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.restaurant and guest_dessert_reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': guest_dessert_reservation.daily_menu.restaurant.centers.first().id if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.restaurant and guest_dessert_reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
        
        data = dessert_options_data[dessert_option_id]
        data['total_reservations'] += 1
        amount = Decimal(str(guest_dessert_reservation.amount or 0))
        data['total_amount'] += amount
        
        if guest_dessert_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_dessert_reservation.status == 'served':
            data['served_count'] += 1
            data['served_amount'] += amount
    
    # گزارش بر اساس BaseDessert
    base_desserts_data = {}
    dessert_options_by_base_dessert = {}
    
    for dessert_reservation in dessert_reservations:
        if not dessert_reservation.dessert_option or not dessert_reservation.dessert_option.base_dessert:
            continue
        
        base_dessert = dessert_reservation.dessert_option.base_dessert
        base_dessert_id = base_dessert.id
        dessert_option_id = dessert_reservation.dessert_option.id
        
        if base_dessert_id not in base_desserts_data:
            base_desserts_data[base_dessert_id] = {
                'base_dessert_id': base_dessert_id,
                'base_dessert_title': base_dessert.title,
                'restaurant_name': dessert_reservation.dessert_option.daily_menu.restaurant.name if dessert_reservation.dessert_option and dessert_reservation.dessert_option.daily_menu and dessert_reservation.dessert_option.daily_menu.restaurant else '',
                'center_name': ', '.join([c.name for c in dessert_reservation.daily_menu.restaurant.centers.all()]) if dessert_reservation.daily_menu and dessert_reservation.daily_menu.restaurant and dessert_reservation.daily_menu.restaurant.centers.exists() else '',
                'dessert_options_count': 0,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
            }
            dessert_options_by_base_dessert[base_dessert_id] = {}
        
        data = base_desserts_data[base_dessert_id]
        data['total_reservations'] += dessert_reservation.quantity or 1
        amount = Decimal(str(dessert_reservation.amount or 0)) * Decimal(str(dessert_reservation.quantity or 1))
        data['total_amount'] += amount
        
        if dessert_reservation.status == 'reserved':
            data['reserved_count'] += dessert_reservation.quantity or 1
        elif dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += dessert_reservation.quantity or 1
        elif dessert_reservation.status == 'served':
            data['served_count'] += dessert_reservation.quantity or 1
        
        if dessert_option_id not in dessert_options_by_base_dessert[base_dessert_id]:
            dessert_options_by_base_dessert[base_dessert_id][dessert_option_id] = {
                'dessert_option_id': dessert_option_id,
                'dessert_option_title': dessert_reservation.dessert_option.title,
                'base_dessert_title': base_dessert.title,
                'restaurant_name': dessert_reservation.dessert_option.daily_menu.restaurant.name if dessert_reservation.dessert_option and dessert_reservation.dessert_option.daily_menu and dessert_reservation.dessert_option.daily_menu.restaurant else '',
                'restaurant_id': dessert_reservation.dessert_option.daily_menu.restaurant.id if dessert_reservation.dessert_option and dessert_reservation.dessert_option.daily_menu and dessert_reservation.dessert_option.daily_menu.restaurant else None,
                'center_name': ', '.join([c.name for c in dessert_reservation.daily_menu.restaurant.centers.all()]) if dessert_reservation.daily_menu and dessert_reservation.daily_menu.restaurant and dessert_reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': dessert_reservation.daily_menu.restaurant.centers.first().id if dessert_reservation.daily_menu and dessert_reservation.daily_menu.restaurant and dessert_reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
            data['dessert_options_count'] += 1
        
        dessert_option_data = dessert_options_by_base_dessert[base_dessert_id][dessert_option_id]
        dessert_option_data['total_reservations'] += dessert_reservation.quantity or 1
        dessert_option_data['total_amount'] += amount
        
        if dessert_reservation.status == 'reserved':
            dessert_option_data['reserved_count'] += dessert_reservation.quantity or 1
            dessert_option_data['reserved_amount'] += amount
        elif dessert_reservation.status == 'cancelled':
            dessert_option_data['cancelled_count'] += dessert_reservation.quantity or 1
        elif dessert_reservation.status == 'served':
            dessert_option_data['served_count'] += dessert_reservation.quantity or 1
            dessert_option_data['served_amount'] += amount
    
    # اضافه کردن رزروهای مهمان دسر به BaseDessert
    for guest_dessert_reservation in guest_dessert_reservations:
        if not guest_dessert_reservation.dessert_option or not guest_dessert_reservation.dessert_option.base_dessert:
            continue
        
        base_dessert = guest_dessert_reservation.dessert_option.base_dessert
        base_dessert_id = base_dessert.id
        dessert_option_id = guest_dessert_reservation.dessert_option.id
        
        if base_dessert_id not in base_desserts_data:
            base_desserts_data[base_dessert_id] = {
                'base_dessert_id': base_dessert_id,
                'base_dessert_title': base_dessert.title,
                'restaurant_name': guest_dessert_reservation.dessert_option.daily_menu.restaurant.name if guest_dessert_reservation.dessert_option and guest_dessert_reservation.dessert_option.daily_menu and guest_dessert_reservation.dessert_option.daily_menu.restaurant else '',
                'center_name': ', '.join([c.name for c in guest_dessert_reservation.daily_menu.restaurant.centers.all()]) if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.restaurant and guest_dessert_reservation.daily_menu.restaurant.centers.exists() else '',
                'dessert_options_count': 0,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
            }
            dessert_options_by_base_dessert[base_dessert_id] = {}
        
        data = base_desserts_data[base_dessert_id]
        data['total_reservations'] += 1
        amount = Decimal(str(guest_dessert_reservation.amount or 0))
        data['total_amount'] += amount
        
        if guest_dessert_reservation.status == 'reserved':
            data['reserved_count'] += 1
        elif guest_dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_dessert_reservation.status == 'served':
            data['served_count'] += 1
        
        if dessert_option_id not in dessert_options_by_base_dessert[base_dessert_id]:
            dessert_options_by_base_dessert[base_dessert_id][dessert_option_id] = {
                'dessert_option_id': dessert_option_id,
                'dessert_option_title': guest_dessert_reservation.dessert_option.title,
                'base_dessert_title': base_dessert.title,
                'restaurant_name': guest_dessert_reservation.dessert_option.daily_menu.restaurant.name if guest_dessert_reservation.dessert_option and guest_dessert_reservation.dessert_option.daily_menu and guest_dessert_reservation.dessert_option.daily_menu.restaurant else '',
                'restaurant_id': guest_dessert_reservation.dessert_option.daily_menu.restaurant.id if guest_dessert_reservation.dessert_option and guest_dessert_reservation.dessert_option.daily_menu and guest_dessert_reservation.dessert_option.daily_menu.restaurant else None,
                'center_name': ', '.join([c.name for c in guest_dessert_reservation.daily_menu.restaurant.centers.all()]) if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.restaurant and guest_dessert_reservation.daily_menu.restaurant.centers.exists() else '',
                'center_id': guest_dessert_reservation.daily_menu.restaurant.centers.first().id if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.restaurant and guest_dessert_reservation.daily_menu.restaurant.centers.exists() else None,
                'total_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'served_amount': Decimal('0'),
            }
            data['dessert_options_count'] += 1
        
        dessert_option_data = dessert_options_by_base_dessert[base_dessert_id][dessert_option_id]
        dessert_option_data['total_reservations'] += 1
        dessert_option_data['total_amount'] += amount
        
        if guest_dessert_reservation.status == 'reserved':
            dessert_option_data['reserved_count'] += 1
            dessert_option_data['reserved_amount'] += amount
        elif guest_dessert_reservation.status == 'cancelled':
            dessert_option_data['cancelled_count'] += 1
        elif guest_dessert_reservation.status == 'served':
            dessert_option_data['served_count'] += 1
            dessert_option_data['served_amount'] += amount
    
    result_base_desserts = []
    for base_dessert_id, data in base_desserts_data.items():
        data['dessert_options'] = list(dessert_options_by_base_dessert[base_dessert_id].values())
        result_base_desserts.append(data)
    
    # گزارش بر اساس کاربر
    users_data = {}
    
    for reservation in reservations:
        user_id = reservation.user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': reservation.user.username,
                'full_name': reservation.user.get_full_name(),
                'employee_number': reservation.user.employee_number or '',
                'center_name': ', '.join([c.name for c in reservation.user.centers.all()]) if reservation.user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
            }
        
        data = users_data[user_id]
        quantity = reservation.quantity or 1
        data['total_reservations'] += quantity
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(quantity))
        data['total_amount'] += amount

        if reservation.status == 'reserved':
            data['reserved_count'] += quantity
            data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += quantity
        elif reservation.status == 'served':
            data['served_count'] += quantity
    
    for guest_reservation in guest_reservations:
        user_id = guest_reservation.host_user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': guest_reservation.host_user.username,
                'full_name': guest_reservation.host_user.get_full_name(),
                'employee_number': guest_reservation.host_user.employee_number or '',
                'center_name': ', '.join([c.name for c in guest_reservation.host_user.centers.all()]) if guest_reservation.host_user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
            }
        
        data = users_data[user_id]
        data['total_guest_reservations'] += 1
        amount = Decimal(str(guest_reservation.amount or 0))
        data['total_amount'] += amount
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    # اضافه کردن رزروهای دسر به گزارش کاربر
    for dessert_reservation in dessert_reservations:
        user_id = dessert_reservation.user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': dessert_reservation.user.username,
                'full_name': dessert_reservation.user.get_full_name(),
                'employee_number': dessert_reservation.user.employee_number or '',
                'center_name': ', '.join([c.name for c in dessert_reservation.user.centers.all()]) if dessert_reservation.user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
            }
        
        data = users_data[user_id]
        data['total_reservations'] += dessert_reservation.quantity or 1
        amount = Decimal(str(dessert_reservation.amount or 0)) * Decimal(str(dessert_reservation.quantity or 1))
        data['total_amount'] += amount
        
        if dessert_reservation.status == 'reserved':
            data['reserved_count'] += dessert_reservation.quantity or 1
            data['reserved_amount'] += amount
        elif dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += dessert_reservation.quantity or 1
        elif dessert_reservation.status == 'served':
            data['served_count'] += dessert_reservation.quantity or 1
    
    # اضافه کردن رزروهای مهمان دسر به گزارش کاربر
    for guest_dessert_reservation in guest_dessert_reservations:
        user_id = guest_dessert_reservation.host_user.id
        
        if user_id not in users_data:
            users_data[user_id] = {
                'user_id': user_id,
                'username': guest_dessert_reservation.host_user.username,
                'full_name': guest_dessert_reservation.host_user.get_full_name(),
                'employee_number': guest_dessert_reservation.host_user.employee_number or '',
                'center_name': ', '.join([c.name for c in guest_dessert_reservation.host_user.centers.all()]) if guest_dessert_reservation.host_user.centers.exists() else '',
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
            }
        
        data = users_data[user_id]
        data['total_guest_reservations'] += 1
        amount = Decimal(str(guest_dessert_reservation.amount or 0))
        data['total_amount'] += amount
        
        if guest_dessert_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_dessert_reservation.status == 'served':
            data['served_count'] += 1
    
    # گزارش بر اساس تاریخ
    dates_data = {}
    
    for reservation in reservations:
        if not reservation.daily_menu:
            continue
        
        date = reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': jdatetime.date.fromgregorian(date=date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'centers': {},
            }
        
        data = dates_data[date]
        quantity = reservation.quantity or 1
        data['total_reservations'] += quantity
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(quantity))
        data['total_amount'] += amount

        center_name = ', '.join([c.name for c in reservation.daily_menu.restaurant.centers.all()]) if reservation.daily_menu and reservation.daily_menu.restaurant and reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': Decimal('0'),
            }

        data['centers'][center_name]['total_reservations'] += quantity
        data['centers'][center_name]['total_amount'] += amount

        if reservation.status == 'reserved':
            data['reserved_count'] += quantity
            data['reserved_amount'] += amount
        elif reservation.status == 'cancelled':
            data['cancelled_count'] += quantity
        elif reservation.status == 'served':
            data['served_count'] += quantity
    
    for guest_reservation in guest_reservations:
        if not guest_reservation.daily_menu:
            continue
        
        date = guest_reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': jdatetime.date.fromgregorian(date=date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_guest_reservations'] += 1
        amount = Decimal(str(guest_reservation.amount or 0))
        data['total_amount'] += amount
        
        center_name = ', '.join([c.name for c in guest_reservation.daily_menu.restaurant.centers.all()]) if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant and guest_reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': Decimal('0'),
            }
        
        if guest_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_reservation.status == 'served':
            data['served_count'] += 1
    
    # اضافه کردن رزروهای دسر به گزارش تاریخ
    for dessert_reservation in dessert_reservations:
        if not dessert_reservation.daily_menu:
            continue
        
        date = dessert_reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': jdatetime.date.fromgregorian(date=date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_reservations'] += dessert_reservation.quantity or 1
        amount = Decimal(str(dessert_reservation.amount or 0)) * Decimal(str(dessert_reservation.quantity or 1))
        data['total_amount'] += amount
        
        center_name = ', '.join([c.name for c in dessert_reservation.daily_menu.restaurant.centers.all()]) if dessert_reservation.daily_menu and dessert_reservation.daily_menu.restaurant and dessert_reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': Decimal('0'),
            }
        
        data['centers'][center_name]['total_reservations'] += dessert_reservation.quantity or 1
        data['centers'][center_name]['total_amount'] += amount
        
        if dessert_reservation.status == 'reserved':
            data['reserved_count'] += dessert_reservation.quantity or 1
            data['reserved_amount'] += amount
        elif dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += dessert_reservation.quantity or 1
        elif dessert_reservation.status == 'served':
            data['served_count'] += dessert_reservation.quantity or 1
    
    # اضافه کردن رزروهای مهمان دسر به گزارش تاریخ
    for guest_dessert_reservation in guest_dessert_reservations:
        if not guest_dessert_reservation.daily_menu:
            continue
        
        date = guest_dessert_reservation.daily_menu.date
        
        if date not in dates_data:
            dates_data[date] = {
                'date': date,
                'jalali_date': jdatetime.date.fromgregorian(date=date).strftime('%Y/%m/%d'),
                'total_reservations': 0,
                'total_guest_reservations': 0,
                'reserved_count': 0,
                'cancelled_count': 0,
                'served_count': 0,
                'total_amount': Decimal('0'),
                'reserved_amount': Decimal('0'),
                'centers': {},
            }
        
        data = dates_data[date]
        data['total_guest_reservations'] += 1
        amount = Decimal(str(guest_dessert_reservation.amount or 0))
        data['total_amount'] += amount
        
        center_name = ', '.join([c.name for c in guest_dessert_reservation.daily_menu.restaurant.centers.all()]) if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.restaurant and guest_dessert_reservation.daily_menu.restaurant.centers.exists() else 'نامشخص'
        if center_name not in data['centers']:
            data['centers'][center_name] = {
                'name': center_name,
                'total_reservations': 0,
                'total_amount': Decimal('0'),
            }
        
        if guest_dessert_reservation.status == 'reserved':
            data['reserved_count'] += 1
            data['reserved_amount'] += amount
        elif guest_dessert_reservation.status == 'cancelled':
            data['cancelled_count'] += 1
        elif guest_dessert_reservation.status == 'served':
            data['served_count'] += 1
    
    result_dates = []
    for date in sorted(dates_data.keys(), reverse=True):
        data = dates_data[date]
        data['centers'] = list(data['centers'].values())
        result_dates.append(data)
    
    # ساخت گزارش جامع
    report_data = {
        'total_reservations': total_reservations,
        'total_guest_reservations': total_guest_reservations,
        'total_dessert_reservations': total_dessert_reservations,
        'total_guest_dessert_reservations': total_guest_dessert_reservations,
        'total_amount': total_amount,
        'reserved_amount': reserved_amount,
        'served_amount': served_amount,
        'cancelled_amount': cancelled_amount,
        'by_meal_option': list(meal_options_data.values()),
        'by_base_meal': result_base_meals,
        'by_dessert_option': list(dessert_options_data.values()),
        'by_base_dessert': result_base_desserts,
        'by_user': list(users_data.values()),
        'by_date': result_dates,
    }
    
    serializer = ComprehensiveReportSerializer(report_data)
    return Response(serializer.data)


@extend_schema(
    summary='Detailed Reservations Report',
    description='Complete detailed report of reservations with all information',
    tags=['Reports'],
    parameters=[
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'user_id',
            'in': 'query',
            'description': 'فیلتر بر اساس کاربر',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'status',
            'in': 'query',
            'description': 'فیلتر بر اساس وضعیت (reserved, cancelled, served)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ]
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def detailed_reservations_report(request):
    """گزارش جزئیات رزروها"""
    user = request.user
    
    # تعیین مراکز قابل دسترسی
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # بررسی دسترسی - فقط System Admin و Food Admin
    if user.role not in ['sys_admin', 'admin_food']:
        return Response({
            'error': 'دسترسی غیرمجاز'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    user_id = request.query_params.get('user_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    status_filter = request.query_params.get('status')
    
    # فیلتر رزروها - حذف رزروهای کنسل شده
    reservations = FoodReservation.objects.select_related(
        'user', 'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').all().exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
    # فیلتر بر اساس مرکز
    if center_id:
        # بررسی دسترسی به مرکز درخواستی
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = filter_reservations_by_center(reservations, center_id=center_id)
    elif accessible_centers is not None:
        # فیلتر بر اساس مراکز قابل دسترسی
        reservations = filter_reservations_by_center(reservations, accessible_centers=accessible_centers)
    
    if user_id:
        reservations = reservations.filter(user_id=user_id)
    
    # فیلتر بر اساس تاریخ
    reservations = filter_reservations_by_date(reservations, start_date=start_date, end_date=end_date)
    
    if status_filter:
        reservations = reservations.filter(status=status_filter)
    
    serializer = DetailedReservationReportSerializer(reservations, many=True, context={'request': request})
    return Response(serializer.data)


@extend_schema(
    summary='User Reservations Report by Date Range',
    description='Get detailed report of a specific user\'s reservations within a date range. If user_id is not provided, returns report for the authenticated user.',
    tags=['Reports'],
    parameters=[
        {
            'name': 'user_id',
            'in': 'query',
            'description': 'شناسه کاربر (اگر ارسال نشود، گزارش برای کاربر لاگین شده نمایش داده می‌شود)',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD) - الزامی',
            'required': True,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD) - الزامی',
            'required': True,
            'schema': {'type': 'string'}
        },
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'},
        404: {'description': 'User not found'}
    }
)
@api_view(['GET'])
@permission_classes([UserReportPermission])
def user_reservations_by_date_range(request):
    """گزارش رزروهای یک کاربر در بازه تاریخ"""
    current_user = request.user
    
    # دریافت user_id - اگر ارسال نشده باشد، از کاربر لاگین شده استفاده می‌شود
    user_id_param = request.query_params.get('user_id')
    if user_id_param:
        try:
            target_user_id = int(user_id_param)
        except (ValueError, TypeError):
            return Response({
                'error': 'شناسه کاربر باید عدد باشد'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # بررسی دسترسی - فقط System Admin و Food Admin می‌توانند گزارش کاربران دیگر را ببینند
        if current_user.role not in ['sys_admin', 'admin_food']:
            return Response({
                'error': 'شما فقط می‌توانید گزارش خودتان را مشاهده کنید'
            }, status=status.HTTP_403_FORBIDDEN)
        
        from apps.accounts.models import User
        try:
            target_user = User.objects.get(id=target_user_id)
        except User.DoesNotExist:
            return Response({
                'error': 'کاربر یافت نشد'
            }, status=status.HTTP_404_NOT_FOUND)
    else:
        # استفاده از کاربر لاگین شده
        target_user = current_user
        target_user_id = current_user.id
    
    # دریافت تاریخ‌ها
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    if not start_date or not end_date:
        return Response({
            'error': 'تاریخ شروع و پایان الزامی است'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    parsed_start_date = parse_date_filter(start_date)
    parsed_end_date = parse_date_filter(end_date)
    
    if not parsed_start_date or not parsed_end_date:
        return Response({
            'error': 'فرمت تاریخ نامعتبر است. از فرمت میلادی (2025-10-24) یا شمسی (1404/08/02) استفاده کنید'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if parsed_start_date > parsed_end_date:
        return Response({
            'error': 'تاریخ شروع باید قبل از تاریخ پایان باشد'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # دریافت فیلتر مرکز (اختیاری)
    center_id = request.query_params.get('center_id')
    
    # فیلتر رزروهای غذا - ابتدا بر اساس کاربر فیلتر می‌کنیم - حذف رزروهای کنسل شده
    reservations = FoodReservation.objects.select_related(
        'user', 'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(
        user_id=target_user_id
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    # فیلتر بر اساس تاریخ با استفاده از تابع helper
    reservations = filter_reservations_by_date(reservations, start_date=parsed_start_date, end_date=parsed_end_date)
    
    # فیلتر رزروهای مهمان - حذف رزروهای کنسل شده
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(
        host_user_id=target_user_id
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    # فیلتر بر اساس تاریخ با استفاده از تابع helper
    guest_reservations = filter_guest_reservations_by_date(guest_reservations, start_date=parsed_start_date, end_date=parsed_end_date)
    
    # فیلتر رزروهای دسر - حذف رزروهای کنسل شده
    dessert_reservations = DessertReservation.objects.select_related(
        'user', 'dessert_option', 'dessert_option__daily_menu',
        'dessert_option__daily_menu__restaurant', 'daily_menu', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(
        user_id=target_user_id
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    # فیلتر بر اساس تاریخ - دسرها هم مثل غذاها هستند
    dessert_reservations = dessert_reservations.filter(
        Q(daily_menu__isnull=True) | 
        (Q(daily_menu__date__gte=parsed_start_date) & Q(daily_menu__date__lte=parsed_end_date))
    )
    
    # فیلتر رزروهای دسر مهمان - حذف رزروهای کنسل شده
    guest_dessert_reservations = GuestDessertReservation.objects.select_related(
        'host_user', 'dessert_option', 'dessert_option__daily_menu',
        'dessert_option__daily_menu__restaurant', 'daily_menu', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers').filter(
        host_user_id=target_user_id
    ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    # فیلتر بر اساس تاریخ
    guest_dessert_reservations = guest_dessert_reservations.filter(
        Q(daily_menu__isnull=True) | 
        (Q(daily_menu__date__gte=parsed_start_date) & Q(daily_menu__date__lte=parsed_end_date))
    )
    
    # فیلتر بر اساس مرکز (فقط اگر center_id ارسال شده باشد)
    if center_id:
        try:
            center_id = int(center_id)
        except (ValueError, TypeError):
            return Response({
                'error': 'شناسه مرکز باید عدد باشد'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # فیلتر بر اساس مرکز
        reservations = reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
        guest_reservations = guest_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
        dessert_reservations = dessert_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
        guest_dessert_reservations = guest_dessert_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
    
    # محاسبه آمار
    total_reservations = reservations.count()
    total_guest_reservations = guest_reservations.count()
    total_dessert_reservations = dessert_reservations.count()
    total_guest_dessert_reservations = guest_dessert_reservations.count()
    
    reserved_count = reservations.filter(status='reserved').count() + guest_reservations.filter(status='reserved').count() + dessert_reservations.filter(status='reserved').count() + guest_dessert_reservations.filter(status='reserved').count()
    cancelled_count = reservations.filter(status='cancelled').count() + guest_reservations.filter(status='cancelled').count() + dessert_reservations.filter(status='cancelled').count() + guest_dessert_reservations.filter(status='cancelled').count()
    served_count = reservations.filter(status='served').count() + guest_reservations.filter(status='served').count() + dessert_reservations.filter(status='served').count() + guest_dessert_reservations.filter(status='served').count()
    
    # محاسبه مبالغ
    total_amount = Decimal('0')
    reserved_amount = Decimal('0')
    
    for reservation in reservations:
        amount = Decimal(str(reservation.amount or 0)) * Decimal(str(reservation.quantity or 1))
        total_amount += amount
        if reservation.status == 'reserved':
            reserved_amount += amount
    
    for guest_reservation in guest_reservations:
        amount = Decimal(str(guest_reservation.amount or 0))
        total_amount += amount
        if guest_reservation.status == 'reserved':
            reserved_amount += amount
    
    for dessert_reservation in dessert_reservations:
        amount = Decimal(str(dessert_reservation.amount or 0)) * Decimal(str(dessert_reservation.quantity or 1))
        total_amount += amount
        if dessert_reservation.status == 'reserved':
            reserved_amount += amount
    
    for guest_dessert_reservation in guest_dessert_reservations:
        amount = Decimal(str(guest_dessert_reservation.amount or 0))
        total_amount += amount
        if guest_dessert_reservation.status == 'reserved':
            reserved_amount += amount
    
    # آماده‌سازی داده‌های رزروهای غذا با اطلاعات رستوران و مرکز - فقط رزروهای غیر کنسل شده
    reservations_data = []
    for reservation in reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if reservation.status == 'cancelled':
            continue
        
        # دریافت اطلاعات رستوران
        restaurant = None
        restaurant_name = None
        restaurant_id = None
        centers = []
        
        if reservation.daily_menu and reservation.daily_menu.restaurant:
            restaurant = reservation.daily_menu.restaurant
            restaurant_name = restaurant.name
            restaurant_id = restaurant.id
            centers = [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
        elif reservation.daily_menu_info:
            # اگر daily_menu حذف شده، از daily_menu_info استفاده می‌کنیم
            restaurant_name = "رستوران حذف شده"
            # سعی می‌کنیم اطلاعات مرکز را از daily_menu_info استخراج کنیم
            if 'مرکز:' in reservation.daily_menu_info:
                center_part = reservation.daily_menu_info.split('مرکز:')[1].split('-')[0].strip()
                centers = [{'id': None, 'name': center_part}]
        
        # دریافت اطلاعات meal_option
        meal_option_data = None
        if reservation.meal_option:
            meal_option_data = {
                'id': reservation.meal_option.id,
                'title': reservation.meal_option.title,
                'description': reservation.meal_option.description or '',
                'price': float(reservation.meal_option.price),
                'base_meal_title': reservation.meal_option.base_meal.title if reservation.meal_option.base_meal else ''
            }
        elif reservation.meal_option_info:
            # اگر meal_option حذف شده، از meal_option_info استفاده می‌کنیم
            meal_option_data = {
                'id': None,
                'title': reservation.meal_option_info,
                'description': '',
                'price': 0,
                'base_meal_title': '',
                'is_deleted': True
            }
        
        # دریافت تاریخ
        date = None
        jalali_date = None
        if reservation.daily_menu and reservation.daily_menu.date:
            date = reservation.daily_menu.date.isoformat()
            jalali_date = jdatetime.date.fromgregorian(date=reservation.daily_menu.date).strftime('%Y/%m/%d')
        elif reservation.daily_menu_info and 'تاریخ:' in reservation.daily_menu_info:
            # استخراج تاریخ از daily_menu_info
            try:
                date_part = reservation.daily_menu_info.split('تاریخ:')[1].strip()
                date = date_part
                # تبدیل به jalali اگر ممکن باشد
                jalali_date = date_part
            except:
                pass
        
        reservations_data.append({
            'id': reservation.id,
            'type': 'food',
            'restaurant': {
                'id': restaurant_id,
                'name': restaurant_name,
                'centers': centers,
                'is_deleted': restaurant is None and reservation.daily_menu_info is not None
            } if restaurant_name else None,
            'daily_menu_info': reservation.daily_menu_info if reservation.daily_menu_info else None,
            'meal_option': meal_option_data,
            'meal_option_info': reservation.meal_option_info if reservation.meal_option_info else None,
            'quantity': reservation.quantity,
            'amount': float(reservation.amount or 0),
            'status': reservation.status,
            'reservation_date': reservation.reservation_date.isoformat() if reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if reservation.reservation_date else None,
            'date': date,
            'jalali_date': jalali_date,
        })
    
    # آماده‌سازی داده‌های رزروهای مهمان - فقط رزروهای غیر کنسل شده
    guest_reservations_data = []
    for guest_reservation in guest_reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if guest_reservation.status == 'cancelled':
            continue
        
        # دریافت اطلاعات رستوران
        restaurant = None
        restaurant_name = None
        restaurant_id = None
        centers = []
        
        if guest_reservation.daily_menu and guest_reservation.daily_menu.restaurant:
            restaurant = guest_reservation.daily_menu.restaurant
            restaurant_name = restaurant.name
            restaurant_id = restaurant.id
            centers = [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
        elif guest_reservation.daily_menu_info:
            restaurant_name = "رستوران حذف شده"
            if 'مرکز:' in guest_reservation.daily_menu_info:
                center_part = guest_reservation.daily_menu_info.split('مرکز:')[1].split('-')[0].strip()
                centers = [{'id': None, 'name': center_part}]
        
        # دریافت اطلاعات meal_option
        meal_option_data = None
        if guest_reservation.meal_option:
            meal_option_data = {
                'id': guest_reservation.meal_option.id,
                'title': guest_reservation.meal_option.title,
                'description': guest_reservation.meal_option.description or '',
                'price': float(guest_reservation.meal_option.price),
                'base_meal_title': guest_reservation.meal_option.base_meal.title if guest_reservation.meal_option.base_meal else ''
            }
        elif guest_reservation.meal_option_info:
            meal_option_data = {
                'id': None,
                'title': guest_reservation.meal_option_info,
                'description': '',
                'price': 0,
                'base_meal_title': '',
                'is_deleted': True
            }
        
        # دریافت تاریخ
        date = None
        jalali_date = None
        if guest_reservation.daily_menu and guest_reservation.daily_menu.date:
            date = guest_reservation.daily_menu.date.isoformat()
            jalali_date = jdatetime.date.fromgregorian(date=guest_reservation.daily_menu.date).strftime('%Y/%m/%d')
        elif guest_reservation.daily_menu_info and 'تاریخ:' in guest_reservation.daily_menu_info:
            try:
                date_part = guest_reservation.daily_menu_info.split('تاریخ:')[1].strip()
                date = date_part
                jalali_date = date_part
            except:
                pass
        
        guest_reservations_data.append({
            'id': guest_reservation.id,
            'type': 'guest_food',
            'guest_first_name': guest_reservation.guest_first_name,
            'guest_last_name': guest_reservation.guest_last_name,
            'guest_name': f"{guest_reservation.guest_first_name} {guest_reservation.guest_last_name}".strip(),
            'restaurant': {
                'id': restaurant_id,
                'name': restaurant_name,
                'centers': centers,
                'is_deleted': restaurant is None and guest_reservation.daily_menu_info is not None
            } if restaurant_name else None,
            'daily_menu_info': guest_reservation.daily_menu_info if guest_reservation.daily_menu_info else None,
            'meal_option': meal_option_data,
            'meal_option_info': guest_reservation.meal_option_info if guest_reservation.meal_option_info else None,
            'amount': float(guest_reservation.amount or 0),
            'status': guest_reservation.status,
            'reservation_date': guest_reservation.reservation_date.isoformat() if guest_reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=guest_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if guest_reservation.reservation_date else None,
            'date': date,
            'jalali_date': jalali_date,
        })
    
    # آماده‌سازی داده‌های رزروهای دسر - فقط رزروهای غیر کنسل شده
    dessert_reservations_data = []
    for dessert_reservation in dessert_reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if dessert_reservation.status == 'cancelled':
            continue
        
        # دریافت اطلاعات رستوران
        restaurant = None
        restaurant_name = None
        restaurant_id = None
        centers = []
        
        if dessert_reservation.daily_menu and dessert_reservation.daily_menu.restaurant:
            restaurant = dessert_reservation.daily_menu.restaurant
            restaurant_name = restaurant.name
            restaurant_id = restaurant.id
            centers = [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
        elif dessert_reservation.daily_menu_info:
            restaurant_name = "رستوران حذف شده"
            if 'مرکز:' in dessert_reservation.daily_menu_info:
                center_part = dessert_reservation.daily_menu_info.split('مرکز:')[1].split('-')[0].strip()
                centers = [{'id': None, 'name': center_part}]
        
        # دریافت اطلاعات dessert_option
        dessert_option_data = None
        if dessert_reservation.dessert_option:
            dessert_option_data = {
                'id': dessert_reservation.dessert_option.id,
                'title': dessert_reservation.dessert_option.title,
                'description': dessert_reservation.dessert_option.description or '',
                'price': float(dessert_reservation.dessert_option.price)
            }
        elif dessert_reservation.dessert_option_info:
            dessert_option_data = {
                'id': None,
                'title': dessert_reservation.dessert_option_info,
                'description': '',
                'price': 0,
                'is_deleted': True
            }
        
        # دریافت تاریخ
        date = None
        jalali_date = None
        if dessert_reservation.daily_menu and dessert_reservation.daily_menu.date:
            date = dessert_reservation.daily_menu.date.isoformat()
            jalali_date = jdatetime.date.fromgregorian(date=dessert_reservation.daily_menu.date).strftime('%Y/%m/%d')
        elif dessert_reservation.daily_menu_info and 'تاریخ:' in dessert_reservation.daily_menu_info:
            try:
                date_part = dessert_reservation.daily_menu_info.split('تاریخ:')[1].strip()
                date = date_part
                jalali_date = date_part
            except:
                pass
        
        dessert_reservations_data.append({
            'id': dessert_reservation.id,
            'type': 'dessert',
            'restaurant': {
                'id': restaurant_id,
                'name': restaurant_name,
                'centers': centers,
                'is_deleted': restaurant is None and dessert_reservation.daily_menu_info is not None
            } if restaurant_name else None,
            'daily_menu_info': dessert_reservation.daily_menu_info if dessert_reservation.daily_menu_info else None,
            'dessert_option': dessert_option_data,
            'dessert_option_info': dessert_reservation.dessert_option_info if dessert_reservation.dessert_option_info else None,
            'quantity': dessert_reservation.quantity,
            'amount': float(dessert_reservation.amount or 0),
            'status': dessert_reservation.status,
            'reservation_date': dessert_reservation.reservation_date.isoformat() if dessert_reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=dessert_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if dessert_reservation.reservation_date else None,
            'date': date,
            'jalali_date': jalali_date,
        })
    
    # آماده‌سازی داده‌های رزروهای دسر مهمان - فقط رزروهای غیر کنسل شده
    guest_dessert_reservations_data = []
    for guest_dessert_reservation in guest_dessert_reservations:
        # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
        if guest_dessert_reservation.status == 'cancelled':
            continue
        
        # دریافت اطلاعات رستوران
        restaurant = None
        restaurant_name = None
        restaurant_id = None
        centers = []
        
        if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.restaurant:
            restaurant = guest_dessert_reservation.daily_menu.restaurant
            restaurant_name = restaurant.name
            restaurant_id = restaurant.id
            centers = [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
        elif guest_dessert_reservation.daily_menu_info:
            restaurant_name = "رستوران حذف شده"
            if 'مرکز:' in guest_dessert_reservation.daily_menu_info:
                center_part = guest_dessert_reservation.daily_menu_info.split('مرکز:')[1].split('-')[0].strip()
                centers = [{'id': None, 'name': center_part}]
        
        # دریافت اطلاعات dessert_option
        dessert_option_data = None
        if guest_dessert_reservation.dessert_option:
            dessert_option_data = {
                'id': guest_dessert_reservation.dessert_option.id,
                'title': guest_dessert_reservation.dessert_option.title,
                'description': guest_dessert_reservation.dessert_option.description or '',
                'price': float(guest_dessert_reservation.dessert_option.price)
            }
        elif guest_dessert_reservation.dessert_option_info:
            dessert_option_data = {
                'id': None,
                'title': guest_dessert_reservation.dessert_option_info,
                'description': '',
                'price': 0,
                'is_deleted': True
            }
        
        # دریافت تاریخ
        date = None
        jalali_date = None
        if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.date:
            date = guest_dessert_reservation.daily_menu.date.isoformat()
            jalali_date = jdatetime.date.fromgregorian(date=guest_dessert_reservation.daily_menu.date).strftime('%Y/%m/%d')
        elif guest_dessert_reservation.daily_menu_info and 'تاریخ:' in guest_dessert_reservation.daily_menu_info:
            try:
                date_part = guest_dessert_reservation.daily_menu_info.split('تاریخ:')[1].strip()
                date = date_part
                jalali_date = date_part
            except:
                pass
        
        guest_dessert_reservations_data.append({
            'id': guest_dessert_reservation.id,
            'type': 'guest_dessert',
            'guest_first_name': guest_dessert_reservation.guest_first_name,
            'guest_last_name': guest_dessert_reservation.guest_last_name,
            'guest_name': f"{guest_dessert_reservation.guest_first_name} {guest_dessert_reservation.guest_last_name}".strip(),
            'restaurant': {
                'id': restaurant_id,
                'name': restaurant_name,
                'centers': centers,
                'is_deleted': restaurant is None and guest_dessert_reservation.daily_menu_info is not None
            } if restaurant_name else None,
            'daily_menu_info': guest_dessert_reservation.daily_menu_info if guest_dessert_reservation.daily_menu_info else None,
            'dessert_option': dessert_option_data,
            'dessert_option_info': guest_dessert_reservation.dessert_option_info if guest_dessert_reservation.dessert_option_info else None,
            'amount': float(guest_dessert_reservation.amount or 0),
            'status': guest_dessert_reservation.status,
            'reservation_date': guest_dessert_reservation.reservation_date.isoformat() if guest_dessert_reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=guest_dessert_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if guest_dessert_reservation.reservation_date else None,
            'date': date,
            'jalali_date': jalali_date,
        })
    
    # ساخت پاسخ
    from apps.reports.serializers import UserReservationsReportSerializer
    
    report_data = {
        'user': {
            'id': target_user.id,
            'username': target_user.username,
            'full_name': target_user.get_full_name(),
            'employee_number': target_user.employee_number or '',
            'centers': [{'id': c.id, 'name': c.name} for c in target_user.centers.all()]
        },
        'start_date': parsed_start_date,
        'end_date': parsed_end_date,
        'jalali_start_date': jdatetime.date.fromgregorian(date=parsed_start_date).strftime('%Y/%m/%d'),
        'jalali_end_date': jdatetime.date.fromgregorian(date=parsed_end_date).strftime('%Y/%m/%d'),
        'total_reservations': total_reservations,
        'total_guest_reservations': total_guest_reservations,
        'total_dessert_reservations': total_dessert_reservations,
        'total_guest_dessert_reservations': total_guest_dessert_reservations,
        'reserved_count': reserved_count,
        'cancelled_count': cancelled_count,
        'served_count': served_count,
        'total_amount': total_amount,
        'reserved_amount': reserved_amount,
        'reservations': reservations_data,
        'guest_reservations': guest_reservations_data,
        'dessert_reservations': dessert_reservations_data,
        'guest_dessert_reservations': guest_dessert_reservations_data
    }
    
    serializer = UserReservationsReportSerializer(report_data, context={'request': request})
    return Response(serializer.data)


# @extend_schema(
#     operation_id='users_by_base_meal',
#     summary='Users Who Reserved a Base Meal',
#     description='Get list of users who reserved a specific base meal with their selected meal option. Only non-cancelled reservations are included.',
#     tags=['Reports'],
#     parameters=[
#         {
#             'name': 'base_meal_id',
#             'in': 'query',
#             'description': 'شناسه غذای پایه',
#             'required': True,
#             'schema': {'type': 'integer'}
#         },
#         {
#             'name': 'center_id',
#             'in': 'query',
#             'description': 'فیلتر بر اساس مرکز',
#             'required': False,
#             'schema': {'type': 'integer'}
#         },
#         {
#             'name': 'start_date',
#             'in': 'query',
#             'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
#             'required': False,
#             'schema': {'type': 'string'}
#         },
#         {
#             'name': 'end_date',
#             'in': 'query',
#             'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
#             'required': False,
#             'schema': {'type': 'string'}
#         },
#     ],
#     responses={
#         200: UserWithMealOptionSerializer(many=True),
#         400: {'description': 'Validation error'},
#         403: {'description': 'Permission denied'},
#         404: {'description': 'Base meal not found'}
#     }
# )
# @api_view(['GET'])
# @permission_classes([StatisticsPermission])
# def users_by_base_meal(request):
#     """لیست کاربرانی که یک غذای پایه را رزرو کرده‌اند به همراه اپشن انتخابی"""
#     user = request.user
    
#     # دریافت base_meal_id
#     base_meal_id = request.query_params.get('base_meal_id')
#     if not base_meal_id:
#         return Response({
#             'error': 'شناسه غذای پایه الزامی است'
#         }, status=status.HTTP_400_BAD_REQUEST)
    
#     try:
#         base_meal_id = int(base_meal_id)
#     except (ValueError, TypeError):
#         return Response({
#             'error': 'شناسه غذای پایه باید عدد باشد'
#         }, status=status.HTTP_400_BAD_REQUEST)
    
#     # بررسی وجود base_meal
#     try:
#         base_meal = BaseMeal.objects.get(id=base_meal_id)
#     except BaseMeal.DoesNotExist:
#         return Response({
#             'error': 'غذای پایه یافت نشد'
#         }, status=status.HTTP_404_NOT_FOUND)
    
#     # تعیین مراکز قابل دسترسی
#     accessible_centers = get_accessible_centers(user)
#     if accessible_centers is not None and not accessible_centers.exists():
#         return Response({
#             'error': 'کاربر مرکز مشخصی ندارد'
#         }, status=status.HTTP_403_FORBIDDEN)
    
#     # دریافت فیلترها
#     center_id = request.query_params.get('center_id')
#     start_date = parse_date_filter(request.query_params.get('start_date'))
#     end_date = parse_date_filter(request.query_params.get('end_date'))
    
#     # فیلتر رزروها - فقط رزروهای غیر کنسل شده
#     reservations = FoodReservation.objects.select_related(
#         'user', 'meal_option', 'meal_option__daily_menu',
#         'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
#     ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').filter(
#         meal_option__base_meal_id=base_meal_id
#     ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
#     guest_reservations = GuestReservation.objects.select_related(
#         'host_user', 'meal_option', 'meal_option__daily_menu',
#         'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
#     ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').filter(
#         meal_option__base_meal_id=base_meal_id
#     ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
#     # فیلتر بر اساس مرکز
#     if center_id:
#         # بررسی دسترسی به مرکز درخواستی
#         if accessible_centers is not None:
#             if not accessible_centers.filter(id=center_id).exists():
#                 return Response({
#                     'error': 'شما دسترسی به این مرکز ندارید'
#                 }, status=status.HTTP_403_FORBIDDEN)
        
#         # شامل رزروهایی که daily_menu=None است یا daily_menu.restaurant.centers شامل center_id است
#         reservations = reservations.filter(
#             Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
#         ).distinct()
#         guest_reservations = guest_reservations.filter(
#             Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
#         ).distinct()
#     elif accessible_centers is not None:
#         # فیلتر بر اساس مراکز قابل دسترسی
#         # شامل رزروهایی که daily_menu=None است یا daily_menu.restaurant.centers در accessible_centers است
#         reservations = reservations.filter(
#             Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__in=accessible_centers)
#         ).distinct()
#         guest_reservations = guest_reservations.filter(
#             Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__in=accessible_centers)
#         ).distinct()
    
#     # فیلتر بر اساس تاریخ
#     if start_date:
#         reservations = reservations.filter(daily_menu__date__gte=start_date)
#         guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
#     if end_date:
#         reservations = reservations.filter(daily_menu__date__lte=end_date)
#         guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
#     # ساخت لیست کاربران با اپشن انتخابی
#     users_data = []
    
#     # پردازش رزروهای معمولی
#     for reservation in reservations:
#         # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
#         if reservation.status == 'cancelled':
#             continue
        
#         # بررسی meal_option
#         if not reservation.meal_option:
#             continue
        
#         user_obj = reservation.user
#         meal_option = reservation.meal_option
#         restaurant = reservation.daily_menu.restaurant if reservation.daily_menu else None
        
#         # اطلاعات اپشن غذا
#         meal_option_data = {
#             'id': meal_option.id,
#             'title': meal_option.title,
#             'description': meal_option.description or '',
#             'price': float(meal_option.price),
#             'base_meal_id': base_meal.id,
#             'base_meal_title': base_meal.title,
#         }
        
#         # اطلاعات رستوران
#         restaurant_data = None
#         if restaurant:
#             restaurant_data = {
#                 'id': restaurant.id,
#                 'name': restaurant.name,
#                 'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
#             }
        
#         users_data.append({
#             'user_id': user_obj.id,
#             'username': user_obj.username,
#             'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
#             'employee_number': getattr(user_obj, 'employee_number', '') or '',
#             'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()],
#             'meal_option': meal_option_data,
#             'quantity': reservation.quantity,
#             'amount': float(reservation.amount or 0),
#             'status': reservation.status,
#             'reservation_date': reservation.reservation_date.isoformat() if reservation.reservation_date else None,
#             'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if reservation.reservation_date else None,
#             'daily_menu_date': reservation.daily_menu.date.isoformat() if reservation.daily_menu and reservation.daily_menu.date else None,
#             'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=reservation.daily_menu.date).strftime('%Y/%m/%d') if reservation.daily_menu and reservation.daily_menu.date else None,
#             'restaurant': restaurant_data,
#             'is_guest': False,
#             'guest_name': None,
#         })
    
#     # پردازش رزروهای مهمان
#     for guest_reservation in guest_reservations:
#         # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
#         if guest_reservation.status == 'cancelled':
#             continue
        
#         # بررسی meal_option
#         if not guest_reservation.meal_option:
#             continue
        
#         user_obj = guest_reservation.host_user
#         meal_option = guest_reservation.meal_option
#         restaurant = guest_reservation.daily_menu.restaurant if guest_reservation.daily_menu else None
        
#         # اطلاعات اپشن غذا
#         meal_option_data = {
#             'id': meal_option.id,
#             'title': meal_option.title,
#             'description': meal_option.description or '',
#             'price': float(meal_option.price),
#             'base_meal_id': base_meal.id,
#             'base_meal_title': base_meal.title,
#         }
        
#         # اطلاعات رستوران
#         restaurant_data = None
#         if restaurant:
#             restaurant_data = {
#                 'id': restaurant.id,
#                 'name': restaurant.name,
#                 'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
#             }
        
#         users_data.append({
#             'user_id': user_obj.id,
#             'username': user_obj.username,
#             'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
#             'employee_number': getattr(user_obj, 'employee_number', '') or '',
#             'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()],
#             'meal_option': meal_option_data,
#             'quantity': 1,  # مهمان همیشه 1 است
#             'amount': float(guest_reservation.amount or 0),
#             'status': guest_reservation.status,
#             'reservation_date': guest_reservation.reservation_date.isoformat() if guest_reservation.reservation_date else None,
#             'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=guest_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if guest_reservation.reservation_date else None,
#             'daily_menu_date': guest_reservation.daily_menu.date.isoformat() if guest_reservation.daily_menu and guest_reservation.daily_menu.date else None,
#             'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=guest_reservation.daily_menu.date).strftime('%Y/%m/%d') if guest_reservation.daily_menu and guest_reservation.daily_menu.date else None,
#             'restaurant': restaurant_data,
#             'is_guest': True,
#             'guest_name': f"{guest_reservation.guest_first_name} {guest_reservation.guest_last_name}".strip(),
#         })
    
#     serializer = UserWithMealOptionSerializer(users_data, many=True)
#     return Response(serializer.data)


# @extend_schema(
#     operation_id='users_by_base_dessert',
#     summary='Users Who Reserved a Base Dessert',
#     description='Get list of users who reserved a specific base dessert with their selected dessert option. Only non-cancelled reservations are included.',
#     tags=['Reports'],
#     parameters=[
#         {
#             'name': 'base_dessert_id',
#             'in': 'query',
#             'description': 'شناسه دسر پایه',
#             'required': True,
#             'schema': {'type': 'integer'}
#         },
#         {
#             'name': 'center_id',
#             'in': 'query',
#             'description': 'فیلتر بر اساس مرکز',
#             'required': False,
#             'schema': {'type': 'integer'}
#         },
#         {
#             'name': 'start_date',
#             'in': 'query',
#             'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
#             'required': False,
#             'schema': {'type': 'string'}
#         },
#         {
#             'name': 'end_date',
#             'in': 'query',
#             'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
#             'required': False,
#             'schema': {'type': 'string'}
#         },
#     ],
#     responses={
#         200: UserWithDessertOptionSerializer(many=True),
#         400: {'description': 'Validation error'},
#         403: {'description': 'Permission denied'},
#         404: {'description': 'Base dessert not found'}
#     }
# )
# @api_view(['GET'])
# @permission_classes([StatisticsPermission])
# def users_by_base_dessert(request):
#     """لیست کاربرانی که یک دسر پایه را رزرو کرده‌اند به همراه اپشن انتخابی"""
#     user = request.user
    
#     # دریافت base_dessert_id
#     base_dessert_id = request.query_params.get('base_dessert_id')
#     if not base_dessert_id:
#         return Response({
#             'error': 'شناسه دسر پایه الزامی است'
#         }, status=status.HTTP_400_BAD_REQUEST)
    
#     try:
#         base_dessert_id = int(base_dessert_id)
#     except (ValueError, TypeError):
#         return Response({
#             'error': 'شناسه دسر پایه باید عدد باشد'
#         }, status=status.HTTP_400_BAD_REQUEST)
    
#     # بررسی وجود base_dessert
#     try:
#         base_dessert = BaseDessert.objects.get(id=base_dessert_id)
#     except BaseDessert.DoesNotExist:
#         return Response({
#             'error': 'دسر پایه یافت نشد'
#         }, status=status.HTTP_404_NOT_FOUND)
    
#     # تعیین مراکز قابل دسترسی
#     accessible_centers = get_accessible_centers(user)
#     if accessible_centers is not None and not accessible_centers.exists():
#         return Response({
#             'error': 'کاربر مرکز مشخصی ندارد'
#         }, status=status.HTTP_403_FORBIDDEN)
    
#     # دریافت فیلترها
#     center_id = request.query_params.get('center_id')
#     start_date = parse_date_filter(request.query_params.get('start_date'))
#     end_date = parse_date_filter(request.query_params.get('end_date'))
    
#     # فیلتر رزروهای دسر - فقط رزروهای غیر کنسل شده
#     dessert_reservations = DessertReservation.objects.select_related(
#         'user', 'dessert_option', 'dessert_option__daily_menu',
#         'dessert_option__daily_menu__restaurant', 'daily_menu__restaurant'
#     ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').filter(
#         dessert_option__base_dessert_id=base_dessert_id
#     ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
#     guest_dessert_reservations = GuestDessertReservation.objects.select_related(
#         'host_user', 'dessert_option', 'dessert_option__daily_menu',
#         'dessert_option__daily_menu__restaurant', 'daily_menu__restaurant'
#     ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').filter(
#         dessert_option__base_dessert_id=base_dessert_id
#     ).exclude(status='cancelled')  # حذف رزروهای کنسل شده
    
#     # فیلتر بر اساس مرکز
#     if center_id:
#         # بررسی دسترسی به مرکز درخواستی
#         if accessible_centers is not None:
#             if not accessible_centers.filter(id=center_id).exists():
#                 return Response({
#                     'error': 'شما دسترسی به این مرکز ندارید'
#                 }, status=status.HTTP_403_FORBIDDEN)
        
#         dessert_reservations = dessert_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
#         guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
#     elif accessible_centers is not None:
#         # فیلتر بر اساس مراکز قابل دسترسی
#         dessert_reservations = dessert_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
#         guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
    
#     # فیلتر بر اساس تاریخ
#     if start_date:
#         dessert_reservations = dessert_reservations.filter(daily_menu__date__gte=start_date)
#         guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__date__gte=start_date)
    
#     if end_date:
#         dessert_reservations = dessert_reservations.filter(daily_menu__date__lte=end_date)
#         guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__date__lte=end_date)
    
#     # ساخت لیست کاربران با اپشن انتخابی
#     users_data = []
    
#     # پردازش رزروهای دسر
#     for dessert_reservation in dessert_reservations:
#         # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
#         if dessert_reservation.status == 'cancelled':
#             continue
        
#         if not dessert_reservation.dessert_option:
#             continue
        
#         user_obj = dessert_reservation.user
#         dessert_option = dessert_reservation.dessert_option
#         restaurant = dessert_reservation.daily_menu.restaurant if dessert_reservation.daily_menu else None
        
#         # اطلاعات اپشن دسر
#         dessert_option_data = {
#             'id': dessert_option.id,
#             'title': dessert_option.title,
#             'description': dessert_option.description or '',
#             'price': float(dessert_option.price),
#             'base_dessert_id': base_dessert.id,
#             'base_dessert_title': base_dessert.title,
#         }
        
#         # اطلاعات رستوران
#         restaurant_data = None
#         if restaurant:
#             restaurant_data = {
#                 'id': restaurant.id,
#                 'name': restaurant.name,
#                 'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
#             }
        
#         users_data.append({
#             'user_id': user_obj.id,
#             'username': user_obj.username,
#             'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
#             'employee_number': getattr(user_obj, 'employee_number', '') or '',
#             'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()],
#             'dessert_option': dessert_option_data,
#             'quantity': dessert_reservation.quantity,
#             'amount': float(dessert_reservation.amount or 0),
#             'status': dessert_reservation.status,
#             'reservation_date': dessert_reservation.reservation_date.isoformat() if dessert_reservation.reservation_date else None,
#             'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=dessert_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if dessert_reservation.reservation_date else None,
#             'daily_menu_date': dessert_reservation.daily_menu.date.isoformat() if dessert_reservation.daily_menu and dessert_reservation.daily_menu.date else None,
#             'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=dessert_reservation.daily_menu.date).strftime('%Y/%m/%d') if dessert_reservation.daily_menu and dessert_reservation.daily_menu.date else None,
#             'restaurant': restaurant_data,
#             'is_guest': False,
#             'guest_name': None,
#         })
    
#     # پردازش رزروهای مهمان دسر
#     for guest_dessert_reservation in guest_dessert_reservations:
#         # چک اضافی برای اطمینان از حذف رزروهای کنسل شده
#         if guest_dessert_reservation.status == 'cancelled':
#             continue
        
#         if not guest_dessert_reservation.dessert_option:
#             continue
        
#         user_obj = guest_dessert_reservation.host_user
#         dessert_option = guest_dessert_reservation.dessert_option
#         restaurant = guest_dessert_reservation.daily_menu.restaurant if guest_dessert_reservation.daily_menu else None
        
#         # اطلاعات اپشن دسر
#         dessert_option_data = {
#             'id': dessert_option.id,
#             'title': dessert_option.title,
#             'description': dessert_option.description or '',
#             'price': float(dessert_option.price),
#             'base_dessert_id': base_dessert.id,
#             'base_dessert_title': base_dessert.title,
#         }
        
#         # اطلاعات رستوران
#         restaurant_data = None
#         if restaurant:
#             restaurant_data = {
#                 'id': restaurant.id,
#                 'name': restaurant.name,
#                 'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
#             }
        
#         users_data.append({
#             'user_id': user_obj.id,
#             'username': user_obj.username,
#             'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
#             'employee_number': getattr(user_obj, 'employee_number', '') or '',
#             'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()],
#             'dessert_option': dessert_option_data,
#             'quantity': 1,  # مهمان همیشه 1 است
#             'amount': float(guest_dessert_reservation.amount or 0),
#             'status': guest_dessert_reservation.status,
#             'reservation_date': guest_dessert_reservation.reservation_date.isoformat() if guest_dessert_reservation.reservation_date else None,
#             'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=guest_dessert_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if guest_dessert_reservation.reservation_date else None,
#             'daily_menu_date': guest_dessert_reservation.daily_menu.date.isoformat() if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.date else None,
#             'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=guest_dessert_reservation.daily_menu.date).strftime('%Y/%m/%d') if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.date else None,
#             'restaurant': restaurant_data,
#             'is_guest': True,
#             'guest_name': f"{guest_dessert_reservation.guest_first_name} {guest_dessert_reservation.guest_last_name}".strip(),
#         })
    
#     serializer = UserWithDessertOptionSerializer(users_data, many=True)
#     return Response(serializer.data)

@extend_schema(
    operation_id='users_by_base_meal',
    summary='Users Who Reserved Base Meals',
    description='Get list of base meals reserved in date range, grouped with users who reserved each meal. Each base meal shows all users and their selected meal options.',
    tags=['Reports'],
    parameters=[
        {
            'name': 'base_meal_id',
            'in': 'query',
            'description': 'شناسه غذای پایه (اختیاری - اگر داده نشود، همه غذاها نمایش داده می‌شوند)',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'},
        404: {'description': 'Base meal not found'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def users_by_base_meal(request):
    """لیست غذاهای پایه با کاربرانی که هر غذا را رزرو کرده‌اند"""
    user = request.user
    
    base_meal_id = request.query_params.get('base_meal_id')
    
    if base_meal_id:
        try:
            base_meal_id = int(base_meal_id)
            BaseMeal.objects.get(id=base_meal_id)
        except (ValueError, TypeError):
            return Response({
                'error': 'شناسه غذای پایه باید عدد باشد'
            }, status=status.HTTP_400_BAD_REQUEST)
        except BaseMeal.DoesNotExist:
            return Response({
                'error': 'غذای پایه یافت نشد'
            }, status=status.HTTP_404_NOT_FOUND)
    
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    reservations = FoodReservation.objects.select_related(
        'user', 'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').exclude(status='cancelled')
    
    if base_meal_id:
        reservations = reservations.filter(meal_option__base_meal_id=base_meal_id)
    
    guest_reservations = GuestReservation.objects.select_related(
        'host_user', 'meal_option', 'meal_option__base_meal', 'meal_option__daily_menu',
        'meal_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').exclude(status='cancelled')
    
    if base_meal_id:
        guest_reservations = guest_reservations.filter(meal_option__base_meal_id=base_meal_id)
    
    if center_id:
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        reservations = reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
        guest_reservations = guest_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__id=center_id)
        ).distinct()
    elif accessible_centers is not None:
        reservations = reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__in=accessible_centers)
        ).distinct()
        guest_reservations = guest_reservations.filter(
            Q(daily_menu__isnull=True) | Q(daily_menu__restaurant__centers__in=accessible_centers)
        ).distinct()
    
    if start_date:
        reservations = reservations.filter(daily_menu__date__gte=start_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        reservations = reservations.filter(daily_menu__date__lte=end_date)
        guest_reservations = guest_reservations.filter(daily_menu__date__lte=end_date)
    
    meals_data = {}
    
    for reservation in reservations:
        if not reservation.meal_option or not reservation.meal_option.base_meal:
            continue
        
        base_meal = reservation.meal_option.base_meal
        restaurant = reservation.daily_menu.restaurant if reservation.daily_menu else None
        
        meal_key = (base_meal.id, restaurant.id if restaurant else None)
        
        if meal_key not in meals_data:
            meals_data[meal_key] = {
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or ''
                },
                'restaurant': {
                    'id': restaurant.id,
                    'name': restaurant.name,
                    'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
                } if restaurant else None,
                'users': {}
            }
        
        user_obj = reservation.user
        
        if user_obj.id not in meals_data[meal_key]['users']:
            meals_data[meal_key]['users'][user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'reservations': []
            }
        
        meal_option = reservation.meal_option
        meals_data[meal_key]['users'][user_obj.id]['reservations'].append({
            'meal_option': {
                'id': meal_option.id,
                'title': meal_option.title,
                'description': meal_option.description or '',
                'price': float(meal_option.price)
            },
            'quantity': reservation.quantity,
            'amount': float(reservation.amount or 0),
            'status': reservation.status,
            'reservation_date': reservation.reservation_date.isoformat() if reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if reservation.reservation_date else None,
            'daily_menu_date': reservation.daily_menu.date.isoformat() if reservation.daily_menu and reservation.daily_menu.date else None,
            'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=reservation.daily_menu.date).strftime('%Y/%m/%d') if reservation.daily_menu and reservation.daily_menu.date else None,
            'is_guest': False
        })
    
    for guest_reservation in guest_reservations:
        if not guest_reservation.meal_option or not guest_reservation.meal_option.base_meal:
            continue
        
        base_meal = guest_reservation.meal_option.base_meal
        restaurant = guest_reservation.daily_menu.restaurant if guest_reservation.daily_menu else None
        
        meal_key = (base_meal.id, restaurant.id if restaurant else None)
        
        if meal_key not in meals_data:
            meals_data[meal_key] = {
                'base_meal': {
                    'id': base_meal.id,
                    'title': base_meal.title,
                    'description': base_meal.description or ''
                },
                'restaurant': {
                    'id': restaurant.id,
                    'name': restaurant.name,
                    'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
                } if restaurant else None,
                'users': {}
            }
        
        user_obj = guest_reservation.host_user
        
        if user_obj.id not in meals_data[meal_key]['users']:
            meals_data[meal_key]['users'][user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'reservations': []
            }
        
        meal_option = guest_reservation.meal_option
        meals_data[meal_key]['users'][user_obj.id]['reservations'].append({
            'meal_option': {
                'id': meal_option.id,
                'title': meal_option.title,
                'description': meal_option.description or '',
                'price': float(meal_option.price)
            },
            'quantity': 1,
            'amount': float(guest_reservation.amount or 0),
            'status': guest_reservation.status,
            'reservation_date': guest_reservation.reservation_date.isoformat() if guest_reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=guest_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if guest_reservation.reservation_date else None,
            'daily_menu_date': guest_reservation.daily_menu.date.isoformat() if guest_reservation.daily_menu and guest_reservation.daily_menu.date else None,
            'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=guest_reservation.daily_menu.date).strftime('%Y/%m/%d') if guest_reservation.daily_menu and guest_reservation.daily_menu.date else None,
            'is_guest': True,
            'guest_name': f"{guest_reservation.guest_first_name} {guest_reservation.guest_last_name}".strip()
        })
    
    result = []
    for meal_data in meals_data.values():
        meal_data['users'] = list(meal_data['users'].values())
        result.append(meal_data)
    
    return Response(result)

@extend_schema(
    operation_id='users_by_base_dessert',
    summary='Users Who Reserved Base Desserts',
    description='Get list of base desserts reserved in date range, grouped with users who reserved each dessert. Each base dessert shows all users and their selected dessert options.',
    tags=['Reports'],
    parameters=[
        {
            'name': 'base_dessert_id',
            'in': 'query',
            'description': 'شناسه دسر پایه (اختیاری - اگر داده نشود، همه دسرها نمایش داده می‌شوند)',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'center_id',
            'in': 'query',
            'description': 'فیلتر بر اساس مرکز',
            'required': False,
            'schema': {'type': 'integer'}
        },
        {
            'name': 'start_date',
            'in': 'query',
            'description': 'تاریخ شروع (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
        {
            'name': 'end_date',
            'in': 'query',
            'description': 'تاریخ پایان (فرمت: YYYY-MM-DD یا YYYY/MM/DD)',
            'required': False,
            'schema': {'type': 'string'}
        },
    ],
    responses={
        200: OpenApiTypes.OBJECT,
        400: {'description': 'Validation error'},
        403: {'description': 'Permission denied'},
        404: {'description': 'Base dessert not found'}
    }
)
@api_view(['GET'])
@permission_classes([StatisticsPermission])
def users_by_base_dessert(request):
    """لیست دسرهای پایه با کاربرانی که هر دسر را رزرو کرده‌اند"""
    user = request.user
    
    base_dessert_id = request.query_params.get('base_dessert_id')
    
    if base_dessert_id:
        try:
            base_dessert_id = int(base_dessert_id)
            BaseDessert.objects.get(id=base_dessert_id)
        except (ValueError, TypeError):
            return Response({
                'error': 'شناسه دسر پایه باید عدد باشد'
            }, status=status.HTTP_400_BAD_REQUEST)
        except BaseDessert.DoesNotExist:
            return Response({
                'error': 'دسر پایه یافت نشد'
            }, status=status.HTTP_404_NOT_FOUND)
    
    accessible_centers = get_accessible_centers(user)
    if accessible_centers is not None and not accessible_centers.exists():
        return Response({
            'error': 'کاربر مرکز مشخصی ندارد'
        }, status=status.HTTP_403_FORBIDDEN)
    
    center_id = request.query_params.get('center_id')
    start_date = parse_date_filter(request.query_params.get('start_date'))
    end_date = parse_date_filter(request.query_params.get('end_date'))
    
    dessert_reservations = DessertReservation.objects.select_related(
        'user', 'dessert_option', 'dessert_option__base_dessert', 'dessert_option__daily_menu',
        'dessert_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'user__centers').exclude(status='cancelled')
    
    if base_dessert_id:
        dessert_reservations = dessert_reservations.filter(dessert_option__base_dessert_id=base_dessert_id)
    
    guest_dessert_reservations = GuestDessertReservation.objects.select_related(
        'host_user', 'dessert_option', 'dessert_option__base_dessert', 'dessert_option__daily_menu',
        'dessert_option__daily_menu__restaurant', 'daily_menu__restaurant'
    ).prefetch_related('daily_menu__restaurant__centers', 'host_user__centers').exclude(status='cancelled')
    
    if base_dessert_id:
        guest_dessert_reservations = guest_dessert_reservations.filter(dessert_option__base_dessert_id=base_dessert_id)
    
    if center_id:
        if accessible_centers is not None:
            if not accessible_centers.filter(id=center_id).exists():
                return Response({
                    'error': 'شما دسترسی به این مرکز ندارید'
                }, status=status.HTTP_403_FORBIDDEN)
        
        dessert_reservations = dessert_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
        guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__restaurant__centers__id=center_id).distinct()
    elif accessible_centers is not None:
        dessert_reservations = dessert_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
        guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__restaurant__centers__in=accessible_centers).distinct()
    
    if start_date:
        dessert_reservations = dessert_reservations.filter(daily_menu__date__gte=start_date)
        guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__date__gte=start_date)
    
    if end_date:
        dessert_reservations = dessert_reservations.filter(daily_menu__date__lte=end_date)
        guest_dessert_reservations = guest_dessert_reservations.filter(daily_menu__date__lte=end_date)
    
    desserts_data = {}
    
    for dessert_reservation in dessert_reservations:
        if not dessert_reservation.dessert_option or not dessert_reservation.dessert_option.base_dessert:
            continue
        
        base_dessert = dessert_reservation.dessert_option.base_dessert
        restaurant = dessert_reservation.daily_menu.restaurant if dessert_reservation.daily_menu else None
        
        dessert_key = (base_dessert.id, restaurant.id if restaurant else None)
        
        if dessert_key not in desserts_data:
            desserts_data[dessert_key] = {
                'base_dessert': {
                    'id': base_dessert.id,
                    'title': base_dessert.title,
                    'description': base_dessert.description or ''
                },
                'restaurant': {
                    'id': restaurant.id,
                    'name': restaurant.name,
                    'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
                } if restaurant else None,
                'users': {}
            }
        
        user_obj = dessert_reservation.user
        
        if user_obj.id not in desserts_data[dessert_key]['users']:
            desserts_data[dessert_key]['users'][user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'reservations': []
            }
        
        dessert_option = dessert_reservation.dessert_option
        desserts_data[dessert_key]['users'][user_obj.id]['reservations'].append({
            'dessert_option': {
                'id': dessert_option.id,
                'title': dessert_option.title,
                'description': dessert_option.description or '',
                'price': float(dessert_option.price)
            },
            'quantity': dessert_reservation.quantity,
            'amount': float(dessert_reservation.amount or 0),
            'status': dessert_reservation.status,
            'reservation_date': dessert_reservation.reservation_date.isoformat() if dessert_reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=dessert_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if dessert_reservation.reservation_date else None,
            'daily_menu_date': dessert_reservation.daily_menu.date.isoformat() if dessert_reservation.daily_menu and dessert_reservation.daily_menu.date else None,
            'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=dessert_reservation.daily_menu.date).strftime('%Y/%m/%d') if dessert_reservation.daily_menu and dessert_reservation.daily_menu.date else None,
            'is_guest': False
        })
    
    for guest_dessert_reservation in guest_dessert_reservations:
        if not guest_dessert_reservation.dessert_option or not guest_dessert_reservation.dessert_option.base_dessert:
            continue
        
        base_dessert = guest_dessert_reservation.dessert_option.base_dessert
        restaurant = guest_dessert_reservation.daily_menu.restaurant if guest_dessert_reservation.daily_menu else None
        
        dessert_key = (base_dessert.id, restaurant.id if restaurant else None)
        
        if dessert_key not in desserts_data:
            desserts_data[dessert_key] = {
                'base_dessert': {
                    'id': base_dessert.id,
                    'title': base_dessert.title,
                    'description': base_dessert.description or ''
                },
                'restaurant': {
                    'id': restaurant.id,
                    'name': restaurant.name,
                    'centers': [{'id': c.id, 'name': c.name} for c in restaurant.centers.all()]
                } if restaurant else None,
                'users': {}
            }
        
        user_obj = guest_dessert_reservation.host_user
        
        if user_obj.id not in desserts_data[dessert_key]['users']:
            desserts_data[dessert_key]['users'][user_obj.id] = {
                'user': {
                    'id': user_obj.id,
                    'username': user_obj.username,
                    'full_name': f"{user_obj.first_name} {user_obj.last_name}".strip(),
                    'employee_number': getattr(user_obj, 'employee_number', '') or '',
                    'centers': [{'id': c.id, 'name': c.name} for c in user_obj.centers.all()]
                },
                'reservations': []
            }
        
        dessert_option = guest_dessert_reservation.dessert_option
        desserts_data[dessert_key]['users'][user_obj.id]['reservations'].append({
            'dessert_option': {
                'id': dessert_option.id,
                'title': dessert_option.title,
                'description': dessert_option.description or '',
                'price': float(dessert_option.price)
            },
            'quantity': 1,
            'amount': float(guest_dessert_reservation.amount or 0),
            'status': guest_dessert_reservation.status,
            'reservation_date': guest_dessert_reservation.reservation_date.isoformat() if guest_dessert_reservation.reservation_date else None,
            'jalali_reservation_date': jdatetime.datetime.fromgregorian(datetime=guest_dessert_reservation.reservation_date).strftime('%Y/%m/%d %H:%M') if guest_dessert_reservation.reservation_date else None,
            'daily_menu_date': guest_dessert_reservation.daily_menu.date.isoformat() if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.date else None,
            'jalali_daily_menu_date': jdatetime.date.fromgregorian(date=guest_dessert_reservation.daily_menu.date).strftime('%Y/%m/%d') if guest_dessert_reservation.daily_menu and guest_dessert_reservation.daily_menu.date else None,
            'is_guest': True,
            'guest_name': f"{guest_dessert_reservation.guest_first_name} {guest_dessert_reservation.guest_last_name}".strip()
        })
    
    result = []
    for dessert_data in desserts_data.values():
        dessert_data['users'] = list(dessert_data['users'].values())
        result.append(dessert_data)
    
    return Response(result)